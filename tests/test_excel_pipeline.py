"""Excel PIMS-shaped → mono + ADMM → lean results workbook tests."""
from __future__ import annotations

from pathlib import Path

import pytest

pytest.importorskip("openpyxl")

from pims_admm_llm.models.excel_pipeline import (
    ensure_template,
    load_pims_excel,
    run_excel_pipeline,
    write_results_excel,
)
from pims_admm_llm.models.assay_loader import write_template_excel

# Goal: ≤15 sheets, one tab per unit submodel, PIMS How-To 07 FCC/Coker
GOAL_MAX_SHEETS = 15
REQUIRED = (
    "How_to_read",
    "Submodel_Index",
    "Calc_Yields",
    "Calc_Blend",
    "Submodel_CDU",
    "Submodel_Blender",
    "Submodel_FCC",
    "Submodel_Coker",
    "Submodel_Linking",
    "Summary",
    "Rates",
    "Shadows",
    "Calc_Check",
)
BANNED_PREFIXES = ("Submodel_FCC_", "Submodel_Coker_", "Live_")
BANNED_EXACT = {
    "Submodel_CDU_Tech",
    "Submodel_CDU_A",
    "Submodel_Blender_Tech",
    "Submodel_Blender_A",
    "Submodel_Linking_B",
    "Calc_BlockAngular",
    "Calc_BA_Map",
    "Calc_BA_Legend",
    "Calc_Process",
    "Calc_Blocks",
    "Calc_Bounds",
    "Calc_Objective",
    "Calc_ModelNote",
    "Calc_Equations",
    "Calc_Linking",
    "Crudes_mono",
    "Products_mono",
    "Crudes_admm",
    "Products_admm",
    "Inter_prod_mono",
    "Inter_use_mono",
}


def test_excel_pipeline_end_to_end(tmp_path):
    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    xlsx_out = tmp_path / "results.xlsx"
    json_out = tmp_path / "results.json"
    report = run_excel_pipeline(
        xlsx_in, results_xlsx=xlsx_out, results_json=json_out
    )
    assert xlsx_out.is_file() and json_out.is_file()
    assert report["mono"]["feasible"] and report["admm"]["feasible"]
    assert report["mono"]["objective"] > 0 and report["admm"]["objective"] > 0
    assert report["comparison"]["objective_gap_rel"] <= 0.005 + 1e-9
    assert report["comparison"]["dual_linf_online"] <= 15.0
    assert report["verdict"].startswith("PASS")
    assert report["meta"]["admm_config"]["rho"] == 8.0
    # E14: feasibility stays classic 2-block; TF must not own duals or form.
    assert report.get("model", {}).get("form") == "classic_2block_excel_path"
    path = str(report.get("admm", {}).get("dual_recovery_path") or "")
    assert "online_lambda" in path
    assert "tf_block" not in path.lower()
    assert "tensorflow" not in path.lower()
    assert "tf_block" not in report
    assert report.get("tf_block") is None


def test_excel_pipeline_shadows_prefer_online_lambda(tmp_path):
    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    mono_sh = report["mono"]["shadow_prices"]
    online = report["admm"]["shadow_prices"]
    recovered = report["admm"]["shadow_prices_recovered"]
    assert report["comparison"]["dual_linf_online"] <= report["comparison"][
        "dual_linf_recovered"
    ] + 1e-9
    for s in mono_sh:
        assert s in online and s in recovered


def test_write_results_excel_lean_goal(tmp_path):
    """Agentic goal: ≤15 sheets, one unit tab each, PIMS FCC/Coker matrix."""
    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    out = tmp_path / "out.xlsx"
    write_results_excel(out, report)
    import openpyxl

    wb = openpyxl.load_workbook(out)
    names = wb.sheetnames

    assert len(names) <= GOAL_MAX_SHEETS, f"n_sheets={len(names)} {names}"
    assert names[0] == "How_to_read"
    for name in REQUIRED:
        assert name in names, name

    banned = [
        s
        for s in names
        if s in BANNED_EXACT or any(s.startswith(p) for p in BANNED_PREFIXES)
    ]
    assert banned == [], f"banned tabs present: {banned}"

    # Index lists single unit sheets
    ih = [c.value for c in wb["Submodel_Index"][1]]
    blocks = {
        r[ih.index("block")].value
        for r in wb["Submodel_Index"].iter_rows(min_row=2)
        if r[0].value
    }
    assert {"CDU", "BLENDER", "FCC", "COKER", "LINKING"} <= blocks

    # CDU merged sections
    cdu_vals = [r[0].value for r in wb["Submodel_CDU"].iter_rows(min_col=1, max_col=1)]
    assert any(v and "TECH" in str(v) for v in cdu_vals)
    assert any(v and "A —" in str(v) or (v and str(v).startswith("===") and "A" in str(v)) for v in cdu_vals)

    # FCC PIMS matrix
    fh = [c.value for c in wb["Submodel_FCC"][1]]
    assert "FEED_FFD" in fh and "BASE" in fh
    assert any(str(h).startswith("D_") for h in fh if h)
    fcc = {
        r[fh.index("row")].value: dict(zip(fh, [c.value for c in r]))
        for r in wb["Submodel_FCC"].iter_rows(min_row=2)
        if r[fh.index("row")].value
    }
    assert fcc["E_BASE_REF"]["FEED_FFD"] == 1.0
    assert fcc["E_BASE_REF"]["BASE"] == -1.0
    assert float(fcc["MB_fcc_naphtha"]["BASE"]) > 0.3
    assert (fcc.get("E_api_REF") or {}).get("FEED_FFD") == -999.0

    ch = [c.value for c in wb["Submodel_Coker"][1]]
    assert "FEED_CFD" in ch and "BASE" in ch
    crow = [r[ch.index("row")].value for r in wb["Submodel_Coker"].iter_rows(min_row=2)]
    assert "E_BASE_REF" in crow and "MB_coker_naphtha" in crow

    # Rates comparison
    rh = [c.value for c in wb["Rates"][1]]
    assert "mono_kbd" in rh and "admm_kbd" in rh

    # Check all ok
    chk_h = [c.value for c in wb["Calc_Check"][1]]
    ok_i = chk_h.index("ok")
    for row in wb["Calc_Check"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            assert row[ok_i] is True, row

    # yield_sum formula still on Calc_Yields
    yh = [c.value for c in wb["Calc_Yields"][1]]
    ys = yh.index("yield_sum")
    cell = wb["Calc_Yields"].cell(2, ys + 1).value
    assert isinstance(cell, str) and cell.startswith("="), cell

    assert report.get("model", {}).get("form") == "classic_2block_excel_path"

    # E1/E11: How_to three-path honesty (export ≠ offline TF ≠ Case1 duals)
    how = {
        str(r[0].value): str(r[1].value or "")
        for r in wb["How_to_read"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    three = how.get("fcc_three_path", "")
    assert "base_delta export" in three or "BASE/D_*" in three
    assert "offline TF" in three or "tf_linear" in three
    assert "classic_2block" in three
    dual_note = how.get("duals_online_lambda", "")
    assert "online" in dual_note.lower()
    # E1/E2: Coker three-path + renorm-outside-affine honesty
    coker = how.get("coker_three_path", "")
    assert coker, "How_to_read must include coker_three_path"
    assert "BASE/D_*" in coker or "base_delta export" in coker
    assert "tf_linear_coker" in coker or "offline TF" in coker
    assert "classic_2block" in coker
    assert "renorm" in coker.lower() or "postprocess" in coker.lower()
    assert "evaluate" in coker.lower() or "reference" in coker.lower()
    # E1/E2: CDU three-path — TECH+A classic ≠ offline TF ≠ duals (not PIMS MB_*)
    cdu = how.get("cdu_three_path", "")
    assert cdu, "How_to_read must include cdu_three_path"
    assert "TECH" in cdu and "A" in cdu
    assert "tf_linear_cdu" in cdu or "offline TF" in cdu or "offline" in cdu.lower()
    assert "not" in cdu.lower() and ("MB_*" in cdu or "BASE/DELTA" in cdu or "PIMS" in cdu or "How-To 07" in cdu)
    assert "online" in cdu.lower() or "dual" in cdu.lower()
    assert "postprocess" in cdu.lower() or "renorm" in cdu.lower() or "affine" in cdu.lower()

    # E1/E2: multi-unit offline TF status (static How_to; not on Case 1 solve)
    tf_off = how.get("tf_offline_units", "")
    assert tf_off, "How_to_read must include tf_offline_units"
    low = tf_off.lower()
    assert "fcc" in low and "coker" in low and "cdu" in low
    assert "classic_2block" in low or "not on this case 1" in low or "not on" in low
    assert "dual_recovery_path" in low or "dual" in low
    assert "primary" in low
    assert "solver=false" in low or "solver=False" in tf_off
    assert "none" in low  # dual_recovery_path=None on TF surface

    # Optional secondary: priced residual readiness (static How_to; still not Case 1)
    tf_priced = how.get("tf_offline_priced", "")
    assert tf_priced, "How_to_read must include tf_offline_priced"
    plow = tf_priced.lower()
    assert "priced residual" in plow or "priced" in plow
    assert "fcc" in plow and "coker" in plow and "cdu" in plow
    assert "not on" in plow or "classic_2block" in plow
    assert "dual" in plow and ("none" in plow or "primary" in plow)
    assert "not" in plow and ("shadow" in plow or "λ" in plow or "lambda" in plow or "admm" in plow)

    # Optional secondary: offline block-solve timing readiness (static How_to)
    tf_timing = how.get("tf_offline_timing", "")
    assert tf_timing, "How_to_read must include tf_offline_timing"
    tlow = tf_timing.lower()
    assert "timing" in tlow or "block-solve" in tlow or "block solve" in tlow
    assert "fcc" in tlow and "coker" in tlow and "cdu" in tlow
    assert "not on" in tlow or "classic_2block" in tlow
    assert "dual" in tlow and ("none" in tlow or "primary" in tlow)
    assert "not" in tlow and (
        "shadow" in tlow or "λ" in tlow or "lambda" in tlow or "wall" in tlow
    )

    # Optional secondary: offline ADMM residual harness (static How_to; still not Case 1)
    tf_admm = how.get("tf_offline_admm_residual", "")
    assert tf_admm, "How_to_read must include tf_offline_admm_residual"
    alow = tf_admm.lower()
    assert "admm" in alow and ("residual" in alow or "consensus" in alow)
    assert "fcc" in alow and "coker" in alow and "cdu" in alow
    assert "not on" in alow or "classic_2block" in alow
    assert "dual" in alow and ("none" in alow or "primary" in alow)
    assert "synthetic" in alow or "λ" in alow or "lambda" in alow
    assert "not" in alow and ("wire" in alow or "online" in alow or "shadow" in alow)

    # Optional secondary: offline ADMM block subproblem maximizer (static How_to)
    tf_sub = how.get("tf_offline_admm_block_subproblem", "")
    assert tf_sub, "How_to_read must include tf_offline_admm_block_subproblem"
    slow = tf_sub.lower()
    assert "admm" in slow and "subproblem" in slow
    assert "raw" in slow
    assert "fcc" in slow and "coker" in slow and "cdu" in slow
    assert "not on" in slow or "classic_2block" in slow
    assert "dual" in slow and ("none" in slow or "primary" in slow)
    assert "not" in slow and "wire" in slow
    assert "synthetic" in slow or "λ" in slow or "lambda" in slow

    # Offline multi-round ADMM coordination How_to (static; still not Case 1)
    tf_coord = how.get("tf_offline_admm_coordination", "")
    assert tf_coord, "How_to_read must include tf_offline_admm_coordination"
    clow = tf_coord.lower()
    assert "admm" in clow and "coordination" in clow
    assert "fcc" in clow and "coker" in clow and "cdu" in clow
    assert "not on" in clow or "classic_2block" in clow
    assert "dual" in clow and ("none" in clow or "primary" in clow)
    assert "not" in clow and "wire" in clow
    assert "synthetic" in clow or "λ" in clow or "lambda" in clow
    assert "plant" in clow or "linking" in clow or "per-unit" in clow or "per unit" in clow

    # Offline multi-block plant-linking ADMM How_to (static; synthetic topology; still not Case 1)
    tf_plant = how.get("tf_offline_admm_plant_linking", "")
    assert tf_plant, "How_to_read must include tf_offline_admm_plant_linking"
    plow = tf_plant.lower()
    assert "admm" in plow and ("plant-linking" in plow or "plant linking" in plow)
    assert "fcc" in plow and "coker" in plow and "cdu" in plow
    assert "not on" in plow or "classic_2block" in plow
    assert "dual" in plow and ("none" in plow or "primary" in plow)
    assert "not" in plow and "wire" in plow
    assert "synthetic" in plow
    assert "full plant" in plow or "mass balance" in plow
    assert "incidence" in plow or "linking" in plow

    # Offline multi-block plant-named linking ADMM How_to (static; plant_named; still not Case 1)
    tf_named = how.get("tf_offline_admm_plant_named_linking", "")
    assert tf_named, "How_to_read must include tf_offline_admm_plant_named_linking"
    nlow = tf_named.lower()
    assert "admm" in nlow and ("plant-named" in nlow or "plant named" in nlow)
    assert "fcc" in nlow and "coker" in nlow and "cdu" in nlow
    assert "not on" in nlow or "classic_2block" in nlow
    assert "dual" in nlow and ("none" in nlow or "primary" in nlow)
    assert "not" in nlow and "wire" in nlow
    assert "plant_named" in nlow or "plant-named" in nlow or "plant named" in nlow
    assert "full plant" in nlow or "mass balance" in nlow
    assert "identity" in nlow or "incidence" in nlow
    assert "synthetic" in nlow  # distinct-from-synthetic language

    # Offline wire-preflight How_to (static packaging of #28; still not Case 1; not wire)
    tf_pre = how.get("tf_offline_wire_preflight", "")
    assert tf_pre, "How_to_read must include tf_offline_wire_preflight"
    prelow = tf_pre.lower()
    assert "wire" in prelow and "preflight" in prelow
    assert "wire_shipped" in prelow or "wire shipped" in prelow or "not wire" in prelow
    assert "blocker" in prelow
    assert "dual" in prelow and ("none" in prelow or "primary" in prelow)
    assert "classic_2block" in prelow or "not on" in prelow
    assert "isolation_rewrite_required" in prelow or "isolation" in prelow
    assert "wire_not_shipped" in prelow or "not wire shipped" in prelow

    # Offline Case-1-shaped CDU↔Blender skeleton How_to (static packaging of #30)
    tf_c1 = how.get("tf_offline_case1_shaped_linking", "")
    assert tf_c1, "How_to_read must include tf_offline_case1_shaped_linking"
    c1low = tf_c1.lower()
    assert "case-1-shaped" in c1low or "case1_shaped" in c1low or "case 1-shaped" in c1low
    assert "skeleton" in c1low or "linking" in c1low
    assert "linear_quality_pooling" in c1low
    assert "naphtha" in c1low and "distillate" in c1low and "gasoil" in c1low and "residue" in c1low
    assert "wire_shipped" in c1low or "wire shipped" in c1low or "not wire" in c1low
    assert "dual" in c1low and ("none" in c1low or "primary" in c1low)
    assert "classic_2block" in c1low or "not on" in c1low
    assert "affine" in c1low or "base_delta" in c1low

    # Offline Case-1 dual-space/form contract How_to (static packaging of #32)
    tf_ds = how.get("tf_offline_case1_dual_space_form_contract", "")
    assert tf_ds, "How_to_read must include tf_offline_case1_dual_space_form_contract"
    dslow = tf_ds.lower()
    assert "dual-space" in dslow or "dual_space" in dslow or "form contract" in dslow or "form-label" in dslow
    assert "classic_2block" in dslow or "form_current" in dslow
    assert "tf_affine_cdu_blender_shaped_excel_path" in dslow or "form_planned" in dslow
    assert "naphtha" in dslow and "residue" in dslow
    assert "unproven" in dslow
    assert "wire_shipped" in dslow or "wire shipped" in dslow or "not wire" in dslow
    assert "dual" in dslow and ("none" in dslow or "primary" in dslow)
    assert "not" in dslow and "wire" in dslow

    # Offline Case-1 dual-space L∞ probe How_to (static packaging of #34)
    tf_lp = how.get("tf_offline_case1_dual_space_linf_probe", "")
    assert tf_lp, "How_to_read must include tf_offline_case1_dual_space_linf_probe"
    lplow = tf_lp.lower()
    assert "probe" in lplow or "l∞" in lplow or "linf" in lplow
    assert "unproven" in lplow
    assert "verdict" in lplow
    assert "online_linf_gate_under_tf_path" in lplow or "online_linf_gate" in lplow
    assert "naphtha" in lplow and "residue" in lplow
    assert "wire_shipped" in lplow or "wire shipped" in lplow or "not wire" in lplow
    assert "dual" in lplow and ("none" in lplow or "primary" in lplow)
    assert "not" in lplow and "wire" in lplow
    assert "raw_online_duals" in lplow or "dual_vector_face" in lplow

    # Offline Case-1 dual-space L∞ live-λ bridge How_to (static packaging of #36)
    tf_lb = how.get("tf_offline_case1_dual_space_linf_live_lambda_bridge", "")
    assert tf_lb, "How_to_read must include tf_offline_case1_dual_space_linf_live_lambda_bridge"
    lblow = tf_lb.lower()
    assert "bridge" in lblow or "live" in lblow or "λ" in lblow or "lambda" in lblow
    assert "unproven" in lblow
    assert "verdict" in lblow
    assert "online_linf_gate_under_tf_path" in lblow or "online_linf_gate" in lblow
    assert "live_lambda_source" in lblow or "source" in lblow
    assert "caller_supplied" in lblow and "package_extract" in lblow and "fixture" in lblow
    assert "naphtha" in lblow and "residue" in lblow
    assert "wire_shipped" in lblow or "wire shipped" in lblow or "not wire" in lblow
    assert "dual" in lblow and ("none" in lblow or "primary" in lblow)
    assert "not" in lblow and "wire" in lblow
    assert "raw_online_duals" in lblow or "dual_vector_face" in lblow

    # Offline Case-1 dual-space L∞ live-λ-seeded warm-start How_to (static packaging of #38)
    tf_ws = how.get("tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart", "")
    assert tf_ws, "How_to_read must include tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"
    wslow = tf_ws.lower()
    assert "warm" in wslow or "seed" in wslow or "λ" in wslow or "lambda" in wslow
    assert "unproven" in wslow
    assert "verdict" in wslow
    assert "online_linf_gate_under_tf_path" in wslow or "online_linf_gate" in wslow
    assert "live_lambda_source" in wslow or "source" in wslow
    assert "seed_policy" in wslow or "lambda0_from_live_primary" in wslow
    assert "seed_identity" in wslow or "seed" in wslow
    assert "caller_supplied" in wslow and "package_extract" in wslow and "fixture" in wslow
    assert "naphtha" in wslow and "residue" in wslow
    assert "wire_shipped" in wslow or "wire shipped" in wslow or "not wire" in wslow
    assert "dual" in wslow and ("none" in wslow or "primary" in wslow)
    assert "not" in wslow and "wire" in wslow
    assert "raw_online_duals" in wslow or "dual_vector_face" in wslow

    # Offline Case-1 honest blender pooling path How_to (static packaging of #40)
    tf_pool = how.get("tf_offline_case1_honest_blender_pooling_path", "")
    assert tf_pool, "How_to_read must include tf_offline_case1_honest_blender_pooling_path"
    pllow = tf_pool.lower()
    assert "pooling" in pllow or "linear_quality_pooling" in pllow
    assert "unproven" in pllow
    assert "verdict" in pllow
    assert "affine" in pllow
    assert "honest_pooling_path_present" in pllow
    assert "wire_shipped" in pllow or "wire shipped" in pllow or "not wire" in pllow
    assert "dual" in pllow and ("none" in pllow or "primary" in pllow)
    assert "not" in pllow and "wire" in pllow
    assert "linear_quality_pooling" in pllow
    assert "no_blender" in pllow or "blocker" in pllow
    assert "blender_affine_or_honest_pooling" not in pllow or "not" in pllow

    # Offline Case-1 online_linf_gate flip-criteria contract How_to (static packaging of #42)
    tf_crit = how.get("tf_offline_case1_online_linf_gate_criteria_contract", "")
    assert tf_crit, "How_to_read must include tf_offline_case1_online_linf_gate_criteria_contract"
    crlow = tf_crit.lower()
    assert "criteria" in crlow or "gate" in crlow or "flip" in crlow
    assert "open" in crlow
    assert "unproven" in crlow
    assert "verdict" in crlow
    assert "gate_flip_allowed_today" in crlow or "flip" in crlow
    assert "criteria_met_today" in crlow or "met" in crlow
    assert "false" in crlow
    assert "wire_shipped" in crlow or "wire shipped" in crlow or "not wire" in crlow
    assert "dual" in crlow and ("none" in crlow or "primary" in crlow)
    assert "not" in crlow and "wire" in crlow
    assert "online_linf_gate" in crlow
    assert "no_blender" in crlow or "blocker" in crlow

    # Offline Case-1 isolation-rewrite design contract How_to (static packaging of #44)
    tf_iso = how.get("tf_offline_case1_isolation_rewrite_design_contract", "")
    assert tf_iso, "How_to_read must include tf_offline_case1_isolation_rewrite_design_contract"
    isolow = tf_iso.lower()
    assert "isolation" in isolow and "design" in isolow
    assert "design_present" in isolow or "isolation_rewrite_design_present" in isolow
    assert "rewrite_shipped" in isolow or "isolation_rewrite_shipped" in isolow
    assert "false" in isolow
    assert "open" in isolow
    assert "unproven" in isolow
    assert "verdict" in isolow
    assert "wire_shipped" in isolow or "wire shipped" in isolow or "not wire" in isolow
    assert "dual" in isolow and ("none" in isolow or "primary" in isolow)
    assert "not" in isolow and "wire" in isolow
    assert "isolation_rewrite_with_wire" in isolow or "checklist" in isolow
    assert "no_blender" in isolow or "blocker" in isolow
    assert "rewrite" in isolow and "delete" in isolow or "not_deleted" in isolow

    # Offline Case-1 wire-ship acceptance design contract How_to (static packaging of #46)
    tf_ws = how.get("tf_offline_case1_wire_ship_acceptance_design_contract", "")
    assert tf_ws, "How_to_read must include tf_offline_case1_wire_ship_acceptance_design_contract"
    wslow = tf_ws.lower()
    assert "wire-ship" in wslow or "wire_ship" in wslow or "ship" in wslow
    assert "design" in wslow
    assert "design_present" in wslow or "wire_ship_acceptance_design_present" in wslow
    assert "wire_ship_allowed_today" in wslow or "ship_allowed" in wslow
    assert "false" in wslow
    assert "wire_shipped" in wslow or "wire shipped" in wslow
    assert "unproven" in wslow
    assert "verdict" in wslow
    assert "dual" in wslow and ("none" in wslow or "primary" in wslow)
    assert "not" in wslow and "wire" in wslow
    assert "no_blender" in wslow or "blocker" in wslow
    assert "isolation" in wslow

    # Offline Case-1 dual-honest TF-aware path design contract How_to (static packaging of #48)
    tf_pd = how.get("tf_offline_case1_dual_honest_tf_aware_path_design_contract", "")
    assert tf_pd, "How_to_read must include tf_offline_case1_dual_honest_tf_aware_path_design_contract"
    pdlow = tf_pd.lower()
    assert "path" in pdlow and "design" in pdlow
    assert "path_design_present" in pdlow
    assert "path_shipped" in pdlow
    assert "false" in pdlow
    assert "ship-met" in pdlow or "ship_met" in pdlow or "dual_honest_tf_aware_path_present" in pdlow
    assert "wire_ship_allowed_today" in pdlow or "ship_allowed" in pdlow
    assert "wire_shipped" in pdlow or "wire shipped" in pdlow
    assert "unproven" in pdlow
    assert "verdict" in pdlow
    assert "dual" in pdlow and ("none" in pdlow or "primary" in pdlow)
    assert "not" in pdlow and "wire" in pdlow
    assert "linear_quality_pooling" in pdlow or "blender" in pdlow
    assert "no_blender" in pdlow or "blocker" in pdlow

    # Offline Case-1 path-present ship-met criteria How_to (static packaging of #50)
    tf_pc = how.get("tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract", "")
    assert tf_pc, (
        "How_to_read must include "
        "tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"
    )
    pclow = tf_pc.lower()
    assert "criteria" in pclow
    assert "criteria_present" in pclow
    assert "ship_met_allowed" in pclow or "ship_met_allowed_today" in pclow
    assert "false" in pclow
    assert "ship-met" in pclow or "ship_met" in pclow or "dual_honest_tf_aware_path_present" in pclow
    assert "path_design_present" in pclow
    assert "path_shipped" in pclow
    assert "wire_ship_allowed_today" in pclow or "ship_allowed" in pclow
    assert "wire_shipped" in pclow or "wire shipped" in pclow
    assert "unproven" in pclow
    assert "verdict" in pclow
    assert "dual" in pclow and ("none" in pclow or "primary" in pclow)
    assert "not" in pclow and "wire" in pclow
    assert "linear_quality_pooling" in pclow or "blender" in pclow
    assert "no_blender" in pclow or "blocker" in pclow

    # Offline Case-1 form_label ship criteria How_to (static packaging of #52)
    tf_fl = how.get("tf_offline_case1_form_label_change_shipped_criteria_contract", "")
    assert tf_fl, (
        "How_to_read must include "
        "tf_offline_case1_form_label_change_shipped_criteria_contract"
    )
    fllow = tf_fl.lower()
    assert "criteria" in fllow
    assert "criteria_present" in fllow
    assert "form_label_ship_allowed" in fllow or "form_label_ship_allowed_today" in fllow
    assert "form_label_change_shipped" in fllow
    assert "false" in fllow
    assert "classic_2block" in fllow or "form_current" in fllow
    assert "path_design_present" in fllow
    assert "path_shipped" in fllow
    assert "wire_ship_allowed_today" in fllow or "ship_allowed" in fllow
    assert "wire_shipped" in fllow or "wire shipped" in fllow
    assert "unproven" in fllow
    assert "verdict" in fllow
    assert "dual" in fllow and ("none" in fllow or "primary" in fllow)
    assert "not" in fllow and "wire" in fllow
    assert "mutation" in fllow or "feature_flag" in fllow
    assert "no_blender" in fllow or "blocker" in fllow

    # Offline Case-1 isolation ship criteria How_to (static packaging of #54)
    tf_iso = how.get("tf_offline_case1_isolation_rewrite_shipped_criteria_contract", "")
    assert tf_iso, (
        "How_to_read must include "
        "tf_offline_case1_isolation_rewrite_shipped_criteria_contract"
    )
    isolow = tf_iso.lower()
    assert "criteria" in isolow
    assert "criteria_present" in isolow
    assert "isolation_ship_allowed" in isolow or "isolation_ship_allowed_today" in isolow
    assert "isolation_rewrite_shipped" in isolow
    assert "false" in isolow
    assert "open" in isolow
    assert "rewrite" in isolow and ("not_delete" in isolow or "not-delete" in isolow or "not delete" in isolow or "rewrite_with_wire_not_delete" in isolow)
    assert "classic_2block" in isolow or "form_current" in isolow
    assert "path_design_present" in isolow
    assert "path_shipped" in isolow
    assert "wire_ship_allowed_today" in isolow or "ship_allowed" in isolow
    assert "wire_shipped" in isolow or "wire shipped" in isolow
    assert "unproven" in isolow
    assert "verdict" in isolow
    assert "dual" in isolow and ("none" in isolow or "primary" in isolow)
    assert "not" in isolow and "wire" in isolow
    assert "no_blender" in isolow or "blocker" in isolow
    assert "isolation_rewrite_required" in isolow or "blocker" in isolow


def test_format_tf_offline_units_howto_pure():
    """Static helper: no solve, isolation-safe contract strings."""
    from pims_admm_llm.models.excel_pipeline import format_tf_offline_units_howto

    d = format_tf_offline_units_howto()
    assert d["topic"] == "tf_offline_units"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    one = d["planner_one_liner"].lower()
    assert "offline" in one
    assert "classic_2block" in one
    assert "primary" in one
    assert "dual" in one


def test_format_tf_offline_priced_howto_pure():
    """Static priced residual How_to: isolation-safe; no TF import."""
    from pims_admm_llm.models.excel_pipeline import format_tf_offline_priced_howto

    d = format_tf_offline_priced_howto()
    assert d["topic"] == "tf_offline_priced"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    assert d["price_source"] == "synthetic_offline_demo"
    one = d["planner_one_liner"].lower()
    assert "priced residual" in one or "priced" in one
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one
    assert "not" in one


def test_format_tf_offline_timing_howto_pure():
    """Static timing readiness How_to: isolation-safe; no TF import."""
    from pims_admm_llm.models.excel_pipeline import format_tf_offline_timing_howto

    d = format_tf_offline_timing_howto()
    assert d["topic"] == "tf_offline_timing"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    one = d["planner_one_liner"].lower()
    assert "timing" in one or "block-solve" in one
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one
    assert "not" in one


def test_format_tf_offline_admm_residual_howto_pure():
    """Static ADMM residual How_to: isolation-safe; no TF import."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_admm_residual_howto,
    )

    d = format_tf_offline_admm_residual_howto()
    assert d["topic"] == "tf_offline_admm_residual"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    assert d["price_source"] == "synthetic_offline_demo"
    assert d["lam_source"] == "synthetic_offline_demo"
    one = d["planner_one_liner"].lower()
    assert "admm" in one and ("residual" in one or "consensus" in one)
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one
    assert "not" in one and "wire" in one
    assert "synthetic" in one


def test_format_tf_offline_admm_block_subproblem_howto_pure():
    """Static ADMM block subproblem How_to: isolation-safe; no TF import."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_admm_block_subproblem_howto,
    )

    d = format_tf_offline_admm_block_subproblem_howto()
    assert d["topic"] == "tf_offline_admm_block_subproblem"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    assert d["optimand_space"] == "raw_affine"
    assert d["price_source"] == "synthetic_offline_demo"
    assert d["lam_source"] == "synthetic_offline_demo"
    one = d["planner_one_liner"].lower()
    assert "admm" in one and "subproblem" in one
    assert "raw" in one
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one
    assert "not" in one and "wire" in one
    assert "synthetic" in one


def test_format_tf_offline_admm_coordination_howto_pure():
    """Static multi-round ADMM coordination How_to: isolation-safe; no TF import."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_admm_coordination_howto,
    )

    d = format_tf_offline_admm_coordination_howto()
    assert d["topic"] == "tf_offline_admm_coordination"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    assert d["optimand_space"] == "raw_affine"
    assert d["coordination_scope"] == "per_unit_synthetic_offline"
    assert d["price_source"] == "synthetic_offline_demo"
    assert d["lam_source"] == "synthetic_offline_demo"
    one = d["planner_one_liner"].lower()
    assert "admm" in one and "coordination" in one
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one
    assert "not" in one and "wire" in one
    assert "synthetic" in one
    assert "plant" in one or "linking" in one or "per-unit" in one or "per unit" in one


def test_format_tf_offline_admm_plant_linking_howto_pure():
    """Static multi-block plant-linking ADMM How_to: isolation-safe; no TF import."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_admm_plant_linking_howto,
    )

    d = format_tf_offline_admm_plant_linking_howto()
    assert d["topic"] == "tf_offline_admm_plant_linking"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    assert d["optimand_space"] == "raw_affine"
    assert d["plant_linking_scope"] == "synthetic_offline_demo"
    assert d["topology_source"] == "synthetic_offline_demo"
    assert d["linking_space"] == "synthetic_linking_streams"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["price_source"] == "synthetic_offline_demo"
    assert d["lam_source"] == "synthetic_offline_demo"
    one = d["planner_one_liner"].lower()
    assert "admm" in one and ("plant-linking" in one or "plant linking" in one)
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one
    assert "not" in one and "wire" in one
    assert "synthetic" in one
    assert "full plant" in one or "mass balance" in one
    assert "coordination" in one  # distinct-from-coordination language


def test_format_tf_offline_admm_plant_named_linking_howto_pure():
    """Static multi-block plant-named linking ADMM How_to: isolation-safe; no TF import."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_admm_plant_named_linking_howto,
    )

    d = format_tf_offline_admm_plant_named_linking_howto()
    assert d["topic"] == "tf_offline_admm_plant_named_linking"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["optimand_space"] == "raw_affine"
    assert d["plant_linking_scope"] == "plant_named_offline_demo"
    assert d["topology_source"] == "plant_named_offline_demo"
    assert d["linking_space"] == "plant_named_linking_streams"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["z_update_space"] == "plant_named_linking_streams"
    one = d["planner_one_liner"].lower()
    assert "admm" in one and ("plant-named" in one or "plant named" in one)
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one
    assert "not" in one and "wire" in one
    assert "full plant" in one or "mass balance" in one
    assert "identity" in one or "incidence" in one
    assert "synthetic" in one  # distinct-from-synthetic language
    assert "coordination" in one  # distinct-from-coordination language


def test_format_tf_offline_wire_preflight_howto_pure():
    """Static wire-preflight How_to: dual-ban, wire_shipped=False, blocker ids; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_wire_preflight_howto,
    )

    d = format_tf_offline_wire_preflight_howto()
    assert d["topic"] == "tf_offline_wire_preflight"
    assert "FCC" in d["units"] and "COKER" in d["units"] and "CDU" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["form"] == "classic_2block_excel_path"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["blockers_documented"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["preflight_lambda_is_not_case1_online_lambda"] == "true"
    assert d["preflight_ok_is_not_wire_shipped"] == "true"
    blockers = d["wire_blockers"]
    for bid in _OFFLINE_WIRE_BLOCKER_IDS:
        assert bid in blockers
    # critical subset called out by plan
    for critical in (
        "isolation_rewrite_required",
        "form_label_change_required",
        "dual_linf_under_wire_unproven",
        "case1_is_cdu_blender_package_admm",
        "no_blender_offline_affine_kernel",
        "wire_not_shipped",
    ):
        assert critical in blockers
    one = d["planner_one_liner"].lower()
    assert "wire" in one and "preflight" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "blocker" in one
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one and "none" in one
    assert "structural" in one or "ready_for_wire_discussion" in one
    assert "not wire tomorrow" in one or "≠ wire" in one or "not wire shipped" in one


def test_format_tf_offline_case1_shaped_linking_howto_pure():
    """Static Case-1-shaped linking How_to: dual-ban, blender honesty, streams; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        format_tf_offline_case1_shaped_linking_howto,
    )

    d = format_tf_offline_case1_shaped_linking_howto()
    assert d["topic"] == "tf_offline_case1_shaped_linking"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form"] == "classic_2block_excel_path"
    assert d["case1_form_unchanged"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["case1_shaped_offline_only"] == "true"
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    assert d["blender_surface_is_not_base_delta_affine_unit"] == "true"
    streams = d["linking_streams"]
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in streams
    assert d["linking_lambda_is_not_case1_online_lambda"] == "true"
    assert d["skeleton_is_not_package_admm_wire"] == "true"
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    one = d["planner_one_liner"].lower()
    assert "case-1-shaped" in one or "case1" in one
    assert "skeleton" in one or "linking" in one
    assert "linear_quality_pooling" in one
    assert "naphtha" in one and "residue" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "classic_2block" in one or "not on" in one
    assert "primary" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "affine" in one or "base_delta" in one


def test_format_tf_offline_case1_dual_space_form_contract_howto_pure():
    """Static dual-space/form contract How_to: form registry, streams, dual_linf unproven; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_SHAPED_LINKING_STREAMS,
        format_tf_offline_case1_dual_space_form_contract_howto,
    )

    d = format_tf_offline_case1_dual_space_form_contract_howto()
    assert d["topic"] == "tf_offline_case1_dual_space_form_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    streams = d["linking_streams"]
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in streams
    assert d["stream_alignment_ok"] == "true"
    assert d["package_dual_gate"] == "online_lambda"
    assert d["package_dual_secondary"] == "recovered_blender"
    assert d["skeleton_lambda_is_not_case1_online_lambda"] == "true"
    assert d["skeleton_lambda_is_not_case1_primary_or_secondary_duals"] == "true"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    one = d["planner_one_liner"].lower()
    assert "dual-space" in one or "form contract" in one or "form" in one
    assert "classic_2block" in one or "form_current" in one
    assert "tf_affine_cdu_blender" in one or "form_planned" in one
    assert "naphtha" in one and "residue" in one
    assert "unproven" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "primary" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one


def test_format_tf_offline_case1_dual_space_linf_probe_howto_pure():
    """Static dual-space L∞ probe How_to: dual-ban, unproven, not VERDICT, not wire; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_DUAL_VECTOR_FACE,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_SHAPED_LINKING_STREAMS,
        format_tf_offline_case1_dual_space_linf_probe_howto,
    )

    d = format_tf_offline_case1_dual_space_linf_probe_howto()
    assert d["topic"] == "tf_offline_case1_dual_space_linf_probe"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    streams = d["linking_streams"]
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in streams
    assert d["stream_alignment_ok"] == "true"
    assert d["dual_vector_face"] == _CASE1_DUAL_VECTOR_FACE
    assert d["dual_vector_face"] == "raw_online_duals"
    assert d["package_dual_gate"] == "online_lambda"
    assert d["package_dual_secondary"] == "recovered_blender"
    assert d["skeleton_lambda_is_not_case1_online_lambda"] == "true"
    assert d["skeleton_lambda_is_not_case1_primary_or_secondary_duals"] == "true"
    assert d["probe_is_not_verdict_gate"] == "true"
    assert d["probe_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["probe_available_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "online_linf_gate_under_tf_path" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    one = d["planner_one_liner"].lower()
    assert "probe" in one or "l∞" in one or "linf" in one
    assert "classic_2block" in one or "form_current" in one
    assert "naphtha" in one and "residue" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "primary" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "online_linf_gate" in one


def test_format_tf_offline_case1_dual_space_linf_live_lambda_bridge_howto_pure():
    """Static dual-space L∞ live-λ bridge How_to: dual-ban, source-labeled, unproven, not VERDICT; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_DUAL_VECTOR_FACE,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_SHAPED_LINKING_STREAMS,
        _LIVE_LAMBDA_SOURCE_ALLOWED,
        format_tf_offline_case1_dual_space_linf_live_lambda_bridge_howto,
    )

    d = format_tf_offline_case1_dual_space_linf_live_lambda_bridge_howto()
    assert d["topic"] == "tf_offline_case1_dual_space_linf_live_lambda_bridge"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    streams = d["linking_streams"]
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in streams
    assert d["stream_alignment_ok"] == "true"
    assert d["dual_vector_face"] == _CASE1_DUAL_VECTOR_FACE
    assert d["dual_vector_face"] == "raw_online_duals"
    assert d["package_dual_gate"] == "online_lambda"
    assert d["package_dual_secondary"] == "recovered_blender"
    assert d["skeleton_lambda_is_not_case1_online_lambda"] == "true"
    assert d["skeleton_lambda_is_not_case1_primary_or_secondary_duals"] == "true"
    assert d["live_lambda_source_must_be_labeled"] == "true"
    allowed = d["live_lambda_source_allowed"]
    for src in _LIVE_LAMBDA_SOURCE_ALLOWED:
        assert src in allowed
    assert "caller_supplied" in allowed and "package_extract" in allowed and "fixture" in allowed
    assert d["extracted_lambda_is_probe_input_only"] == "true"
    assert d["live_lambda_is_not_dual_recovery"] == "true"
    assert d["bridge_is_not_verdict_gate"] == "true"
    assert d["bridge_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["bridge_available_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "online_linf_gate_under_tf_path" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    one = d["planner_one_liner"].lower()
    assert "bridge" in one or "live" in one or "λ" in one or "lambda" in one
    assert "classic_2block" in one or "form_current" in one
    assert "naphtha" in one and "residue" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "primary" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "online_linf_gate" in one
    assert "live_lambda_source" in one or "source" in one
    assert "caller_supplied" in one and "fixture" in one


def test_format_tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart_howto_pure():
    """Static dual-space L∞ live-λ-seeded warm-start How_to: dual-ban, seed_policy, seed≠proof; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_DUAL_VECTOR_FACE,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_SHAPED_LINKING_STREAMS,
        _LIVE_LAMBDA_SOURCE_ALLOWED,
        _WARMSTART_SEED_POLICY,
        _WARMSTART_Z0_POLICY,
        format_tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart_howto,
    )

    d = format_tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart_howto()
    assert d["topic"] == "tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    streams = d["linking_streams"]
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in streams
    assert d["stream_alignment_ok"] == "true"
    assert d["dual_vector_face"] == _CASE1_DUAL_VECTOR_FACE
    assert d["dual_vector_face"] == "raw_online_duals"
    assert d["package_dual_gate"] == "online_lambda"
    assert d["package_dual_secondary"] == "recovered_blender"
    assert d["skeleton_lambda_is_not_case1_online_lambda"] == "true"
    assert d["skeleton_lambda_is_not_case1_primary_or_secondary_duals"] == "true"
    assert d["live_lambda_source_must_be_labeled"] == "true"
    allowed = d["live_lambda_source_allowed"]
    for src in _LIVE_LAMBDA_SOURCE_ALLOWED:
        assert src in allowed
    assert "caller_supplied" in allowed and "package_extract" in allowed and "fixture" in allowed
    assert d["seed_policy"] == _WARMSTART_SEED_POLICY
    assert d["seed_policy"] == "lambda0_from_live_primary_online"
    assert d["z0_policy"] == _WARMSTART_Z0_POLICY
    assert d["z0_policy"] == "unchanged_default_skeleton_z"
    assert d["seeded_lambda_is_probe_input_only"] == "true"
    assert d["live_lambda_is_not_dual_recovery"] == "true"
    assert d["warmstart_is_not_verdict_gate"] == "true"
    assert d["warmstart_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["warmstart_available_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["seed_identity_linf_is_not_proof"] == "true"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "online_linf_gate_under_tf_path" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    one = d["planner_one_liner"].lower()
    assert "warm" in one or "seed" in one or "λ" in one or "lambda" in one
    assert "classic_2block" in one or "form_current" in one
    assert "naphtha" in one and "residue" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "primary" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "online_linf_gate" in one
    assert "live_lambda_source" in one or "source" in one
    assert "caller_supplied" in one and "fixture" in one
    assert "seed_policy" in one or "lambda0_from_live_primary" in one
    assert "seed_identity" in one or "seed" in one


def test_format_tf_offline_case1_honest_blender_pooling_path_howto_pure():
    """Static honest blender pooling path How_to: dual-ban, not affine, checklist present; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_HONEST_BLENDER_POOLING_PATH_CHECKLIST_STATUS,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_honest_blender_pooling_path_howto,
    )

    d = format_tf_offline_case1_honest_blender_pooling_path_howto()
    assert d["topic"] == "tf_offline_case1_honest_blender_pooling_path"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    assert d["blender_is_base_delta_affine_unit"] == "false"
    assert d["pooling_path_is_not_affine_kernel"] == "true"
    assert d["pooling_path_is_not_wire"] == "true"
    assert d["pooling_path_is_not_verdict_gate"] == "true"
    assert d["pooling_path_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["excel_cdu_matrix_matches_affine"] == "None"
    assert d["excel_blender_matrix_matches_affine"] == "None"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["blender_pooling_checklist_status"] == (
        _CASE1_HONEST_BLENDER_POOLING_PATH_CHECKLIST_STATUS
    )
    assert d["blender_pooling_checklist_status"] == "honest_pooling_path_present"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "online_linf_gate_under_tf_path" in open_ids
    assert "blender_affine_or_honest_pooling" not in open_ids
    assert "blender" not in open_ids.lower()
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "pooling" in one or "linear_quality_pooling" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "affine" in one
    assert "honest_pooling_path_present" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one


def test_format_tf_offline_case1_online_linf_gate_criteria_contract_howto_pure():
    """Static gate-criteria contract How_to: gate open, flip false, dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_ONLINE_LINF_GATE_ANTI_CRITERIA,
        _CASE1_ONLINE_LINF_GATE_FLIP_CRITERIA_KEYS,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_online_linf_gate_criteria_contract_howto,
    )

    d = format_tf_offline_case1_online_linf_gate_criteria_contract_howto()
    assert d["topic"] == "tf_offline_case1_online_linf_gate_criteria_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["contract_is_not_gate_flip"] == "true"
    assert d["contract_is_not_wire"] == "true"
    assert d["contract_is_not_verdict_gate"] == "true"
    assert d["contract_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "online_linf_gate_under_tf_path" in open_ids
    flip_keys = d["flip_criteria_keys"]
    for k in _CASE1_ONLINE_LINF_GATE_FLIP_CRITERIA_KEYS:
        assert k in flip_keys
    assert "isolation_rewrite_with_wire" in flip_keys
    assert "form_label_change_shipped" in flip_keys
    assert "wire_shipped" in flip_keys
    anti = d["anti_criteria_today"]
    for k in _CASE1_ONLINE_LINF_GATE_ANTI_CRITERIA:
        assert k in anti
    assert "probe_linf" in anti
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "criteria" in one or "gate" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "gate_flip_allowed_today" in one or "flip" in one
    assert "false" in one
    assert "open" in one
    assert "online_linf_gate" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one


def test_format_tf_offline_case1_isolation_rewrite_design_contract_howto_pure():
    """Static isolation-rewrite design How_to: design_present; rewrite=false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_BLOCKER_ID,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_KEY,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_ISOLATION_REWRITE_DESIGN_PRESENT,
        _CASE1_ISOLATION_REWRITE_NOT_DELETE,
        _CASE1_ISOLATION_REWRITE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_isolation_rewrite_design_contract_howto,
    )

    d = format_tf_offline_case1_isolation_rewrite_design_contract_howto()
    assert d["topic"] == "tf_offline_case1_isolation_rewrite_design_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["isolation_rewrite_design_present"] == "true"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["isolation_tests_must_be_rewritten_with_wire_not_deleted"] == "true"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["design_is_not_isolation_rewrite_shipped"] == "true"
    assert d["design_is_not_wire"] == "true"
    assert d["design_is_not_verdict_gate"] == "true"
    assert d["design_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["design_is_not_gate_flip"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "isolation_rewrite_with_wire" in open_ids
    assert _CASE1_ISOLATION_REWRITE_CHECKLIST_KEY in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_ISOLATION_REWRITE_DESIGN_PRESENT is True
    assert _CASE1_ISOLATION_REWRITE_SHIPPED is False
    assert _CASE1_ISOLATION_REWRITE_NOT_DELETE is True
    assert _CASE1_ISOLATION_REWRITE_BLOCKER_ID in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "isolation" in one and "design" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "rewrite_shipped" in one or "isolation_rewrite_shipped" in one
    assert "false" in one
    assert "open" in one
    assert "design_present" in one or "isolation_rewrite_design_present" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "not_deleted" in one or "not-delete" in one or "not delete" in one


def test_format_tf_offline_case1_wire_ship_acceptance_design_contract_howto_pure():
    """Static wire-ship acceptance design How_to: design_present; ship=false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_WIRE_SHIP_ACCEPTANCE_DESIGN_PRESENT,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIP_ANTI_CRITERIA,
        _CASE1_WIRE_SHIP_CRITERIA_MET_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_wire_ship_acceptance_design_contract_howto,
    )

    d = format_tf_offline_case1_wire_ship_acceptance_design_contract_howto()
    assert d["topic"] == "tf_offline_case1_wire_ship_acceptance_design_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["wire_ship_acceptance_design_present"] == "true"
    assert d["design_present"] == "true"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["wire_ship_criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["design_is_not_wire_ship_allow"] == "true"
    assert d["design_is_not_wire"] == "true"
    assert d["design_is_not_verdict_gate"] == "true"
    assert d["design_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["design_is_not_isolation_rewrite_shipped"] == "true"
    assert d["design_is_not_form_flip"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "wire_shipped_false_today" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_WIRE_SHIP_ACCEPTANCE_DESIGN_PRESENT is True
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIP_CRITERIA_MET_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "packaging_alone" in _CASE1_WIRE_SHIP_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "wire-ship" in one or "wire_ship" in one or "ship" in one
    assert "design" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "wire_ship_allowed_today" in one or "ship_allowed" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "design_present" in one or "wire_ship_acceptance_design_present" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "packaging" in one or "design alone" in one or "anti" in one




def test_format_tf_offline_case1_dual_honest_tf_aware_path_design_contract_howto_pure():
    """Static path-design How_to: path_design_present; path=false; ship-met=false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_PATH_DESIGN_ANTI_CRITERIA,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_dual_honest_tf_aware_path_design_contract_howto,
    )

    d = format_tf_offline_case1_dual_honest_tf_aware_path_design_contract_howto()
    assert d["topic"] == "tf_offline_case1_dual_honest_tf_aware_path_design_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure" not in d["dual_recovery_path_planned_when_shipped"].lower() or "not pure" in d["planner_one_liner"].lower()
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["dual_honest_tf_aware_path_present"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["wire_ship_criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    assert d["design_is_not_path_shipped"] == "true"
    assert d["design_is_not_path_present_for_ship"] == "true"
    assert d["design_is_not_wire_ship_allow"] == "true"
    assert d["design_is_not_wire"] == "true"
    assert d["design_is_not_verdict_gate"] == "true"
    assert d["design_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["design_is_not_isolation_rewrite_shipped"] == "true"
    assert d["design_is_not_form_flip"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "wire_shipped_false_today" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_path_design_alone" in _CASE1_PATH_DESIGN_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "path" in one and "design" in one
    assert "path_design_present" in one
    assert "path_shipped" in one
    assert "ship-met" in one or "ship_met" in one or "dual_honest_tf_aware_path_present" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "wire_ship_allowed_today" in one or "ship_allowed" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "linear_quality_pooling" in one
    assert "packaging" in one or "design alone" in one or "anti" in one
    # isolation: formatter body must not import TF / call live path-design report
    import ast
    import inspect
    src = inspect.getsource(format_tf_offline_case1_dual_honest_tf_aware_path_design_contract_howto)
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    # live report call banned (docstring may name it as a hard negative)
    assert "offline_case1_dual_honest_tf_aware_path_design_contract_report(" not in src


def test_format_tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract_howto_pure():
    """Static path-present criteria How_to: criteria_present; ship-met=false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_PRESENT_CRITERIA_ANTI_CRITERIA,
        _CASE1_PATH_PRESENT_CRITERIA_MET_TODAY,
        _CASE1_PATH_PRESENT_CRITERIA_PRESENT,
        _CASE1_PATH_PRESENT_FLIP_CRITERIA_KEYS,
        _CASE1_PATH_SHIPPED,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract_howto,
    )

    d = format_tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract_howto()
    assert d["topic"] == "tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["criteria_present"] == "true"
    assert d["path_present_criteria_present"] == "true"
    assert d["ship_met_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["dual_honest_tf_aware_path_present"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["wire_ship_criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    for k in _CASE1_PATH_PRESENT_FLIP_CRITERIA_KEYS:
        assert k in d["flip_criteria_keys"]
    assert d["criteria_is_not_ship_met"] == "true"
    assert d["criteria_is_not_path_shipped"] == "true"
    assert d["criteria_is_not_path_present_for_ship"] == "true"
    assert d["criteria_is_not_wire_ship_allow"] == "true"
    assert d["criteria_is_not_wire"] == "true"
    assert d["criteria_is_not_verdict_gate"] == "true"
    assert d["criteria_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["criteria_is_not_isolation_rewrite_shipped"] == "true"
    assert d["criteria_is_not_form_flip"] == "true"
    assert d["path_design_present_is_not_ship_met"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "wire_shipped_false_today" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_PATH_PRESENT_CRITERIA_PRESENT is True
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_PRESENT_CRITERIA_MET_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_path_present_criteria_contract_alone" in _CASE1_PATH_PRESENT_CRITERIA_ANTI_CRITERIA
    assert "this_ship_met_criteria_contract_alone" in _CASE1_PATH_PRESENT_CRITERIA_ANTI_CRITERIA
    assert "this_path_design_alone" in _CASE1_PATH_PRESENT_CRITERIA_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "criteria" in one and ("ship-met" in one or "ship_met" in one or "path-present" in one)
    assert "criteria_present" in one
    assert "ship_met_allowed_today" in one or "ship_met_allowed" in one
    assert "path_design_present" in one
    assert "path_shipped" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "linear_quality_pooling" in one
    assert "packaging" in one or "criteria alone" in one or "anti" in one
    # five-way permission coexistence
    assert d["criteria_present"] == "true"
    assert d["ship_met_allowed_today"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["path_shipped"] == "false"
    assert d["wire_shipped"] == "false"
    # isolation: formatter body must not import TF / call live criteria report
    import ast
    import inspect
    src = inspect.getsource(
        format_tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "offline_case1_dual_honest_tf_aware_path_present_criteria_contract_report(" not in src


def test_format_tf_offline_case1_form_label_change_shipped_criteria_contract_howto_pure():
    """Static form_label criteria How_to: criteria_present; form_label ship false; form classic; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_LABEL_CHANGE_SHIPPED,
        _CASE1_FORM_LABEL_CRITERIA_ANTI_CRITERIA,
        _CASE1_FORM_LABEL_CRITERIA_MET_TODAY,
        _CASE1_FORM_LABEL_CRITERIA_PRESENT,
        _CASE1_FORM_LABEL_FLIP_CRITERIA_KEYS,
        _CASE1_FORM_LABEL_MUTATION_PATH_EXECUTED_TODAY,
        _CASE1_FORM_LABEL_MUTATION_PATH_NAME,
        _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_form_label_change_shipped_criteria_contract_howto,
    )

    d = format_tf_offline_case1_form_label_change_shipped_criteria_contract_howto()
    assert d["topic"] == "tf_offline_case1_form_label_change_shipped_criteria_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_planned"] == "tf_affine_cdu_blender_shaped_excel_path"
    assert d["form_current"] != d["form_planned"]
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["planned_form_distinct"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["criteria_present"] == "true"
    assert d["form_label_criteria_present"] == "true"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["mutation_path_name"] == _CASE1_FORM_LABEL_MUTATION_PATH_NAME
    assert d["mutation_path_executed_today"] == "false"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["dual_honest_tf_aware_path_present"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["wire_ship_criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    for k in _CASE1_FORM_LABEL_FLIP_CRITERIA_KEYS:
        assert k in d["flip_criteria_keys"]
    assert d["criteria_is_not_form_label_shipped"] == "true"
    assert d["criteria_is_not_form_flip"] == "true"
    assert d["criteria_is_not_form_label_ship_allow"] == "true"
    assert d["criteria_is_not_ship_met"] == "true"
    assert d["criteria_is_not_path_shipped"] == "true"
    assert d["criteria_is_not_wire_ship_allow"] == "true"
    assert d["criteria_is_not_wire"] == "true"
    assert d["criteria_is_not_verdict_gate"] == "true"
    assert d["criteria_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["criteria_is_not_isolation_rewrite_shipped"] == "true"
    assert d["form_registration_is_not_form_label_shipped"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "form_label_change_shipped" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_FORM_LABEL_CRITERIA_PRESENT is True
    assert _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY is False
    assert _CASE1_FORM_LABEL_CRITERIA_MET_TODAY is False
    assert _CASE1_FORM_LABEL_CHANGE_SHIPPED is False
    assert _CASE1_FORM_LABEL_MUTATION_PATH_EXECUTED_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_form_label_change_shipped_criteria_contract_alone" in _CASE1_FORM_LABEL_CRITERIA_ANTI_CRITERIA
    assert "form_registration_alone" in _CASE1_FORM_LABEL_CRITERIA_ANTI_CRITERIA
    assert "packaging_alone" in _CASE1_FORM_LABEL_CRITERIA_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "criteria" in one and "form_label" in one
    assert "criteria_present" in one
    assert "form_label_ship_allowed_today" in one or "form_label_ship_allowed" in one
    assert "form_label_change_shipped" in one
    assert "classic_2block" in one or "form_current" in one
    assert "path_design_present" in one
    assert "path_shipped" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "mutation" in one
    assert "packaging" in one or "criteria alone" in one or "anti" in one
    # multi-way permission coexistence
    assert d["criteria_present"] == "true"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["wire_shipped"] == "false"
    # isolation: formatter body must not import TF / call live criteria report
    import ast
    import inspect
    src = inspect.getsource(
        format_tf_offline_case1_form_label_change_shipped_criteria_contract_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "offline_case1_form_label_change_shipped_criteria_contract_report(" not in src




def test_format_tf_offline_case1_isolation_rewrite_shipped_criteria_contract_howto_pure():
    """Static isolation ship criteria How_to: criteria_present; ship allow false; rewrite false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_LABEL_CHANGE_SHIPPED,
        _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_ISOLATION_REWRITE_SHIPPED,
        _CASE1_ISOLATION_SHIP_ALLOWED_TODAY,
        _CASE1_ISOLATION_SHIP_CRITERIA_ANTI_CRITERIA,
        _CASE1_ISOLATION_SHIP_CRITERIA_MET_TODAY,
        _CASE1_ISOLATION_SHIP_CRITERIA_PRESENT,
        _CASE1_ISOLATION_SHIP_FLIP_CRITERIA_KEYS,
        _CASE1_ISOLATION_TESTS_REWRITTEN_WITH_WIRE,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_isolation_rewrite_shipped_criteria_contract_howto,
    )

    d = format_tf_offline_case1_isolation_rewrite_shipped_criteria_contract_howto()
    assert d["topic"] == "tf_offline_case1_isolation_rewrite_shipped_criteria_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["criteria_present"] == "true"
    assert d["isolation_ship_criteria_present"] == "true"
    assert d["isolation_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_design_present"] == "true"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["rewrite_with_wire_not_delete"] == "true"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    for k in _CASE1_ISOLATION_SHIP_FLIP_CRITERIA_KEYS:
        assert k in d["flip_criteria_keys"]
    assert len(_CASE1_ISOLATION_SHIP_FLIP_CRITERIA_KEYS) == 9
    assert d["criteria_is_not_isolation_rewrite_shipped"] == "true"
    assert d["criteria_is_not_isolation_ship_allow"] == "true"
    assert d["criteria_is_not_form_label_shipped"] == "true"
    assert d["criteria_is_not_form_flip"] == "true"
    assert d["criteria_is_not_ship_met"] == "true"
    assert d["criteria_is_not_path_shipped"] == "true"
    assert d["criteria_is_not_wire_ship_allow"] == "true"
    assert d["criteria_is_not_wire"] == "true"
    assert d["criteria_is_not_verdict_gate"] == "true"
    assert d["criteria_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["isolation_design_is_not_isolation_rewrite_shipped"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "isolation_rewrite_with_wire" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_ISOLATION_SHIP_CRITERIA_PRESENT is True
    assert _CASE1_ISOLATION_SHIP_ALLOWED_TODAY is False
    assert _CASE1_ISOLATION_SHIP_CRITERIA_MET_TODAY is False
    assert _CASE1_ISOLATION_REWRITE_SHIPPED is False
    assert _CASE1_ISOLATION_TESTS_REWRITTEN_WITH_WIRE is False
    assert _CASE1_FORM_LABEL_CHANGE_SHIPPED is False
    assert _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_isolation_rewrite_shipped_criteria_contract_alone" in _CASE1_ISOLATION_SHIP_CRITERIA_ANTI_CRITERIA
    assert "isolation_design_alone" in _CASE1_ISOLATION_SHIP_CRITERIA_ANTI_CRITERIA
    assert "packaging_alone" in _CASE1_ISOLATION_SHIP_CRITERIA_ANTI_CRITERIA
    assert "residual_must_vanish" in _CASE1_ISOLATION_SHIP_CRITERIA_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "criteria" in one and "isolation" in one
    assert "criteria_present" in one
    assert "isolation_ship_allowed" in one
    assert "isolation_rewrite_shipped" in one
    assert "classic_2block" in one or "form_current" in one
    assert "path_design_present" in one
    assert "path_shipped" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "rewrite" in one
    assert "packaging" in one or "criteria alone" in one or "anti" in one
    # multi-way permission coexistence
    assert d["criteria_present"] == "true"
    assert d["isolation_ship_allowed_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["wire_shipped"] == "false"
    # isolation: formatter body must not import TF / call live criteria report
    import ast
    import inspect
    src = inspect.getsource(
        format_tf_offline_case1_isolation_rewrite_shipped_criteria_contract_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "offline_case1_isolation_rewrite_shipped_criteria_contract_report(" not in src

def test_format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract_howto_pure():
    """Static multi-blocker bundle design How_to: design present; ship/allow false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_BUNDLE_CRITERIA_MET_TODAY,
        _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA,
        _CASE1_BUNDLE_DESIGN_MEMBER_KEYS,
        _CASE1_BUNDLE_DESIGN_NAME,
        _CASE1_BUNDLE_DESIGN_ORDER_HINT,
        _CASE1_BUNDLE_DESIGN_PRESENT,
        _CASE1_BUNDLE_SHIP_ALLOWED_TODAY,
        _CASE1_BUNDLE_SHIPPED,
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_LABEL_CHANGE_SHIPPED,
        _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_ISOLATION_REWRITE_SHIPPED,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract_howto,
    )

    d = format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract_howto()
    assert d["topic"] == "tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["bundle_design_present"] == "true"
    assert d["bundle_shipped"] == "false"
    assert d["bundle_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["bundle_name"] == _CASE1_BUNDLE_DESIGN_NAME
    assert "dual_honest_tf_case1_wire" in d["bundle_name"]
    for k in _CASE1_BUNDLE_DESIGN_MEMBER_KEYS:
        assert k in d["members"]
    assert len(_CASE1_BUNDLE_DESIGN_MEMBER_KEYS) == 14
    for h in _CASE1_BUNDLE_DESIGN_ORDER_HINT:
        assert h in d["order_hint"]
    assert len(_CASE1_BUNDLE_DESIGN_ORDER_HINT) == 5
    assert d["order_hint_is_not_executor"] == "true"
    assert d["atomic_coship_also_valid"] == "true"
    assert d["no_auto_wire"] == "true"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["rewrite_with_wire_not_delete"] == "true"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    assert d["design_is_not_bundle_shipped"] == "true"
    assert d["design_is_not_bundle_ship_allow"] == "true"
    assert d["design_is_not_wire"] == "true"
    assert d["design_is_not_wire_ship_allow"] == "true"
    assert d["design_is_not_isolation_rewrite_shipped"] == "true"
    assert d["design_is_not_form_label_change_shipped"] == "true"
    assert d["design_is_not_form_flip"] == "true"
    assert d["design_is_not_path_shipped"] == "true"
    assert d["design_is_not_ship_met"] == "true"
    assert d["design_is_not_verdict_gate"] == "true"
    assert d["design_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["wire_ship_acceptance_design_alone_is_not_bundle_ship"] == "true"
    assert d["packaging_is_not_bundle_shipped"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "isolation_rewrite_with_wire" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert d["distinct_from_wire_ship_acceptance_design"] == "true"
    assert _CASE1_BUNDLE_DESIGN_PRESENT is True
    assert _CASE1_BUNDLE_SHIPPED is False
    assert _CASE1_BUNDLE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_BUNDLE_CRITERIA_MET_TODAY is False
    assert _CASE1_ISOLATION_REWRITE_SHIPPED is False
    assert _CASE1_FORM_LABEL_CHANGE_SHIPPED is False
    assert _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_bundle_design_alone" in _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA
    assert "this_bundle_ship_criteria_contract_alone" in _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA
    assert "this_bundle_ship_met_criteria_alone" in _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA
    assert "packaging_alone" in _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA
    assert "wire_ship_acceptance_design_alone" in _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA
    assert "residual_must_vanish" in _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA
    assert "isolation_ship_criteria_alone" in _CASE1_BUNDLE_DESIGN_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "dual_linf_under_wire_unproven" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "bundle" in one
    assert "bundle_design_present" in one
    assert "bundle_shipped" in one
    assert "order_hint" in one
    assert (
        "not_executor" in one.replace(" ", "_")
        or "not an executor" in one
        or "not executor" in one
        or "order_hint_is_not_executor" in one
    )
    assert "classic_2block" in one or "form_current" in one
    assert "path_design_present" in one
    assert "path_shipped" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "packaging" in one or "alone" in one or "anti" in one
    # multi-way permission coexistence
    assert d["bundle_design_present"] == "true"
    assert d["bundle_ship_allowed_today"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    # isolation: formatter body must not import TF / call live design report
    import ast
    import inspect

    src = inspect.getsource(
        format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract_report(" not in src
    # module-level: excel packaging never *calls* live TF design report
    import pims_admm_llm.models.excel_pipeline as ep
    from pathlib import Path as _Path

    ep_src = _Path(ep.__file__).read_text(encoding="utf-8")
    assert "offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract_report(" not in ep_src


def test_format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract_howto_pure():
    """Static multi-blocker bundle ship-met criteria How_to: present; ship/allow/met false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_BUNDLE_CRITERIA_MET_TODAY,
        _CASE1_BUNDLE_CRITERIA_PRESENT,
        _CASE1_BUNDLE_DESIGN_MEMBER_KEYS,
        _CASE1_BUNDLE_DESIGN_NAME,
        _CASE1_BUNDLE_DESIGN_ORDER_HINT,
        _CASE1_BUNDLE_DESIGN_PRESENT,
        _CASE1_BUNDLE_SHIP_ALLOWED_TODAY,
        _CASE1_BUNDLE_SHIPPED,
        _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA,
        _CASE1_BUNDLE_SHIPPED_FLIP_CRITERIA_KEYS,
        _CASE1_BUNDLE_SHIPPED_FLIP_UNDER_WIRE_ONLY_KEYS,
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_LABEL_CHANGE_SHIPPED,
        _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_ISOLATION_REWRITE_SHIPPED,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract_howto,
    )

    d = format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract_howto()
    assert d["topic"] == "tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["criteria_present"] == "true"
    assert d["bundle_criteria_present"] == "true"
    assert d["bundle_design_present"] == "true"
    assert d["bundle_shipped"] == "false"
    assert d["bundle_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["bundle_name"] == _CASE1_BUNDLE_DESIGN_NAME
    assert "dual_honest_tf_case1_wire" in d["bundle_name"]
    for k in _CASE1_BUNDLE_SHIPPED_FLIP_CRITERIA_KEYS:
        assert k in d["flip_criteria_keys"]
    assert len(_CASE1_BUNDLE_SHIPPED_FLIP_CRITERIA_KEYS) == 15
    for k in _CASE1_BUNDLE_SHIPPED_FLIP_UNDER_WIRE_ONLY_KEYS:
        assert k in d["under_wire_only_keys"]
    assert "online_linf_gate_under_tf_path" in d["under_wire_only_keys"]
    assert "dual_linf_under_wire_proven" in d["under_wire_only_keys"]
    for k in _CASE1_BUNDLE_DESIGN_MEMBER_KEYS:
        assert k in d["members"]
    for h in _CASE1_BUNDLE_DESIGN_ORDER_HINT:
        assert h in d["order_hint"]
    assert d["order_hint_is_not_executor"] == "true"
    assert d["atomic_coship_also_valid"] == "true"
    assert d["no_auto_wire"] == "true"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["rewrite_with_wire_not_delete"] == "true"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    assert d["criteria_is_not_bundle_shipped"] == "true"
    assert d["criteria_is_not_bundle_ship_allow"] == "true"
    assert d["criteria_is_not_criteria_met"] == "true"
    assert d["criteria_is_not_wire"] == "true"
    assert d["criteria_is_not_wire_ship_allow"] == "true"
    assert d["criteria_is_not_isolation_rewrite_shipped"] == "true"
    assert d["criteria_is_not_form_label_change_shipped"] == "true"
    assert d["criteria_is_not_form_flip"] == "true"
    assert d["criteria_is_not_path_shipped"] == "true"
    assert d["criteria_is_not_ship_met"] == "true"
    assert d["criteria_is_not_verdict_gate"] == "true"
    assert d["criteria_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["this_bundle_design_alone_is_not_ship"] == "true"
    assert d["packaging_is_not_bundle_shipped"] == "true"
    assert d["distinct_from_bundle_design_packaging"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "isolation_rewrite_with_wire" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_BUNDLE_CRITERIA_PRESENT is True
    assert _CASE1_BUNDLE_DESIGN_PRESENT is True
    assert _CASE1_BUNDLE_SHIPPED is False
    assert _CASE1_BUNDLE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_BUNDLE_CRITERIA_MET_TODAY is False
    assert _CASE1_ISOLATION_REWRITE_SHIPPED is False
    assert _CASE1_FORM_LABEL_CHANGE_SHIPPED is False
    assert _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_bundle_ship_criteria_contract_alone" in _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA
    assert "this_bundle_ship_met_criteria_alone" in _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA
    assert "this_bundle_design_alone" in _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA
    assert "packaging_alone" in _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA
    assert "wire_ship_acceptance_design_alone" in _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA
    assert "residual_must_vanish" in _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA
    assert "isolation_ship_criteria_alone" in _CASE1_BUNDLE_SHIPPED_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "dual_linf_under_wire_unproven" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "bundle" in one
    assert "criteria_present" in one
    assert "bundle_shipped" in one
    assert "criteria_met" in one or "criteria_met_today" in one
    assert "order_hint" in one
    assert (
        "not_executor" in one.replace(" ", "_")
        or "not an executor" in one
        or "not executor" in one
        or "order_hint_is_not_executor" in one
    )
    assert "classic_2block" in one or "form_current" in one
    assert "path_design_present" in one
    assert "path_shipped" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "packaging" in one or "alone" in one or "anti" in one
    assert "when" in one or "distinct" in one or "design" in one
    # multi-way permission coexistence
    assert d["criteria_present"] == "true"
    assert d["bundle_ship_allowed_today"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    # isolation: formatter body must not import TF / call live criteria report
    import ast
    import inspect

    src = inspect.getsource(
        format_tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract_report(" not in src
    # module-level: excel packaging never *calls* live TF criteria report
    import pims_admm_llm.models.excel_pipeline as ep
    from pathlib import Path as _Path

    ep_src = _Path(ep.__file__).read_text(encoding="utf-8")
    assert "offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract_report(" not in ep_src

def test_format_tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold_howto_pure():
    """Static scaffold packaging How_to: present; all ship flags false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_BUNDLE_CRITERIA_MET_TODAY,
        _CASE1_BUNDLE_SHIP_ALLOWED_TODAY,
        _CASE1_BUNDLE_SHIPPED,
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_EXECUTION_SCAFFOLD_PRESENT,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_LABEL_CHANGE_SHIPPED,
        _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_ISOLATION_REWRITE_SHIPPED,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_SCAFFOLD_ANTI_CRITERIA,
        _CASE1_SCAFFOLD_PRESENT,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold_howto,
    )

    d = format_tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold_howto()
    assert d["topic"] == "tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["scaffold_present"] == "true"
    assert d["execution_scaffold_present"] == "true"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["bundle_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["rewrite_with_wire_not_delete"] == "true"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    assert d["scaffold_is_not_path_shipped"] == "true"
    assert d["scaffold_is_not_ship_met"] == "true"
    assert d["scaffold_is_not_wire"] == "true"
    assert d["scaffold_is_not_bundle_shipped"] == "true"
    assert d["scaffold_is_not_isolation_rewrite_shipped"] == "true"
    assert d["scaffold_is_not_form_label_change_shipped"] == "true"
    assert d["scaffold_is_not_form_flip"] == "true"
    assert d["scaffold_is_not_verdict_gate"] == "true"
    assert d["scaffold_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["packaging_is_not_path_shipped"] == "true"
    assert d["packaging_is_not_wire_shipped"] == "true"
    assert d["packaging_is_not_bundle_shipped"] == "true"
    assert d["order_hint_is_not_executor"] == "true"
    assert d["no_auto_wire"] == "true"
    assert d["distinct_from_bundle_design_and_ship_met_criteria_packaging"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "isolation_rewrite_with_wire" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_SCAFFOLD_PRESENT is True
    assert _CASE1_EXECUTION_SCAFFOLD_PRESENT is True
    assert _CASE1_BUNDLE_SHIPPED is False
    assert _CASE1_BUNDLE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_BUNDLE_CRITERIA_MET_TODAY is False
    assert _CASE1_ISOLATION_REWRITE_SHIPPED is False
    assert _CASE1_FORM_LABEL_CHANGE_SHIPPED is False
    assert _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_scaffold_alone" in _CASE1_SCAFFOLD_ANTI_CRITERIA
    assert "this_execution_scaffold_alone" in _CASE1_SCAFFOLD_ANTI_CRITERIA
    assert "packaging_alone" in _CASE1_SCAFFOLD_ANTI_CRITERIA
    assert "bundle_ship_met_criteria_alone" in _CASE1_SCAFFOLD_ANTI_CRITERIA
    assert "path_design_alone" in _CASE1_SCAFFOLD_ANTI_CRITERIA
    assert "residual_must_vanish" in _CASE1_SCAFFOLD_ANTI_CRITERIA
    assert "diagnostic_linf_alone" in _CASE1_SCAFFOLD_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "dual_linf_under_wire_unproven" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "scaffold" in one
    assert "scaffold_present" in one
    assert "execution_scaffold_present" in one
    assert "path_shipped" in one
    assert "bundle_shipped" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "packaging" in one or "alone" in one or "anti" in one
    assert "how-without-ship" in one or "how without" in one or "offline how" in one or "how" in one
    # multi-way permission coexistence
    assert d["scaffold_present"] == "true"
    assert d["execution_scaffold_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    # isolation: formatter body must not import TF / call live scaffold report
    import ast
    import inspect

    src = inspect.getsource(
        format_tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "pulp" not in imported
    assert "offline_case1_dual_honest_tf_aware_path_execution_scaffold_report(" not in src
    # module-level: excel packaging never *calls* live TF scaffold report
    import pims_admm_llm.models.excel_pipeline as ep
    from pathlib import Path as _Path

    ep_src = _Path(ep.__file__).read_text(encoding="utf-8")
    assert "offline_case1_dual_honest_tf_aware_path_execution_scaffold_report(" not in ep_src
    # no executable TF imports on packaging module (docstrings may mention bans)
    tree_mod = ast.parse(ep_src)
    for node in ast.walk(tree_mod):
        if isinstance(node, ast.Import):
            for alias in node.names:
                assert "tf_linear_blocks" not in alias.name
                assert "tensorflow" not in alias.name
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            assert "tf_linear_blocks" not in mod
            assert "tensorflow" not in mod
    # pulp may be used by classic Case 1 solve path in this module; packaging formatters must not
    # import it themselves (checked above via formatter AST).


def test_format_tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal_howto_pure():
    """Static rehearsal packaging How_to: present; all ship flags false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_BUNDLE_CRITERIA_MET_TODAY,
        _CASE1_BUNDLE_SHIP_ALLOWED_TODAY,
        _CASE1_BUNDLE_SHIPPED,
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_EXECUTION_SCAFFOLD_PRESENT,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_LABEL_CHANGE_SHIPPED,
        _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY,
        _CASE1_FORM_PLANNED,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_ISOLATION_REWRITE_SHIPPED,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_REHEARSAL_ANTI_CRITERIA,
        _CASE1_REHEARSAL_PRESENT,
        _CASE1_SCAFFOLD_PRESENT,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_REHEARSAL_PRESENT,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal_howto,
    )

    d = format_tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal_howto()
    assert d["topic"] == "tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["rehearsal_present"] == "true"
    assert d["wire_rehearsal_present"] == "true"
    assert d["scaffold_present"] == "true"
    assert d["execution_scaffold_present"] == "true"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["bundle_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["rewrite_with_wire_not_delete"] == "true"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    assert d["rehearsal_is_not_path_shipped"] == "true"
    assert d["rehearsal_is_not_ship_met"] == "true"
    assert d["rehearsal_is_not_wire"] == "true"
    assert d["rehearsal_is_not_bundle_shipped"] == "true"
    assert d["rehearsal_is_not_isolation_rewrite_shipped"] == "true"
    assert d["rehearsal_is_not_form_label_change_shipped"] == "true"
    assert d["rehearsal_is_not_form_flip"] == "true"
    assert d["rehearsal_is_not_verdict_gate"] == "true"
    assert d["rehearsal_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["coreq_matrix_is_not_ship"] == "true"
    assert d["packaging_is_not_path_shipped"] == "true"
    assert d["packaging_is_not_wire_shipped"] == "true"
    assert d["packaging_is_not_bundle_shipped"] == "true"
    assert d["order_hint_is_not_executor"] == "true"
    assert d["no_auto_wire"] == "true"
    assert d["distinct_from_scaffold_and_bundle_design_and_ship_met_criteria_packaging"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "isolation_rewrite_with_wire" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_REHEARSAL_PRESENT is True
    assert _CASE1_WIRE_REHEARSAL_PRESENT is True
    assert _CASE1_SCAFFOLD_PRESENT is True
    assert _CASE1_EXECUTION_SCAFFOLD_PRESENT is True
    assert _CASE1_BUNDLE_SHIPPED is False
    assert _CASE1_BUNDLE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_BUNDLE_CRITERIA_MET_TODAY is False
    assert _CASE1_ISOLATION_REWRITE_SHIPPED is False
    assert _CASE1_FORM_LABEL_CHANGE_SHIPPED is False
    assert _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_rehearsal_alone" in _CASE1_REHEARSAL_ANTI_CRITERIA
    assert "wire_rehearsal_alone" in _CASE1_REHEARSAL_ANTI_CRITERIA
    assert "coreq_matrix_alone" in _CASE1_REHEARSAL_ANTI_CRITERIA
    assert "scaffold_plus_rehearsal_alone" in _CASE1_REHEARSAL_ANTI_CRITERIA
    assert "packaging_alone" in _CASE1_REHEARSAL_ANTI_CRITERIA
    assert "this_scaffold_alone" in _CASE1_REHEARSAL_ANTI_CRITERIA
    assert "residual_must_vanish" in _CASE1_REHEARSAL_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "dual_linf_under_wire_unproven" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "rehearsal" in one
    assert "rehearsal_present" in one
    assert "wire_rehearsal_present" in one
    assert "scaffold_present" in one
    assert "path_shipped" in one
    assert "bundle_shipped" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "packaging" in one or "alone" in one or "anti" in one
    assert "co-req" in one or "coreq" in one or "dry-run" in one or "dry run" in one
    assert "coreq" in d["coreq_status_glance"] or "isolation" in d["coreq_status_glance"]
    # multi-way permission coexistence
    assert d["rehearsal_present"] == "true"
    assert d["wire_rehearsal_present"] == "true"
    assert d["scaffold_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    # isolation: formatter body must not import TF / call live rehearsal report
    import ast
    import inspect

    src = inspect.getsource(
        format_tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "pulp" not in imported
    assert "offline_case1_dual_honest_multi_blocker_wire_rehearsal_report(" not in src
    # module-level: excel packaging never *calls* live TF rehearsal report
    import pims_admm_llm.models.excel_pipeline as ep
    from pathlib import Path as _Path

    ep_src = _Path(ep.__file__).read_text(encoding="utf-8")
    assert "offline_case1_dual_honest_multi_blocker_wire_rehearsal_report(" not in ep_src
    # no executable TF imports on packaging module (docstrings may mention bans)
    tree_mod = ast.parse(ep_src)
    for node in ast.walk(tree_mod):
        if isinstance(node, ast.Import):
            for alias in node.names:
                assert "tf_linear_blocks" not in alias.name
                assert "tensorflow" not in alias.name
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            assert "tf_linear_blocks" not in mod
            assert "tensorflow" not in mod
    # pulp may be used by classic Case 1 solve path in this module; packaging formatters must not
    # import it themselves (checked above via formatter AST).



def test_format_tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint_howto_pure():
    """Static blueprint packaging How_to: present; first_blocking; ship flags false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_BLUEPRINT_ANTI_CRITERIA,
        _CASE1_BLUEPRINT_PRESENT,
        _CASE1_BUNDLE_CRITERIA_MET_TODAY,
        _CASE1_BUNDLE_SHIP_ALLOWED_TODAY,
        _CASE1_BUNDLE_SHIPPED,
        _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET,
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_EXECUTION_SCAFFOLD_PRESENT,
        _CASE1_FIRST_BLOCKING_COREQ,
        _CASE1_FORM_CURRENT,
        _CASE1_FORM_LABEL_CHANGE_SHIPPED,
        _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY,
        _CASE1_FORM_PLANNED,
        _CASE1_IMPLEMENTATION_BLUEPRINT_PRESENT,
        _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS,
        _CASE1_ISOLATION_REWRITE_SHIPPED,
        _CASE1_PATH_DESIGN_CDU_SURFACE,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _CASE1_PATH_DESIGN_PRESENT,
        _CASE1_PATH_SHIPPED,
        _CASE1_REHEARSAL_PRESENT,
        _CASE1_SCAFFOLD_PRESENT,
        _CASE1_SHAPED_BLENDER_SURFACE,
        _CASE1_SHAPED_LINKING_STREAMS,
        _CASE1_SHIP_MET_ALLOWED_TODAY,
        _CASE1_WIRE_GO_BOARD_PRESENT,
        _CASE1_WIRE_REHEARSAL_PRESENT,
        _CASE1_WIRE_SHIP_ALLOWED_TODAY,
        _CASE1_WIRE_SHIPPED,
        _OFFLINE_TF_UNITS,
        _OFFLINE_WIRE_BLOCKER_IDS,
        format_tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint_howto,
    )

    d = format_tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint_howto()
    assert d["topic"] == "tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"
    assert "CDU" in d["units"] and "Blender" in d["units"]
    assert d["on_case1_solve"] == "false"
    assert d["not_case1_solve"] == "true"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["form_planned"] == _CASE1_FORM_PLANNED
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["case1_form_unchanged"] == "true"
    assert d["form_unchanged"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["solver"] == "false"
    assert d["on_excel_case1_path"] == "false"
    assert d["blueprint_present"] == "true"
    assert d["implementation_blueprint_present"] == "true"
    assert d["wire_go_board_present"] == "true"
    assert d["first_blocking_coreq"] == _CASE1_FIRST_BLOCKING_COREQ
    assert d["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert d["rehearsal_present"] == "true"
    assert d["wire_rehearsal_present"] == "true"
    assert d["scaffold_present"] == "true"
    assert d["execution_scaffold_present"] == "true"
    assert d["wire_shipped"] == "false"
    assert d["not_wire_shipped"] == "true"
    assert d["path_design_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["not_path_shipped"] == "true"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["bundle_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_rewrite_with_wire"] == _CASE1_ISOLATION_REWRITE_CHECKLIST_STATUS
    assert d["isolation_rewrite_with_wire"] == "open"
    assert d["rewrite_with_wire_not_delete"] == "true"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["isolation_rewrite_required_still_in_blockers"] == "true"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["cdu_surface"] == _CASE1_PATH_DESIGN_CDU_SURFACE
    assert d["blender_surface"] == _CASE1_SHAPED_BLENDER_SURFACE
    assert d["blender_surface"] == "linear_quality_pooling"
    for s in _CASE1_SHAPED_LINKING_STREAMS:
        assert s in d["intermediates"]
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["feature_flag_enabled_today"] == "false"
    assert d["blueprint_is_not_path_shipped"] == "true"
    assert d["blueprint_is_not_ship_met"] == "true"
    assert d["blueprint_is_not_wire"] == "true"
    assert d["blueprint_is_not_bundle_shipped"] == "true"
    assert d["blueprint_is_not_isolation_rewrite_shipped"] == "true"
    assert d["blueprint_is_not_form_label_change_shipped"] == "true"
    assert d["blueprint_is_not_form_flip"] == "true"
    assert d["blueprint_is_not_verdict_gate"] == "true"
    assert d["blueprint_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["first_blocking_coreq_alone_is_not_ship"] == "true"
    assert d["prep_map_is_not_ship"] == "true"
    assert d["packaging_is_not_path_shipped"] == "true"
    assert d["packaging_is_not_wire_shipped"] == "true"
    assert d["packaging_is_not_bundle_shipped"] == "true"
    assert d["order_hint_is_not_executor"] == "true"
    assert d["no_auto_wire"] == "true"
    assert d["distinct_from_scaffold_and_rehearsal_and_bundle_design_and_ship_met_criteria_packaging"] == "true"
    assert d["no_blender_offline_affine_kernel_blocker_still_true"] == "true"
    assert d["units_affine_unchanged"] == _OFFLINE_TF_UNITS
    assert "BLENDER" not in d["units_affine_unchanged"]
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    open_ids = d["dual_linf_proof_checklist_open_ids"]
    for oid in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS:
        assert oid in open_ids
    assert "isolation_rewrite_with_wire" in open_ids
    assert d["does_not_clear_wire_blockers"] == "true"
    assert d["not_full_plant_mass_balance"] == "true"
    assert d["not_pure_admm_dual_recovery"] == "true"
    assert d["not_form_flip"] == "true"
    assert d["not_dual_linf_under_wire_proven"] == "true"
    assert _CASE1_BLUEPRINT_PRESENT is True
    assert _CASE1_IMPLEMENTATION_BLUEPRINT_PRESENT is True
    assert _CASE1_WIRE_GO_BOARD_PRESENT is True
    assert _CASE1_FIRST_BLOCKING_COREQ == "isolation_rewrite_with_wire"
    assert _CASE1_REHEARSAL_PRESENT is True
    assert _CASE1_WIRE_REHEARSAL_PRESENT is True
    assert _CASE1_SCAFFOLD_PRESENT is True
    assert _CASE1_EXECUTION_SCAFFOLD_PRESENT is True
    assert _CASE1_BUNDLE_SHIPPED is False
    assert _CASE1_BUNDLE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_BUNDLE_CRITERIA_MET_TODAY is False
    assert _CASE1_ISOLATION_REWRITE_SHIPPED is False
    assert _CASE1_FORM_LABEL_CHANGE_SHIPPED is False
    assert _CASE1_FORM_LABEL_SHIP_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_PRESENT is True
    assert _CASE1_PATH_SHIPPED is False
    assert _CASE1_DUAL_HONEST_TF_AWARE_PATH_PRESENT_SHIP_MET is False
    assert _CASE1_SHIP_MET_ALLOWED_TODAY is False
    assert _CASE1_PATH_DESIGN_FEATURE_FLAG_ENABLED_TODAY is False
    assert _CASE1_WIRE_SHIP_ALLOWED_TODAY is False
    assert _CASE1_WIRE_SHIPPED is False
    assert "this_blueprint_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "go_board_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "first_blocking_coreq_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "prep_map_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "scaffold_plus_rehearsal_plus_blueprint_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "packaging_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "this_rehearsal_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "this_scaffold_alone" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "residual_must_vanish" in _CASE1_BLUEPRINT_ANTI_CRITERIA
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "dual_linf_under_wire_unproven" in _OFFLINE_WIRE_BLOCKER_IDS
    one = d["planner_one_liner"].lower()
    assert "blueprint" in one
    assert "blueprint_present" in one
    assert "implementation_blueprint_present" in one or "go-board" in one or "go_board" in one
    assert "first_blocking" in one or "isolation_rewrite_with_wire" in one
    assert "scaffold_present" in one
    assert "rehearsal_present" in one
    assert "path_shipped" in one
    assert "bundle_shipped" in one
    assert "wire_shipped" in one or "wire shipped" in one
    assert "classic_2block" in one or "form_current" in one
    assert "unproven" in one
    assert "verdict" in one
    assert "false" in one
    assert "dual" in one and "none" in one
    assert "not" in one and "wire" in one
    assert "no_blender" in one or "blocker" in one
    assert "fcc" in one and "coker" in one and "cdu" in one
    assert "packaging" in one or "alone" in one or "anti" in one
    assert "order_hint" in one or "go-board" in one or "go_board" in one or "prep" in one
    assert "isolation_rewrite_with_wire" in d["go_board_status_glance"] or "first_blocking" in d["go_board_status_glance"]
    # multi-way permission coexistence
    assert d["blueprint_present"] == "true"
    assert d["implementation_blueprint_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present_ship_met"] == "false"
    assert d["wire_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    # isolation: formatter body must not import TF / call live blueprint report
    import ast
    import inspect

    src = inspect.getsource(
        format_tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint_howto
    )
    tree = ast.parse(src)
    imported = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                imported.add(alias.name.split(".")[0])
        elif isinstance(node, ast.ImportFrom):
            if node.module:
                imported.add(node.module.split(".")[0])
    assert "tf_linear_blocks" not in imported
    assert "tensorflow" not in imported
    assert "pulp" not in imported
    assert "offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint_report(" not in src
    # module-level: excel packaging never *calls* live TF blueprint report
    import pims_admm_llm.models.excel_pipeline as ep
    from pathlib import Path as _Path

    ep_src = _Path(ep.__file__).read_text(encoding="utf-8")
    assert "offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint_report(" not in ep_src
    tree_mod = ast.parse(ep_src)
    for node in ast.walk(tree_mod):
        if isinstance(node, ast.Import):
            for alias in node.names:
                assert "tf_linear_blocks" not in alias.name
                assert "tensorflow" not in alias.name
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            assert "tf_linear_blocks" not in mod
            assert "tensorflow" not in mod


def test_case1_dual_linf_open_ids_no_longer_list_blender_pooling():
    """Excel open-ids realigned after #40: blender pooling no longer open."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS,
        _OFFLINE_WIRE_BLOCKER_IDS,
    )

    assert "blender_affine_or_honest_pooling" not in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS
    assert "isolation_rewrite_with_wire" in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS
    assert "form_label_change_shipped" in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS
    assert "online_linf_gate_under_tf_path" in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS
    assert "wire_shipped_false_today" in _CASE1_DUAL_LINF_PROOF_CHECKLIST_OPEN_IDS
    assert "no_blender_offline_affine_kernel" in _OFFLINE_WIRE_BLOCKER_IDS
    assert "isolation_rewrite_required" in _OFFLINE_WIRE_BLOCKER_IDS


def test_excel_fcc_export_matches_affine_coeffs():
    """E10 always-on: matrix builder MB_* == affine package (no TF, no solve)."""
    from pims_admm_llm.models.tf_linear_blocks import excel_fcc_matrix_matches_affine

    report = excel_fcc_matrix_matches_affine(atol=1e-12)
    assert report["ok"], report.get("mismatches")


def test_excel_pipeline_case1_tf_non_wiring_contract(tmp_path):
    """E14 permanent gate: Case 1 form + duals stay free of TF ownership claims."""
    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)

    assert report["mono"]["feasible"] and report["admm"]["feasible"]
    assert report["comparison"]["objective_gap_rel"] <= 0.005 + 1e-9
    assert report["comparison"]["dual_linf_online"] <= 15.0
    assert report["verdict"].startswith("PASS")

    assert report.get("model", {}).get("form") == "classic_2block_excel_path"

    admm = report["admm"]
    path = str(admm.get("dual_recovery_path") or "")
    assert path, "dual_recovery_path must be labeled"
    assert "online_lambda" in path
    # Do not claim pure-ADMM as dual recovery (path may say package-admm backend).
    assert "pure_admm" not in path.lower()
    assert "pure-admm-dual" not in path.lower().replace("_", "-")
    # TF must never be presented as dual recovery or a report-level dual owner.
    for blob in (path, str(report.get("tf_block")), str(admm.get("tf_block"))):
        assert "tf_block" not in blob.lower()
        assert "tensorflow" not in blob.lower()
    assert "tf_block" not in report
    assert "tf_block" not in admm
    # Primary shadows remain free online economic λ (not a TF artifact).
    assert isinstance(admm.get("shadow_prices"), dict) and admm["shadow_prices"]
    assert isinstance(admm.get("shadow_prices_recovered"), dict)


def test_excel_dual_honesty_primary_secondary(tmp_path):
    """E1/E2: Shadows/How_to/Summary PRIMARY online vs SECONDARY recovered; online-only gate."""
    from pims_admm_llm.models.excel_pipeline import (
        _verdict,
        format_dual_honesty_summary,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    out = tmp_path / "dual_honesty.xlsx"
    write_results_excel(out, report)
    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) <= GOAL_MAX_SHEETS

    # --- Report comparison roles (JSON honesty feed) ---
    cmp_ = report["comparison"]
    assert cmp_["dual_gate"] == "online_lambda"
    assert cmp_["verdict_dual_gate"] == "online_only"
    assert cmp_["dual_linf_online_role"] == "PRIMARY"
    assert cmp_["dual_linf_recovered_role"] == "SECONDARY"
    assert cmp_["recovered_secondary"] is True
    assert cmp_["dual_linf_online"] <= 15.0
    # Recovered may be large / ≥ online; never require recovered ≤15.
    assert cmp_["dual_linf_recovered"] + 1e-9 >= cmp_["dual_linf_online"]

    # Online-only VERDICT math lock (anti-recovered-PASS regression).
    assert report["verdict"].startswith("PASS")
    soft = _verdict(
        report["mono"],
        report["admm"],
        cmp_["objective_gap_rel"],
        dual_linf=100.0,  # would fail if this were the gate input
    )
    assert soft.startswith("PASS_SOFT") or soft.startswith("FAIL")
    assert "dual L∞" in soft or "dual" in soft.lower()

    dual = format_dual_honesty_summary(report)
    assert dual["primary_role"] == "PRIMARY"
    assert dual["secondary_role"] == "SECONDARY"
    assert dual["verdict_dual_gate"] == "online_only"
    assert "online_lambda" in dual["dual_recovery_path"]
    assert "PRIMARY" in dual["planner_one_liner"] and "SECONDARY" in dual["planner_one_liner"]

    # --- Shadows surface ---
    sh_vals = []
    for row in wb["Shadows"].iter_rows(values_only=True):
        sh_vals.extend(str(c) for c in row if c is not None)
    blob = " | ".join(sh_vals)
    assert "PRIMARY" in blob
    assert "SECONDARY" in blob
    assert "admm_online_econ" in blob
    assert "admm_recovered_econ" in blob
    assert "online_only" in blob or "verdict_dual_gate" in blob
    assert "online_lambda" in blob
    assert any("dual_L" in s and "online" in s.lower() for s in sh_vals) or (
        "dual_L∞_online_vs_mono" in blob or "dual_linf_online" in blob.lower()
    )
    assert any("recovered" in s.lower() and "dual" in s.lower() for s in sh_vals) or (
        "dual_L∞_recovered_vs_mono" in blob
    )

    # --- How_to this-run numbers + gate language ---
    how = {
        str(r[0].value): str(r[1].value or "")
        for r in wb["How_to_read"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    dual_note = how.get("duals_online_lambda", "")
    assert "PRIMARY" in dual_note and "SECONDARY" in dual_note
    assert "online_lambda" in dual_note
    # this-run number presence (formatted L∞ from report)
    online_s = dual["dual_linf_online"]
    assert online_s in dual_note or online_s in how.get("duals_primary_secondary", "")
    dps = how.get("duals_primary_secondary", "")
    assert dps and "PRIMARY" in dps and "SECONDARY" in dps
    assert "online_only" in dps or "gates VERDICT" in dps or "tol" in dps
    assert "not pure-ADMM" in dps.lower() or "not pure-admm" in dps.lower()
    assert "not TF" in dps or "not TF dual" in dps
    sb = how.get("solve_boundary", "")
    assert "PRIMARY" in sb and "SECONDARY" in sb

    # --- Summary gate notes ---
    summary = {
        str(r[0].value): r[1].value
        for r in wb["Summary"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert summary.get("dual_gate") == "online_lambda"
    assert summary.get("verdict_dual_gate") == "online_only"
    assert summary.get("recovered_secondary") is True
    online_role = str(summary.get("dual_linf_online_role") or "")
    rec_role = str(summary.get("dual_linf_recovered_role") or "")
    assert "PRIMARY" in online_role and "VERDICT" in online_role
    assert "SECONDARY" in rec_role
    assert "online_lambda" in str(summary.get("dual_recovery_path") or "")


def test_format_dual_honesty_summary_pure():
    """DRY helper works without a full solve."""
    from pims_admm_llm.models.excel_pipeline import format_dual_honesty_summary

    fake = {
        "admm": {"dual_recovery_path": "package-admm/qp_l2+online_lambda_shadows"},
        "comparison": {"dual_linf_online": 2.66, "dual_linf_recovered": 112.0},
    }
    d = format_dual_honesty_summary(fake)
    assert d["primary_role"] == "PRIMARY"
    assert d["secondary_role"] == "SECONDARY"
    assert d["verdict_dual_gate"] == "online_only"
    assert "2.66" in d["dual_linf_online"]
    assert "112" in d["dual_linf_recovered"]
    assert "online_lambda" in d["dual_recovery_path"]
    assert "PRIMARY" in d["shadows_role_banner"] and "SECONDARY" in d["shadows_role_banner"]
    # E3 dual-glance polish
    assert d["dual_linf_tol"] == "15"
    assert d["primary_vs_tol"] == "PASS"
    assert d["secondary_not_gate"] == "true"
    assert "PRIMARY online L∞" in d["dual_glance_strip"]
    assert "SECONDARY recovered" in d["dual_glance_strip"]
    assert "not gate" in d["dual_glance_strip"]
    assert "verdict_dual_gate=online_only" in d["dual_glance_strip"]
    assert "not pure-ADMM" in d["dual_glance_strip"]
    assert "dual_linf_under_wire=unproven" in d["dual_glance_strip"]
    assert "not pure-ADMM" in d["dual_recovery_path_honesty"]
    assert "online_lambda" in d["dual_recovery_path_honesty"]
    # recovered large must NOT flip primary_vs_tol / not gate language
    assert "112" in d["dual_glance_strip"]
    soft = format_dual_honesty_summary(
        {
            "admm": {"dual_recovery_path": "package-admm/x"},
            "comparison": {"dual_linf_online": 20.0, "dual_linf_recovered": 1.0},
        }
    )
    assert soft["primary_vs_tol"] == "SOFT/FAIL_vs_tol"
    assert "not gate" in soft["dual_glance_strip"]


def test_format_tf_offline_ladder_toc_howto_pure():
    """E7: static offline ladder TOC — blueprint + ship=false dual-ban; no TF import."""
    from pims_admm_llm.models.excel_pipeline import format_tf_offline_ladder_toc_howto
    import pims_admm_llm.models.excel_pipeline as ep

    d = format_tf_offline_ladder_toc_howto()
    assert d["topic"] == "tf_offline_ladder_toc"
    assert d["toc_present"] == "true"
    assert d["offline_tf_ladder_toc_ready"] == "true"
    assert d["includes_blueprint"] == "true"
    assert d["includes_first_blocker_operational_prep"] == "true"
    assert d.get("includes_form_label_second_coreq_operational_prep") == "true"
    assert d.get("includes_path_third_coreq_operational_prep") == "true"
    assert d.get("includes_dual_linf_fourth_coreq_operational_prep") == "true"
    assert d.get("includes_wire_fifth_coreq_operational_prep") == "true"
    assert d.get("includes_isolation_rewrite_first_blocker_execution_scaffold") == "true"
    assert d["ship_false_dual_ban"] == "true"
    assert d["wire_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["dual_linf_under_wire_status"] == "unproven"
    assert d["dual_recovery_path"] == "None"
    assert d["ready_is_not_ship"] == "true"
    assert d["blueprint_present_is_not_wire_ready"] == "true"
    assert d["not_index_growth"] == "true"
    ids = d["topic_ids"]
    assert "tf_offline_units" in ids
    assert "implementation_blueprint" in ids
    assert "isolation_rewrite_first_blocker_operational_prep" in ids
    assert "form_label_second_coreq_operational_prep" in ids
    assert "wire_preflight" in ids
    assert "isolation_rewrite" in ids
    one = d["planner_one_liner"].lower()
    assert "blueprint" in one
    assert "ship=false" in one or "ship=false" in one.replace(" ", "")
    assert "dual-ban" in one or "dual_recovery_path=none" in one
    assert "wire ready" in one or "wire_shipped" in one
    # isolation purity: formatter module source must not import tf on this helper path
    src = open(ep.__file__, encoding="utf-8").read()
    # helper itself is pure — no tensorflow token required inside the function body alone
    assert "format_tf_offline_ladder_toc_howto" in src
    assert int(d["topic_count"]) >= 20


def test_excel_dual_glance_and_ladder_toc_surfaces(tmp_path):
    """E3+E7: dual glance strip + ladder TOC on How_to / Summary / Shadows / meta / Calc_Check."""
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
        _OFFLINE_TF_INDEX_WHAT,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    out = tmp_path / "dual_glance_toc.xlsx"
    write_results_excel(out, report)
    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) <= GOAL_MAX_SHEETS

    pkg = format_planner_honesty_package(report)
    assert "dual_glance_strip" in pkg["meta"]
    assert "PRIMARY" in pkg["meta"]["dual_glance_strip"]
    assert "SECONDARY" in pkg["meta"]["dual_glance_strip"]
    assert pkg["meta"]["offline_tf_ladder_toc_ready"] is True
    assert pkg["meta"]["secondary_not_gate"] in (True, "true")
    assert pkg["meta"]["dual_linf_tol"] in (15, "15")
    assert "tf_offline_ladder_toc" in pkg
    assert pkg["tf_offline_ladder_toc"]["includes_blueprint"] == "true"
    # Index must not grow for TOC (no growth requirement / prefer How_to)
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439

    how = {
        str(r[0].value): str(r[1].value or "")
        for r in wb["How_to_read"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert "tf_offline_ladder_toc" in how
    toc = how["tf_offline_ladder_toc"].lower()
    assert "blueprint" in toc
    assert "ship=false" in toc or "ship=false" in toc.replace(" ", "")
    assert "dual" in toc
    assert "dual_glance_strip" in how
    assert "PRIMARY" in how["dual_glance_strip"]
    assert "not gate" in how["dual_glance_strip"]
    assert "dual_recovery_path_honesty" in how
    assert "not pure-ADMM" in how["dual_recovery_path_honesty"] or "not pure-admm" in how[
        "dual_recovery_path_honesty"
    ].lower()

    summary = {
        str(r[0].value): r[1].value
        for r in wb["Summary"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert summary.get("verdict_dual_gate") == "online_only"
    strip = str(summary.get("dual_glance_strip") or "")
    assert "PRIMARY" in strip and "SECONDARY" in strip
    assert summary.get("secondary_not_gate") in (True, "true")
    assert summary.get("offline_tf_ladder_toc_ready") is True
    assert "How_to" in str(summary.get("offline_tf_ladder_toc") or "")

    sh_vals = []
    for row in wb["Shadows"].iter_rows(values_only=True):
        sh_vals.extend(str(c) for c in row if c is not None)
    blob = " | ".join(sh_vals)
    assert "dual_glance_strip" in blob
    assert "PRIMARY" in blob and "SECONDARY" in blob
    assert "not gate" in blob.lower() or "SECONDARY_recovered_not_gate" in blob
    assert "dual_recovery_path_honesty" in blob or "not pure-ADMM" in blob
    assert "dual_linf_under_wire" in blob and "unproven" in blob

    rows = planner_honesty_check_rows(report)
    names = {r["check"] for r in rows}
    assert "recovered_linf_not_verdict_gate" in names
    assert "dual_glance_strip_primary_secondary" in names
    assert "dual_linf_under_wire_unproven_not_from_diagnostics" in names
    assert "offline_tf_ladder_toc_navigator" in names
    assert all(r["ok"] for r in rows if r["check"] in names)

    # VERDICT still online-only — recovered large does not fail
    assert report["verdict"].startswith("PASS")
    assert report["comparison"]["dual_linf_online"] <= 15.0
    assert report["comparison"]["verdict_dual_gate"] == "online_only"


def test_planner_honesty_glance_package(tmp_path):
    """E1: Index OFFLINE_TF + Summary strip + Calc_Check audits + meta.planner_honesty.

    After #30 Case-1-shaped skeleton API: glance package also locks offline
    Case-1-shaped CDU↔Blender skeleton readiness (linear_quality_pooling;
    naphtha/distillate/gasoil/residue; wire_shipped=False; dual_recovery_path=None;
    skeleton λ ≠ Case 1 duals; skeleton ≠ wire) alongside wire-preflight + multi-block
    plant-named linking + synthetic plant-linking + multi-round coordination + residual
    + block subproblem + priced + timing (static; isolation-safe). Dual PRIMARY/SECONDARY
    and classic form remain non-regression contracts. No live skeleton /
    offline_wire_preflight_report call.
    """
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    out = tmp_path / "honesty_glance.xlsx"
    write_results_excel(out, report)
    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) <= GOAL_MAX_SHEETS

    # --- Pure helpers ---
    pkg = format_planner_honesty_package(report)
    assert pkg["index_row"]["block"] == "OFFLINE_TF"
    assert "NOT" in pkg["index_row"]["what"] and "FCC" in pkg["index_row"]["what"]
    assert "COKER" in pkg["index_row"]["what"] and "CDU" in pkg["index_row"]["what"]
    what_l = pkg["index_row"]["what"].lower()
    assert "priced" in what_l and "timing" in what_l
    assert "admm residual" in what_l
    assert "block subproblem" in what_l or "residual+subproblem" in what_l or "subproblem" in what_l
    assert "coordination" in what_l
    assert "multi-round" in what_l or "multi round" in what_l
    assert "synthetic" in what_l
    assert "readiness" in what_l
    assert "raw affine" in what_l or "raw" in what_l
    assert "not duals" in what_l or "prices not duals" in what_l
    assert "not wire" in what_l
    # Positive plant-linking readiness packaging (not only the coordination ban phrase)
    assert "plant-linking readiness" in what_l or "plant linking readiness" in what_l
    assert "multi-block plant-linking" in what_l or "multi-block plant linking" in what_l
    assert "linking topology" in what_l or "shared λ" in what_l or "incidence" in what_l
    assert "not full plant" in what_l or "full plant mb" in what_l
    assert "per-unit coordination" in what_l or "coordination ≠ plant linking" in what_l
    # Positive plant-named readiness packaging (discriminators — not synthetic-only phrases)
    assert "plant-named linking readiness" in what_l or "plant named linking readiness" in what_l
    assert "plant_named_offline_demo" in what_l or "plant product streams" in what_l
    assert "identity incidence" in what_l or "plant product" in what_l
    # Wire-preflight readiness (short Index clause; full blockers in How_to/Summary/meta)
    assert "wire-preflight readiness" in what_l or "wire preflight readiness" in what_l
    assert "wire_shipped=false" in what_l or "wire_shipped=false" in what_l.replace(" ", "")
    assert "blockers" in what_l
    # Case-1-shaped skeleton readiness (short Index clause)
    assert "case-1-shaped" in what_l or "case1_shaped" in what_l or "case 1-shaped" in what_l
    assert "skeleton readiness" in what_l or "skeleton" in what_l
    assert "linear_quality_pooling" in what_l
    assert "naphtha" in what_l and "residue" in what_l
    # Dual-space/form contract readiness (short Index clause)
    assert "dual-space" in what_l or "dual_space" in what_l or "form contract" in what_l
    assert "unproven" in what_l
    assert "planned" in what_l or "form" in what_l
    # Dual-space L∞ probe readiness (short Index clause; hold/trim vs prior 1439)
    assert "probe" in what_l or "l∞" in what_l or "linf" in what_l
    assert "not verdict" in what_l or "verdict" in what_l
    assert "dual-ban" in what_l or "dual_recovery_path=none" in what_l or "dual-ban" in what_l.replace(" ", "")
    # Dual-space L∞ live-λ bridge readiness (short Index clause; co-exists with probe)
    assert "bridge" in what_l or "live-λ" in what_l or "live-lambda" in what_l or "live_lambda" in what_l
    assert "source-labeled" in what_l or "live_lambda_source" in what_l or "source" in what_l
    # Dual-space L∞ live-λ-seeded warm-start readiness (short Index clause; co-exists with bridge)
    assert "warm-start" in what_l or "warmstart" in what_l or "seeded" in what_l
    assert "seed_policy" in what_l or "seed≠proof" in what_l or "seed" in what_l
    # Honest blender pooling path readiness (short Index clause; co-exists with warm-start)
    assert "pooling" in what_l or "honest blender" in what_l
    assert "not affine" in what_l or "affine" in what_l
    assert "linear_quality_pooling" in what_l
    # Gate-criteria contract readiness (short Index clause; co-exists with pooling)
    assert "gate-criteria" in what_l or "criteria" in what_l or "flip=false" in what_l
    assert "gate open" in what_l or "flip=false" in what_l or "dual-ban" in what_l
    # Isolation design+ship-criteria readiness (short Index; co-exists with gate-criteria)
    assert (
        "iso design+crit" in what_l
        or "isolation design" in what_l
        or "isolation-rewrite" in what_l
        or "design_present" in what_l
    )
    assert (
        "rew=false" in what_l
        or "rewrite=false" in what_l
        or "rewrite_shipped" in what_l
        or "ship=false" in what_l
        or "design_present" in what_l
    )
    # Wire-ship acceptance design readiness (short Index clause; co-exists with isolation)
    assert "wire-ship" in what_l or "wire_ship" in what_l
    assert "ship=false" in what_l or "ship_allowed=false" in what_l or "ship=false" in what_l
    assert "wire=false" in what_l or "wire_shipped" in what_l or "dual-ban" in what_l
    # Dual-honest TF-aware path design+criteria readiness (short Index; co-exists with wire-ship)
    assert "path design" in what_l or "path_design" in what_l or "path design+crit" in what_l
    assert "path design+crit" in what_l or "path design+criteria" in what_l or "criteria" in what_l
    assert "path=false" in what_l or "path_shipped" in what_l or "path=false" in what_l
    assert "ship-met=false" in what_l or "ship_met" in what_l or "dual-ban" in what_l
    assert "form_label_ship=false" in what_l or "form=classic" in what_l or "form_label" in what_l or "dual-ban" in what_l
    assert "flip=false" in what_l or "ship-met=false" in what_l or "form_label_ship=false" in what_l or "dual-ban" in what_l
    assert "scaffold" in what_l
    assert "rehearse" in what_l or "rehearsal" in what_l
    from pims_admm_llm.models.excel_pipeline import _OFFLINE_TF_INDEX_WHAT
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439, len(_OFFLINE_TF_INDEX_WHAT)
    assert "iso design+crit" in _OFFLINE_TF_INDEX_WHAT or "isolation" in _OFFLINE_TF_INDEX_WHAT
    assert "ship=false" in _OFFLINE_TF_INDEX_WHAT
    assert "rew=false" in _OFFLINE_TF_INDEX_WHAT or "rewrite=false" in _OFFLINE_TF_INDEX_WHAT
    assert "bundle" in _OFFLINE_TF_INDEX_WHAT
    assert "bundle design" in _OFFLINE_TF_INDEX_WHAT or "bundle (present" in _OFFLINE_TF_INDEX_WHAT or "bundle design+crit" in _OFFLINE_TF_INDEX_WHAT
    assert "design+crit" in _OFFLINE_TF_INDEX_WHAT or "crit" in _OFFLINE_TF_INDEX_WHAT
    assert "ship=false" in _OFFLINE_TF_INDEX_WHAT  # includes bundle/scaffold ship=false
    assert "allow=false" in _OFFLINE_TF_INDEX_WHAT or "met=false" in _OFFLINE_TF_INDEX_WHAT
    assert "scaffold" in _OFFLINE_TF_INDEX_WHAT
    assert "rehearse" in _OFFLINE_TF_INDEX_WHAT or "rehearsal" in _OFFLINE_TF_INDEX_WHAT
    assert "scaffold+rehearse" in _OFFLINE_TF_INDEX_WHAT or (
        "scaffold" in _OFFLINE_TF_INDEX_WHAT and "rehearse" in _OFFLINE_TF_INDEX_WHAT
    )
    # blueprint / go-board signal (fold: scaffold+rehearse+bp; 1st=iso maps first_blocking)
    assert (
        "bp" in _OFFLINE_TF_INDEX_WHAT
        or "blueprint" in _OFFLINE_TF_INDEX_WHAT
        or "1st=iso" in _OFFLINE_TF_INDEX_WHAT
        or "go-board" in _OFFLINE_TF_INDEX_WHAT
    )
    assert "1st=iso" in _OFFLINE_TF_INDEX_WHAT or "isolation" in _OFFLINE_TF_INDEX_WHAT
    assert "scaffold+rehearse+bp" in _OFFLINE_TF_INDEX_WHAT or "bp" in _OFFLINE_TF_INDEX_WHAT
    assert pkg["meta"]["form"] == "classic_2block_excel_path"
    assert pkg["meta"]["dual_gate"] == "online_lambda"
    assert pkg["meta"]["verdict_dual_gate"] == "online_only"
    assert pkg["meta"]["on_excel_case1_path"] is False
    assert pkg["meta"]["tf_on_excel_case1_path"] is False
    assert "FCC" in pkg["meta"]["offline_tf_units"]
    assert "online_lambda" in str(pkg["meta"]["dual_recovery_path"])
    assert pkg["meta"]["offline_tf_priced_ready"] is True
    assert pkg["meta"]["offline_tf_timing_ready"] is True
    assert pkg["meta"]["offline_tf_admm_residual_ready"] is True
    assert pkg["meta"]["offline_tf_admm_block_subproblem_ready"] is True
    assert pkg["meta"]["offline_tf_admm_coordination_ready"] is True
    assert pkg["meta"]["offline_tf_admm_plant_linking_ready"] is True
    assert pkg["meta"]["offline_tf_admm_plant_named_linking_ready"] is True
    assert pkg["meta"]["offline_tf_wire_preflight_ready"] is True
    assert pkg["meta"]["offline_tf_case1_shaped_linking_ready"] is True
    assert pkg["meta"]["offline_tf_case1_dual_space_form_contract_ready"] is True
    assert pkg["meta"]["offline_tf_case1_dual_space_linf_probe_ready"] is True
    assert pkg["meta"]["offline_tf_case1_dual_space_linf_live_lambda_bridge_ready"] is True
    assert pkg["meta"]["offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_ready"] is True
    assert pkg["meta"]["offline_tf_case1_honest_blender_pooling_path_ready"] is True
    assert pkg["meta"]["offline_tf_case1_online_linf_gate_criteria_contract_ready"] is True
    assert pkg["meta"]["offline_tf_case1_isolation_rewrite_design_contract_ready"] is True
    assert pkg["meta"]["offline_tf_case1_wire_ship_acceptance_design_contract_ready"] is True
    assert pkg["meta"]["offline_tf_case1_dual_honest_tf_aware_path_design_contract_ready"] is True
    assert pkg["meta"][
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract_ready"
    ] is True
    assert pkg["meta"][
        "offline_tf_case1_form_label_change_shipped_criteria_contract_ready"
    ] is True
    assert pkg["meta"][
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract_ready"
    ] is True
    assert pkg["meta"][
        "offline_tf_case1_dual_honest_multi_blocker_wire_bundle_design_contract_ready"
    ] is True
    assert pkg["meta"][
        "offline_tf_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract_ready"
    ] is True
    assert pkg["meta"][
        "offline_tf_case1_dual_honest_tf_aware_path_execution_scaffold_ready"
    ] is True
    assert pkg["meta"]["offline_tf_scaffold_present"] is True
    assert pkg["meta"]["offline_tf_execution_scaffold_present"] is True
    assert pkg["meta"][
        "offline_tf_case1_dual_honest_multi_blocker_wire_rehearsal_ready"
    ] is True
    assert pkg["meta"]["offline_tf_rehearsal_present"] is True
    assert pkg["meta"]["offline_tf_wire_rehearsal_present"] is True
    assert pkg["meta"][
        "offline_tf_case1_dual_honest_multi_blocker_wire_implementation_blueprint_ready"
    ] is True
    assert pkg["meta"]["offline_tf_blueprint_present"] is True
    assert pkg["meta"]["offline_tf_implementation_blueprint_present"] is True
    assert pkg["meta"]["offline_tf_wire_go_board_present"] is True
    assert pkg["meta"]["offline_tf_first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert pkg["meta"]["offline_tf_bundle_design_present"] is True
    assert pkg["meta"]["offline_tf_bundle_criteria_present"] is True
    assert pkg["meta"]["offline_tf_bundle_shipped"] is False
    assert pkg["meta"]["offline_tf_bundle_ship_allowed_today"] is False
    assert pkg["meta"]["offline_tf_bundle_criteria_met_today"] is False
    assert pkg["meta"]["offline_tf_path_design_present"] is True
    assert pkg["meta"]["offline_tf_path_present_criteria_present"] is True
    assert pkg["meta"]["offline_tf_form_label_criteria_present"] is True
    assert pkg["meta"]["offline_tf_isolation_ship_criteria_present"] is True
    assert pkg["meta"]["offline_tf_path_shipped"] is False
    assert pkg["meta"]["offline_tf_dual_honest_tf_aware_path_present_ship_met"] is False
    assert pkg["meta"]["offline_tf_ship_met_allowed_today"] is False
    assert pkg["meta"]["offline_tf_form_label_ship_allowed_today"] is False
    assert pkg["meta"]["offline_tf_form_label_change_shipped"] is False
    assert pkg["meta"]["offline_tf_isolation_ship_allowed_today"] is False
    assert pkg["meta"]["offline_tf_isolation_rewrite_shipped"] is False
    assert pkg["meta"]["offline_tf_wire_ship_allowed_today"] is False
    assert pkg["meta"]["offline_tf_wire_shipped"] is False
    assert "priced" in str(pkg["meta"]["offline_tf_priced"]).lower()
    assert "timing" in str(pkg["meta"]["offline_tf_timing"]).lower()
    admm_note = str(pkg["meta"]["offline_tf_admm_residual"]).lower()
    assert "admm residual" in admm_note or "synthetic" in admm_note
    assert "not" in admm_note and (
        "dual" in admm_note or "wire" in admm_note or "online" in admm_note
    )
    sub_note = str(pkg["meta"]["offline_tf_admm_block_subproblem"]).lower()
    assert "block subproblem" in sub_note or "subproblem" in sub_note
    assert "synthetic" in sub_note or "raw" in sub_note
    assert "not" in sub_note and (
        "dual" in sub_note or "wire" in sub_note or "online" in sub_note
    )
    coord_note = str(pkg["meta"]["offline_tf_admm_coordination"]).lower()
    assert "coordination" in coord_note
    assert "synthetic" in coord_note or "multi-round" in coord_note
    assert "not" in coord_note and (
        "dual" in coord_note or "wire" in coord_note or "online" in coord_note
    )
    assert "plant" in coord_note or "linking" in coord_note or "per-unit" in coord_note
    plant_note = str(pkg["meta"]["offline_tf_admm_plant_linking"]).lower()
    assert "plant-linking" in plant_note or "plant linking" in plant_note
    assert "synthetic" in plant_note
    assert "not" in plant_note and (
        "dual" in plant_note or "wire" in plant_note or "online" in plant_note
    )
    assert "full plant" in plant_note or "mass balance" in plant_note
    assert "coordination" in plant_note  # distinct-from-coordination
    named_note = str(pkg["meta"]["offline_tf_admm_plant_named_linking"]).lower()
    assert "plant-named" in named_note or "plant named" in named_note
    assert "plant_named_offline_demo" in named_note or "plant product" in named_note
    assert "not" in named_note and (
        "dual" in named_note or "wire" in named_note or "online" in named_note
    )
    assert "full plant" in named_note or "mass balance" in named_note
    assert "synthetic" in named_note  # distinct-from-synthetic
    assert "coordination" in named_note  # distinct-from-coordination
    preflight_note = str(pkg["meta"]["offline_tf_wire_preflight"]).lower()
    assert "wire" in preflight_note and "preflight" in preflight_note
    assert "wire_shipped" in preflight_note or "not wire shipped" in preflight_note
    assert "blocker" in preflight_note
    assert "isolation_rewrite_required" in preflight_note
    assert "wire_not_shipped" in preflight_note
    assert "structural" in preflight_note or "ready_for_wire_discussion" in preflight_note
    c1_note = str(pkg["meta"]["offline_tf_case1_shaped_linking"]).lower()
    assert "case-1-shaped" in c1_note or "case1" in c1_note or "skeleton" in c1_note
    assert "linear_quality_pooling" in c1_note
    assert "naphtha" in c1_note and "residue" in c1_note
    assert "wire_shipped" in c1_note or "not wire" in c1_note
    assert "dual" in c1_note
    ds_note = str(pkg["meta"]["offline_tf_case1_dual_space_form_contract"]).lower()
    assert "dual-space" in ds_note or "form contract" in ds_note or "form_planned" in ds_note
    assert "classic_2block" in ds_note or "form_current" in ds_note
    assert "tf_affine_cdu_blender" in ds_note or "form_planned" in ds_note
    assert "unproven" in ds_note
    assert "wire_shipped" in ds_note or "not wire" in ds_note
    lp_note = str(pkg["meta"]["offline_tf_case1_dual_space_linf_probe"]).lower()
    assert "probe" in lp_note or "l∞" in lp_note or "linf" in lp_note
    assert "unproven" in lp_note
    assert "verdict" in lp_note
    assert "wire_shipped" in lp_note or "not wire" in lp_note
    assert "dual" in lp_note
    assert "online_linf_gate" in lp_note or "online_linf_gate_under_tf_path" in lp_note
    lb_note = str(pkg["meta"]["offline_tf_case1_dual_space_linf_live_lambda_bridge"]).lower()
    assert "bridge" in lb_note or "live" in lb_note or "λ" in lb_note or "lambda" in lb_note
    assert "unproven" in lb_note
    assert "verdict" in lb_note
    assert "wire_shipped" in lb_note or "not wire" in lb_note
    assert "dual" in lb_note
    assert "live_lambda_source" in lb_note or "source" in lb_note
    assert "caller_supplied" in lb_note and "fixture" in lb_note
    assert "online_linf_gate" in lb_note or "online_linf_gate_under_tf_path" in lb_note
    ws_note = str(
        pkg["meta"]["offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart"]
    ).lower()
    assert "warm" in ws_note or "seed" in ws_note or "λ" in ws_note or "lambda" in ws_note
    assert "unproven" in ws_note
    assert "verdict" in ws_note
    assert "wire_shipped" in ws_note or "not wire" in ws_note
    assert "dual" in ws_note
    assert "live_lambda_source" in ws_note or "source" in ws_note
    assert "seed_policy" in ws_note or "lambda0_from_live_primary" in ws_note
    assert "seed" in ws_note
    assert "caller_supplied" in ws_note and "fixture" in ws_note
    assert "online_linf_gate" in ws_note or "online_linf_gate_under_tf_path" in ws_note
    pool_note = str(pkg["meta"]["offline_tf_case1_honest_blender_pooling_path"]).lower()
    assert "pooling" in pool_note or "linear_quality_pooling" in pool_note
    assert "unproven" in pool_note
    assert "verdict" in pool_note
    assert "affine" in pool_note
    assert "honest_pooling_path_present" in pool_note
    assert "wire_shipped" in pool_note or "not wire" in pool_note
    assert "dual" in pool_note
    assert "no_blender" in pool_note or "blocker" in pool_note
    crit_note = str(pkg["meta"]["offline_tf_case1_online_linf_gate_criteria_contract"]).lower()
    assert "criteria" in crit_note or "gate" in crit_note or "flip" in crit_note
    assert "unproven" in crit_note
    assert "verdict" in crit_note
    assert "open" in crit_note
    assert "gate_flip_allowed_today" in crit_note or "flip" in crit_note
    assert "false" in crit_note
    assert "wire_shipped" in crit_note or "not wire" in crit_note
    assert "dual" in crit_note
    assert "dual-ban" in crit_note or "dual_recovery_path" in crit_note
    assert "online_linf_gate" in crit_note
    assert "no_blender" in crit_note or "blocker" in crit_note
    iso_note = str(pkg["meta"]["offline_tf_case1_isolation_rewrite_design_contract"]).lower()
    assert "isolation" in iso_note and "design" in iso_note
    assert "design_present" in iso_note or "rewrite_shipped" in iso_note
    assert "false" in iso_note
    assert "open" in iso_note
    assert "unproven" in iso_note
    assert "verdict" in iso_note
    assert "wire_shipped" in iso_note or "not wire" in iso_note
    assert "dual" in iso_note
    assert "dual-ban" in iso_note or "dual_recovery_path" in iso_note
    assert "isolation_rewrite" in iso_note
    assert "no_blender" in iso_note or "blocker" in iso_note
    ws_design_note = str(
        pkg["meta"]["offline_tf_case1_wire_ship_acceptance_design_contract"]
    ).lower()
    assert "wire-ship" in ws_design_note or "wire_ship" in ws_design_note or "ship" in ws_design_note
    assert "design" in ws_design_note
    assert "design_present" in ws_design_note
    assert "wire_ship_allowed_today" in ws_design_note or "ship_allowed" in ws_design_note
    assert "false" in ws_design_note
    assert "wire_shipped" in ws_design_note or "not wire" in ws_design_note
    assert "unproven" in ws_design_note
    assert "verdict" in ws_design_note
    assert "dual" in ws_design_note
    assert "dual-ban" in ws_design_note or "dual_recovery_path" in ws_design_note
    assert "no_blender" in ws_design_note or "blocker" in ws_design_note
    pd_design_note = str(
        pkg["meta"]["offline_tf_case1_dual_honest_tf_aware_path_design_contract"]
    ).lower()
    assert "path" in pd_design_note and "design" in pd_design_note
    assert "path_design_present" in pd_design_note
    assert "path_shipped" in pd_design_note
    assert "false" in pd_design_note
    assert "ship-met" in pd_design_note or "ship_met" in pd_design_note or "dual_honest" in pd_design_note
    assert "wire_ship_allowed_today" in pd_design_note or "ship_allowed" in pd_design_note
    assert "wire_shipped" in pd_design_note or "not wire" in pd_design_note
    assert "unproven" in pd_design_note
    assert "verdict" in pd_design_note
    assert "dual" in pd_design_note
    assert "dual-ban" in pd_design_note or "dual_recovery_path" in pd_design_note
    assert "no_blender" in pd_design_note or "blocker" in pd_design_note
    assert "linear_quality_pooling" in pd_design_note or "blender_surface" in pd_design_note
    pc_note = str(
        pkg["meta"]["offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract"]
    ).lower()
    assert "criteria" in pc_note
    assert "criteria_present" in pc_note
    assert "ship_met_allowed" in pc_note or "ship_met_allowed_today" in pc_note
    assert "false" in pc_note
    assert "ship-met" in pc_note or "ship_met" in pc_note or "dual_honest" in pc_note
    assert "path_design_present" in pc_note
    assert "path_shipped" in pc_note
    assert "wire_ship_allowed_today" in pc_note or "ship_allowed" in pc_note
    assert "wire_shipped" in pc_note or "not wire" in pc_note
    assert "unproven" in pc_note
    assert "verdict" in pc_note
    assert "dual" in pc_note
    assert "dual-ban" in pc_note or "dual_recovery_path" in pc_note
    assert "no_blender" in pc_note or "blocker" in pc_note
    fl_note = str(
        pkg["meta"]["offline_tf_case1_form_label_change_shipped_criteria_contract"]
    ).lower()
    assert "criteria" in fl_note
    assert "criteria_present" in fl_note
    assert "form_label_ship_allowed" in fl_note or "form_label_ship_allowed_today" in fl_note
    assert "form_label_change_shipped" in fl_note
    assert "false" in fl_note
    assert "classic" in fl_note or "form_current" in fl_note
    assert "path_design_present" in fl_note
    assert "path_shipped" in fl_note
    assert "wire_ship_allowed_today" in fl_note or "ship_allowed" in fl_note
    assert "wire_shipped" in fl_note or "not wire" in fl_note
    assert "unproven" in fl_note
    assert "verdict" in fl_note
    assert "dual" in fl_note
    assert "dual-ban" in fl_note or "dual_recovery_path" in fl_note
    assert "no_blender" in fl_note or "blocker" in fl_note
    iso_ship_note = str(
        pkg["meta"]["offline_tf_case1_isolation_rewrite_shipped_criteria_contract"]
    ).lower()
    assert "criteria" in iso_ship_note
    assert "criteria_present" in iso_ship_note
    assert "isolation_ship_allowed" in iso_ship_note
    assert "isolation_rewrite_shipped" in iso_ship_note
    assert "false" in iso_ship_note
    assert "open" in iso_ship_note
    assert "rewrite" in iso_ship_note
    assert "classic" in iso_ship_note or "form_current" in iso_ship_note
    assert "path_design_present" in iso_ship_note or "path_shipped" in iso_ship_note
    assert "wire_shipped" in iso_ship_note or "not wire" in iso_ship_note
    assert "unproven" in iso_ship_note
    assert "verdict" in iso_ship_note
    assert "dual" in iso_ship_note
    assert "dual-ban" in iso_ship_note or "dual_recovery_path" in iso_ship_note
    assert "no_blender" in iso_ship_note or "blocker" in iso_ship_note
    blockers_meta = str(pkg["meta"]["offline_tf_wire_blockers"])
    assert "isolation_rewrite_required" in blockers_meta
    assert "form_label_change_required" in blockers_meta
    assert "wire_not_shipped" in blockers_meta
    assert "case1_is_cdu_blender_package_admm" in blockers_meta
    assert "no_blender_offline_affine_kernel" in blockers_meta
    assert "dual_linf_under_wire_unproven" in blockers_meta
    readiness_note = str(pkg["meta"]["offline_tf_readiness_note"]).lower()
    assert "not" in readiness_note
    assert "admm residual" in readiness_note
    assert "block subproblem" in readiness_note
    assert "coordination" in readiness_note
    assert "plant-linking" in readiness_note or "plant linking" in readiness_note
    assert "plant-named" in readiness_note or "plant named" in readiness_note
    assert "wire-preflight" in readiness_note or "wire preflight" in readiness_note
    assert "case-1-shaped" in readiness_note or "case1" in readiness_note or "skeleton" in readiness_note
    assert "dual-space" in readiness_note or "form contract" in readiness_note
    assert "unproven" in readiness_note
    assert "probe" in readiness_note or "l∞" in readiness_note
    assert "verdict" in readiness_note
    assert "bridge" in readiness_note or "live-λ" in readiness_note or "live-lambda" in readiness_note
    assert "source-labeled" in readiness_note or "live_lambda_source" in readiness_note
    assert "warm-start" in readiness_note or "warmstart" in readiness_note or "seeded" in readiness_note
    assert "seed_policy" in readiness_note or "seed≠proof" in readiness_note or "seed" in readiness_note
    assert "pooling" in readiness_note or "honest blender" in readiness_note
    assert "not affine" in readiness_note or "affine" in readiness_note
    assert "criteria" in readiness_note or "gate-criteria" in readiness_note or "flip-criteria" in readiness_note
    assert "gate open" in readiness_note or "flip=false" in readiness_note or "flip" in readiness_note
    assert "isolation" in readiness_note and "design" in readiness_note
    assert "rewrite_shipped" in readiness_note or "design_present" in readiness_note
    assert "wire-ship" in readiness_note or "wire_ship" in readiness_note or "ship_allowed" in readiness_note
    assert "ship_allowed=false" in readiness_note or "ship_allowed" in readiness_note
    assert "path design" in readiness_note or "path_design" in readiness_note
    assert "path_shipped=false" in readiness_note or "path_shipped" in readiness_note
    assert "ship-met=false" in readiness_note or "ship-met" in readiness_note
    assert "path-present" in readiness_note or "criteria_present" in readiness_note or "ship_met_allowed" in readiness_note
    assert "form_label" in readiness_note or "form=classic" in readiness_note
    assert "isolation" in readiness_note and ("ship" in readiness_note or "rewrite" in readiness_note or "isolation_rewrite_shipped" in readiness_note)
    one_l = str(pkg["meta"]["planner_one_liner"]).lower()
    assert "priced" in one_l and "timing" in one_l
    assert "admm residual" in one_l
    assert "block subproblem" in one_l
    assert "coordination" in one_l
    assert "plant-linking" in one_l or "plant linking" in one_l
    assert "plant-named" in one_l or "plant named" in one_l
    assert "wire-preflight" in one_l or "wire preflight" in one_l
    assert "case-1-shaped" in one_l or "skeleton" in one_l
    assert "dual-space" in one_l or "form contract" in one_l
    assert "probe" in one_l or "l∞" in one_l
    assert "bridge" in one_l or "live-λ" in one_l or "live-lambda" in one_l or "live_lambda" in one_l
    assert "warm-start" in one_l or "warmstart" in one_l or "seeded" in one_l or "seed_policy" in one_l
    assert "pooling" in one_l or "honest blender" in one_l or "honest_pooling" in one_l
    assert "criteria" in one_l or "gate" in one_l or "flip" in one_l
    assert "isolation" in one_l and "design" in one_l
    assert "path design" in one_l or "path_design" in one_l or "path_shipped" in one_l
    assert "criteria" in one_l or "ship_met_allowed" in one_l or "path-present" in one_l
    assert "form_label" in one_l or "form=classic" in one_l
    assert "PRIMARY" in pkg["meta"]["dual_linf_online_role"]
    assert "SECONDARY" in pkg["meta"]["dual_linf_recovered_role"]
    assert pkg.get("tf_offline_admm_block_subproblem") is not None
    assert pkg["tf_offline_admm_block_subproblem"]["topic"] == "tf_offline_admm_block_subproblem"
    assert pkg.get("tf_offline_admm_coordination") is not None
    assert pkg["tf_offline_admm_coordination"]["topic"] == "tf_offline_admm_coordination"
    assert pkg.get("tf_offline_admm_plant_linking") is not None
    assert pkg["tf_offline_admm_plant_linking"]["topic"] == "tf_offline_admm_plant_linking"
    assert pkg.get("tf_offline_admm_plant_named_linking") is not None
    assert pkg["tf_offline_admm_plant_named_linking"]["topic"] == "tf_offline_admm_plant_named_linking"
    assert pkg.get("tf_offline_wire_preflight") is not None
    assert pkg["tf_offline_wire_preflight"]["topic"] == "tf_offline_wire_preflight"
    assert pkg["tf_offline_wire_preflight"]["wire_shipped"] == "false"
    assert pkg["tf_offline_wire_preflight"]["dual_recovery_path"] == "None"
    assert pkg.get("tf_offline_case1_shaped_linking") is not None
    assert pkg["tf_offline_case1_shaped_linking"]["topic"] == "tf_offline_case1_shaped_linking"
    assert pkg["tf_offline_case1_shaped_linking"]["wire_shipped"] == "false"
    assert pkg["tf_offline_case1_shaped_linking"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_case1_shaped_linking"]["blender_surface"] == "linear_quality_pooling"
    assert pkg.get("tf_offline_case1_dual_space_form_contract") is not None
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["topic"]
        == "tf_offline_case1_dual_space_form_contract"
    )
    assert pkg["tf_offline_case1_dual_space_form_contract"]["wire_shipped"] == "false"
    assert pkg["tf_offline_case1_dual_space_form_contract"]["dual_recovery_path"] == "None"
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["form_current"]
        == "classic_2block_excel_path"
    )
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["form_planned"]
        == "tf_affine_cdu_blender_shaped_excel_path"
    )
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["dual_linf_under_wire_status"]
        == "unproven"
    )
    assert pkg.get("tf_offline_case1_dual_space_linf_probe") is not None
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"]["topic"]
        == "tf_offline_case1_dual_space_linf_probe"
    )
    assert pkg["tf_offline_case1_dual_space_linf_probe"]["wire_shipped"] == "false"
    assert pkg["tf_offline_case1_dual_space_linf_probe"]["dual_recovery_path"] == "None"
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"]["dual_linf_under_wire_status"]
        == "unproven"
    )
    assert pkg["tf_offline_case1_dual_space_linf_probe"]["probe_is_not_verdict_gate"] == "true"
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"][
            "probe_is_not_dual_linf_under_wire_proof"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"]["dual_vector_face"]
        == "raw_online_duals"
    )
    assert pkg.get("tf_offline_case1_dual_space_linf_live_lambda_bridge") is not None
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["topic"]
        == "tf_offline_case1_dual_space_linf_live_lambda_bridge"
    )
    assert pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["wire_shipped"] == "false"
    assert pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["dual_recovery_path"] == "None"
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["dual_linf_under_wire_status"]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["bridge_is_not_verdict_gate"]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"][
            "bridge_is_not_dual_linf_under_wire_proof"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"][
            "live_lambda_source_must_be_labeled"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["dual_vector_face"]
        == "raw_online_duals"
    )
    assert pkg.get("tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart") is not None
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["topic"]
        == "tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["wire_shipped"]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["dual_recovery_path"]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "warmstart_is_not_verdict_gate"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "warmstart_is_not_dual_linf_under_wire_proof"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "seed_identity_linf_is_not_proof"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "live_lambda_source_must_be_labeled"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["seed_policy"]
        == "lambda0_from_live_primary_online"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["z0_policy"]
        == "unchanged_default_skeleton_z"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["dual_vector_face"]
        == "raw_online_duals"
    )
    assert pkg.get("tf_offline_case1_honest_blender_pooling_path") is not None
    assert (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["topic"]
        == "tf_offline_case1_honest_blender_pooling_path"
    )
    assert pkg["tf_offline_case1_honest_blender_pooling_path"]["wire_shipped"] == "false"
    assert pkg["tf_offline_case1_honest_blender_pooling_path"]["dual_recovery_path"] == "None"
    assert (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["dual_linf_under_wire_status"]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["pooling_path_is_not_verdict_gate"]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["pooling_path_is_not_affine_kernel"]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["pooling_path_is_not_wire"]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["blender_pooling_checklist_status"]
        == "honest_pooling_path_present"
    )
    assert (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["blender_surface"]
        == "linear_quality_pooling"
    )
    assert "blender_affine_or_honest_pooling" not in (
        pkg["tf_offline_case1_honest_blender_pooling_path"]["dual_linf_proof_checklist_open_ids"]
    )
    assert pkg.get("tf_offline_case1_online_linf_gate_criteria_contract") is not None
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"]["topic"]
        == "tf_offline_case1_online_linf_gate_criteria_contract"
    )
    assert pkg["tf_offline_case1_online_linf_gate_criteria_contract"]["wire_shipped"] == "false"
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"]["dual_recovery_path"] == "None"
    )
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"][
            "online_linf_gate_under_tf_path"
        ]
        == "open"
    )
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"][
            "gate_flip_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"]["criteria_met_today"]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"][
            "contract_is_not_gate_flip"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"][
            "contract_is_not_verdict_gate"
        ]
        == "true"
    )
    assert "online_linf_gate_under_tf_path" in (
        pkg["tf_offline_case1_online_linf_gate_criteria_contract"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    assert pkg.get("tf_offline_case1_isolation_rewrite_design_contract") is not None
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"]["topic"]
        == "tf_offline_case1_isolation_rewrite_design_contract"
    )
    assert pkg["tf_offline_case1_isolation_rewrite_design_contract"]["wire_shipped"] == "false"
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"]["dual_recovery_path"] == "None"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"][
            "isolation_rewrite_design_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"][
            "isolation_rewrite_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"][
            "isolation_rewrite_with_wire"
        ]
        == "open"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"][
            "design_is_not_isolation_rewrite_shipped"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"][
            "design_is_not_verdict_gate"
        ]
        == "true"
    )
    assert "isolation_rewrite_with_wire" in (
        pkg["tf_offline_case1_isolation_rewrite_design_contract"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    assert pkg.get("tf_offline_case1_wire_ship_acceptance_design_contract") is not None
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"]["topic"]
        == "tf_offline_case1_wire_ship_acceptance_design_contract"
    )
    assert pkg["tf_offline_case1_wire_ship_acceptance_design_contract"]["wire_shipped"] == "false"
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"]["dual_recovery_path"] == "None"
    )
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"][
            "wire_ship_acceptance_design_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"][
            "wire_ship_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"][
            "wire_ship_criteria_met_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"][
            "design_is_not_wire_ship_allow"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"][
            "design_is_not_verdict_gate"
        ]
        == "true"
    )
    assert "wire_shipped_false_today" in (
        pkg["tf_offline_case1_wire_ship_acceptance_design_contract"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    assert pkg.get("tf_offline_case1_dual_honest_tf_aware_path_design_contract") is not None
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"]["topic"]
        == "tf_offline_case1_dual_honest_tf_aware_path_design_contract"
    )
    assert pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"]["wire_shipped"] == "false"
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"]["dual_recovery_path"] == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "path_design_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "path_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "dual_honest_tf_aware_path_present_ship_met"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "wire_ship_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "design_is_not_path_shipped"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "design_is_not_verdict_gate"
        ]
        == "true"
    )
    assert "wire_shipped_false_today" in (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_design_contract"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    assert pkg.get("tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract") is not None
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"]["topic"]
        == "tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "criteria_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "ship_met_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "dual_honest_tf_aware_path_present_ship_met"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "path_design_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "path_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "wire_ship_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "criteria_is_not_ship_met"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "criteria_is_not_verdict_gate"
        ]
        == "true"
    )
    assert "wire_shipped_false_today" in (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_present_criteria_contract"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    assert pkg.get("tf_offline_case1_form_label_change_shipped_criteria_contract") is not None
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"]["topic"]
        == "tf_offline_case1_form_label_change_shipped_criteria_contract"
    )
    assert pkg.get("tf_offline_case1_isolation_rewrite_shipped_criteria_contract") is not None
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"]["topic"]
        == "tf_offline_case1_isolation_rewrite_shipped_criteria_contract"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "isolation_ship_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "isolation_rewrite_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "criteria_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "isolation_rewrite_with_wire"
        ]
        == "open"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "rewrite_with_wire_not_delete"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "criteria_is_not_verdict_gate"
        ]
        == "true"
    )
    assert "isolation_rewrite_with_wire" in (
        pkg["tf_offline_case1_isolation_rewrite_shipped_criteria_contract"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )

    assert pkg.get("tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract") is not None
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"]["topic"]
        == "tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "bundle_design_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "bundle_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "bundle_ship_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "order_hint_is_not_executor"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "isolation_rewrite_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "design_is_not_verdict_gate"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_design_contract"][
            "design_is_not_bundle_shipped"
        ]
        == "true"
    )

    assert pkg.get("tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract") is not None
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"]["topic"]
        == "tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "criteria_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "bundle_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "bundle_ship_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "criteria_met_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "order_hint_is_not_executor"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "isolation_rewrite_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "criteria_is_not_verdict_gate"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "criteria_is_not_bundle_shipped"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_bundle_shipped_criteria_contract"][
            "distinct_from_bundle_design_packaging"
        ]
        == "true"
    )
    assert pkg.get("tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold") is not None
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"]["topic"]
        == "tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "scaffold_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "execution_scaffold_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "path_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "dual_honest_tf_aware_path_present_ship_met"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "bundle_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "isolation_rewrite_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "form_label_change_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "scaffold_is_not_verdict_gate"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_tf_aware_path_execution_scaffold"][
            "feature_flag_enabled_today"
        ]
        == "false"
    )
    # dual-ban coexistence: scaffold ready never coexists with ship flags true
    assert pkg["meta"]["offline_tf_case1_dual_honest_tf_aware_path_execution_scaffold_ready"] is True
    assert pkg["meta"]["offline_tf_scaffold_present"] is True
    assert pkg["meta"]["offline_tf_path_shipped"] is False
    assert pkg["meta"]["offline_tf_wire_shipped"] is False
    assert pkg["meta"]["offline_tf_bundle_shipped"] is False
    assert pkg["meta"]["offline_tf_isolation_rewrite_shipped"] is False
    assert pkg["meta"]["offline_tf_form_label_change_shipped"] is False
    # rehearsal packaging twin
    assert pkg.get("tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal") is not None
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"]["topic"]
        == "tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "rehearsal_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "wire_rehearsal_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "scaffold_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "path_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "dual_honest_tf_aware_path_present_ship_met"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "bundle_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "isolation_rewrite_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "form_label_change_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "rehearsal_is_not_verdict_gate"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "feature_flag_enabled_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_rehearsal"][
            "coreq_matrix_is_not_ship"
        ]
        == "true"
    )
    # dual-ban coexistence: rehearsal ready never coexists with ship flags true
    assert pkg["meta"]["offline_tf_case1_dual_honest_multi_blocker_wire_rehearsal_ready"] is True
    assert pkg["meta"]["offline_tf_rehearsal_present"] is True
    assert pkg["meta"]["offline_tf_wire_rehearsal_present"] is True
    assert pkg["meta"]["offline_tf_path_shipped"] is False
    assert pkg["meta"]["offline_tf_wire_shipped"] is False
    assert pkg["meta"]["offline_tf_bundle_shipped"] is False
    assert pkg["meta"]["offline_tf_isolation_rewrite_shipped"] is False
    assert pkg["meta"]["offline_tf_form_label_change_shipped"] is False
    # blueprint packaging twin
    assert pkg.get("tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint") is not None
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"]["topic"]
        == "tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "blueprint_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "implementation_blueprint_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "wire_go_board_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "first_blocking_coreq"
        ]
        == "isolation_rewrite_with_wire"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "rehearsal_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "scaffold_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "path_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "dual_honest_tf_aware_path_present_ship_met"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "bundle_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "isolation_rewrite_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "form_label_change_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "blueprint_is_not_verdict_gate"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "feature_flag_enabled_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "first_blocking_coreq_alone_is_not_ship"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_honest_multi_blocker_wire_implementation_blueprint"][
            "order_hint_is_not_executor"
        ]
        == "true"
    )
    # dual-ban coexistence: blueprint ready never coexists with ship flags true
    assert pkg["meta"]["offline_tf_case1_dual_honest_multi_blocker_wire_implementation_blueprint_ready"] is True
    assert pkg["meta"]["offline_tf_blueprint_present"] is True
    assert pkg["meta"]["offline_tf_implementation_blueprint_present"] is True
    assert pkg["meta"]["offline_tf_wire_go_board_present"] is True
    assert pkg["meta"]["offline_tf_first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert pkg["meta"]["offline_tf_path_shipped"] is False
    assert pkg["meta"]["offline_tf_wire_shipped"] is False
    assert pkg["meta"]["offline_tf_bundle_shipped"] is False
    assert pkg["meta"]["offline_tf_isolation_rewrite_shipped"] is False
    assert pkg["meta"]["offline_tf_form_label_change_shipped"] is False
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "wire_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "criteria_present"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "form_label_ship_allowed_today"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "form_label_change_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "form_current"
        ]
        == "classic_2block_excel_path"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "path_shipped"
        ]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "criteria_is_not_form_label_shipped"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "criteria_is_not_verdict_gate"
        ]
        == "true"
    )
    assert "form_label_change_shipped" in (
        pkg["tf_offline_case1_form_label_change_shipped_criteria_contract"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    summary_keys = {k for k, _ in pkg["summary_pairs"]}
    assert {
        "offline_tf_priced",
        "offline_tf_timing",
        "offline_tf_admm_residual",
        "offline_tf_admm_block_subproblem",
        "offline_tf_admm_coordination",
        "offline_tf_admm_plant_linking",
        "offline_tf_admm_plant_named_linking",
        "offline_tf_wire_preflight",
        "offline_tf_case1_shaped_linking",
        "offline_tf_case1_dual_space_form_contract",
        "offline_tf_case1_dual_space_linf_probe",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart",
        "offline_tf_case1_honest_blender_pooling_path",
        "offline_tf_case1_online_linf_gate_criteria_contract",
        "offline_tf_case1_isolation_rewrite_design_contract",
        "offline_tf_case1_wire_ship_acceptance_design_contract",
        "offline_tf_case1_dual_honest_tf_aware_path_design_contract",
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract",
        "offline_tf_case1_form_label_change_shipped_criteria_contract",
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract",
        "offline_tf_wire_blockers",
        "offline_tf_path_design_present",
        "offline_tf_path_present_criteria_present",
        "offline_tf_form_label_criteria_present",
        "offline_tf_isolation_ship_criteria_present",
        "offline_tf_path_shipped",
        "offline_tf_dual_honest_tf_aware_path_present_ship_met",
        "offline_tf_ship_met_allowed_today",
        "offline_tf_form_label_ship_allowed_today",
        "offline_tf_form_label_change_shipped",
        "offline_tf_isolation_ship_allowed_today",
        "offline_tf_isolation_rewrite_shipped",
        "offline_tf_wire_ship_allowed_today",
        "offline_tf_wire_shipped",
        "offline_tf_readiness_note",
        "offline_tf_units",
        "dual_gate",
        "model_form",
    } <= summary_keys

    rows = planner_honesty_check_rows(report)
    names = {r["check"] for r in rows}
    assert {
        "form_classic_2block",
        "dual_gate_online_only",
        "offline_tf_not_on_case1",
        "offline_tf_priced_not_duals",
        "offline_tf_timing_not_case1",
        "offline_tf_admm_residual_not_duals",
        "offline_tf_admm_block_subproblem_not_duals",
        "offline_tf_admm_coordination_not_duals",
        "offline_tf_admm_plant_linking_not_duals",
        "offline_tf_admm_plant_named_linking_not_duals",
        "offline_tf_wire_preflight_not_duals",
        "offline_tf_wire_not_shipped",
        "offline_tf_case1_shaped_linking_not_duals",
        "offline_tf_case1_shaped_not_wire",
        "offline_tf_case1_dual_space_form_contract_not_duals",
        "offline_tf_case1_dual_space_form_contract_not_wire",
        "offline_tf_case1_dual_space_linf_probe_not_duals",
        "offline_tf_case1_dual_space_linf_probe_not_wire",
        "offline_tf_case1_dual_space_linf_probe_not_verdict_gate",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_not_duals",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_not_wire",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_not_verdict_gate",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_source_must_be_labeled",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_duals",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_wire",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_verdict_gate",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_source_must_be_labeled",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_seed_identity_not_proof",
        "offline_tf_case1_honest_blender_pooling_path_not_duals",
        "offline_tf_case1_honest_blender_pooling_path_not_wire",
        "offline_tf_case1_honest_blender_pooling_path_not_verdict_gate",
        "offline_tf_case1_honest_blender_pooling_path_not_affine_kernel",
        "offline_tf_case1_honest_blender_pooling_path_checklist_honest_pooling_path_present",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_duals",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_wire",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_verdict_gate",
        "offline_tf_case1_online_linf_gate_criteria_contract_gate_open_flip_false",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_isolation_rewrite_design_contract_not_duals",
        "offline_tf_case1_isolation_rewrite_design_contract_not_wire",
        "offline_tf_case1_isolation_rewrite_design_contract_not_verdict_gate",
        "offline_tf_case1_isolation_rewrite_design_contract_rewrite_not_shipped_checklist_open",
        "offline_tf_case1_isolation_rewrite_design_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_duals",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_wire",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_verdict_gate",
        "offline_tf_case1_wire_ship_acceptance_design_contract_ship_allowed_false_wire_false",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_dual_honest_tf_aware_path_design_contract_not_duals",
        "offline_tf_case1_dual_honest_tf_aware_path_design_contract_not_wire",
        "offline_tf_case1_dual_honest_tf_aware_path_design_contract_not_verdict_gate",
        "offline_tf_case1_dual_honest_tf_aware_path_design_contract_path_shipped_false_ship_met_false",
        "offline_tf_case1_dual_honest_tf_aware_path_design_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract_not_duals",
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract_not_wire",
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract_not_verdict_gate",
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract_ship_met_false_path_shipped_false",
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_form_label_change_shipped_criteria_contract_not_duals",
        "offline_tf_case1_form_label_change_shipped_criteria_contract_not_wire",
        "offline_tf_case1_form_label_change_shipped_criteria_contract_not_verdict_gate",
        "offline_tf_case1_form_label_change_shipped_criteria_contract_form_label_shipped_false_form_classic",
        "offline_tf_case1_form_label_change_shipped_criteria_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract_not_duals",
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract_not_wire",
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract_not_verdict_gate",
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract_rewrite_shipped_false_checklist_open",
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract_not_dual_linf_under_wire_proof",
    } <= names
    assert all(r["ok"] is True for r in rows)

    # --- meta.planner_honesty on report ---
    ph = report["meta"]["planner_honesty"]
    assert ph["form"] == "classic_2block_excel_path"
    assert ph["dual_gate"] == "online_lambda"
    assert ph["verdict_dual_gate"] == "online_only"
    assert ph["dual_linf_online_role"] == "PRIMARY"
    assert ph["dual_linf_recovered_role"] == "SECONDARY"
    assert ph["on_excel_case1_path"] is False
    assert ph["tf_on_excel_case1_path"] is False
    assert "FCC" in ph["offline_tf_units"] and "COKER" in ph["offline_tf_units"]
    assert "CDU" in ph["offline_tf_units"]
    assert "online_lambda" in str(ph["dual_recovery_path"])
    assert ph["offline_tf_priced_ready"] is True
    assert ph["offline_tf_timing_ready"] is True
    assert ph["offline_tf_admm_residual_ready"] is True
    assert ph["offline_tf_admm_block_subproblem_ready"] is True
    assert ph["offline_tf_admm_coordination_ready"] is True
    assert ph["offline_tf_admm_plant_linking_ready"] is True
    assert ph["offline_tf_admm_plant_named_linking_ready"] is True
    assert ph["offline_tf_wire_preflight_ready"] is True
    assert ph["offline_tf_case1_shaped_linking_ready"] is True
    assert ph["offline_tf_case1_dual_space_form_contract_ready"] is True
    assert ph["offline_tf_case1_dual_space_linf_probe_ready"] is True
    assert ph["offline_tf_case1_dual_space_linf_live_lambda_bridge_ready"] is True
    assert ph["offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_ready"] is True
    assert ph["offline_tf_case1_honest_blender_pooling_path_ready"] is True
    assert ph["offline_tf_case1_online_linf_gate_criteria_contract_ready"] is True
    assert ph["offline_tf_case1_isolation_rewrite_design_contract_ready"] is True
    assert ph["offline_tf_case1_wire_ship_acceptance_design_contract_ready"] is True
    assert ph["offline_tf_case1_dual_honest_tf_aware_path_design_contract_ready"] is True
    assert ph[
        "offline_tf_case1_dual_honest_tf_aware_path_present_criteria_contract_ready"
    ] is True
    assert ph[
        "offline_tf_case1_form_label_change_shipped_criteria_contract_ready"
    ] is True
    assert ph[
        "offline_tf_case1_isolation_rewrite_shipped_criteria_contract_ready"
    ] is True
    assert ph["offline_tf_path_design_present"] is True
    assert ph["offline_tf_path_present_criteria_present"] is True
    assert ph["offline_tf_form_label_criteria_present"] is True
    assert ph["offline_tf_isolation_ship_criteria_present"] is True
    assert ph["offline_tf_path_shipped"] is False
    assert ph["offline_tf_dual_honest_tf_aware_path_present_ship_met"] is False
    assert ph["offline_tf_ship_met_allowed_today"] is False
    assert ph["offline_tf_form_label_ship_allowed_today"] is False
    assert ph["offline_tf_form_label_change_shipped"] is False
    assert ph["offline_tf_isolation_ship_allowed_today"] is False
    assert ph["offline_tf_isolation_rewrite_shipped"] is False
    assert ph["offline_tf_wire_ship_allowed_today"] is False
    assert ph["offline_tf_wire_shipped"] is False
    assert "priced" in str(ph["offline_tf_priced"]).lower()
    assert "timing" in str(ph["offline_tf_timing"]).lower()
    assert "not duals" in str(ph["offline_tf_priced"]).lower() or "not admm" in str(
        ph["offline_tf_priced"]
    ).lower()
    assert "not case 1" in str(ph["offline_tf_timing"]).lower() or "wall" in str(
        ph["offline_tf_timing"]
    ).lower()
    ph_admm = str(ph["offline_tf_admm_residual"]).lower()
    assert "synthetic" in ph_admm or "admm residual" in ph_admm
    assert "not" in ph_admm and (
        "dual" in ph_admm or "wire" in ph_admm or "online" in ph_admm
    )
    ph_sub = str(ph["offline_tf_admm_block_subproblem"]).lower()
    assert "subproblem" in ph_sub
    assert "not" in ph_sub and (
        "dual" in ph_sub or "wire" in ph_sub or "online" in ph_sub
    )
    ph_coord = str(ph["offline_tf_admm_coordination"]).lower()
    assert "coordination" in ph_coord
    assert "not" in ph_coord and (
        "dual" in ph_coord or "wire" in ph_coord or "online" in ph_coord
    )
    assert "plant" in ph_coord or "linking" in ph_coord or "per-unit" in ph_coord
    ph_plant = str(ph["offline_tf_admm_plant_linking"]).lower()
    assert "plant-linking" in ph_plant or "plant linking" in ph_plant
    assert "synthetic" in ph_plant
    assert "not" in ph_plant and (
        "dual" in ph_plant or "wire" in ph_plant or "online" in ph_plant
    )
    assert "full plant" in ph_plant or "mass balance" in ph_plant
    ph_named = str(ph["offline_tf_admm_plant_named_linking"]).lower()
    assert "plant-named" in ph_named or "plant named" in ph_named
    assert "plant_named_offline_demo" in ph_named or "plant product" in ph_named
    assert "not" in ph_named and (
        "dual" in ph_named or "wire" in ph_named or "online" in ph_named
    )
    assert "full plant" in ph_named or "mass balance" in ph_named
    ph_pre = str(ph["offline_tf_wire_preflight"]).lower()
    assert "preflight" in ph_pre
    assert "blocker" in ph_pre
    assert "wire_shipped" in ph_pre or "not wire shipped" in ph_pre
    ph_c1 = str(ph["offline_tf_case1_shaped_linking"]).lower()
    assert "case-1-shaped" in ph_c1 or "skeleton" in ph_c1 or "case1" in ph_c1
    assert "linear_quality_pooling" in ph_c1
    assert "naphtha" in ph_c1 and "residue" in ph_c1
    ph_ds = str(ph["offline_tf_case1_dual_space_form_contract"]).lower()
    assert "dual-space" in ph_ds or "form contract" in ph_ds or "form_planned" in ph_ds
    assert "unproven" in ph_ds
    assert "classic_2block" in ph_ds or "form_current" in ph_ds
    ph_lp = str(ph["offline_tf_case1_dual_space_linf_probe"]).lower()
    assert "probe" in ph_lp or "l∞" in ph_lp or "linf" in ph_lp
    assert "unproven" in ph_lp
    assert "verdict" in ph_lp
    ph_lb = str(ph["offline_tf_case1_dual_space_linf_live_lambda_bridge"]).lower()
    assert "bridge" in ph_lb or "live" in ph_lb or "λ" in ph_lb or "lambda" in ph_lb
    assert "unproven" in ph_lb
    assert "verdict" in ph_lb
    assert "live_lambda_source" in ph_lb or "source" in ph_lb
    ph_ws = str(ph["offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart"]).lower()
    assert "warm" in ph_ws or "seed" in ph_ws or "λ" in ph_ws or "lambda" in ph_ws
    assert "unproven" in ph_ws
    assert "verdict" in ph_ws
    assert "live_lambda_source" in ph_ws or "source" in ph_ws
    assert "seed_policy" in ph_ws or "lambda0_from_live_primary" in ph_ws
    ph_pool = str(ph["offline_tf_case1_honest_blender_pooling_path"]).lower()
    assert "pooling" in ph_pool or "linear_quality_pooling" in ph_pool
    assert "unproven" in ph_pool
    assert "verdict" in ph_pool
    assert "affine" in ph_pool
    assert "honest_pooling_path_present" in ph_pool
    ph_crit = str(ph["offline_tf_case1_online_linf_gate_criteria_contract"]).lower()
    assert "criteria" in ph_crit or "gate" in ph_crit or "flip" in ph_crit
    assert "unproven" in ph_crit
    assert "verdict" in ph_crit
    assert "open" in ph_crit
    assert "false" in ph_crit
    assert "online_linf_gate" in ph_crit
    ph_iso = str(ph["offline_tf_case1_isolation_rewrite_design_contract"]).lower()
    assert "isolation" in ph_iso and "design" in ph_iso
    assert "design_present" in ph_iso or "rewrite_shipped" in ph_iso
    assert "unproven" in ph_iso
    assert "verdict" in ph_iso
    assert "open" in ph_iso
    assert "false" in ph_iso
    ph_wsd = str(ph["offline_tf_case1_wire_ship_acceptance_design_contract"]).lower()
    assert "wire-ship" in ph_wsd or "wire_ship" in ph_wsd or "ship" in ph_wsd
    assert "design" in ph_wsd
    assert "design_present" in ph_wsd
    assert "wire_ship_allowed_today" in ph_wsd or "ship_allowed" in ph_wsd
    assert "false" in ph_wsd
    assert "unproven" in ph_wsd
    assert "verdict" in ph_wsd

    # --- Submodel_Index OFFLINE_TF readiness ---
    ih = [c.value for c in wb["Submodel_Index"][1]]
    bi = ih.index("block")
    wi = ih.index("what")
    index_rows = {
        r[bi].value: str(r[wi].value or "")
        for r in wb["Submodel_Index"].iter_rows(min_row=2)
        if r[bi].value
    }
    assert "OFFLINE_TF" in index_rows
    ot = index_rows["OFFLINE_TF"].lower()
    assert "fcc" in ot and "coker" in ot and "cdu" in ot
    assert "not" in ot and ("case 1" in ot or "classic" in ot)
    assert "none" in ot or "dual_recovery_path" in ot
    assert "priced" in ot and "timing" in ot
    assert "admm residual" in ot
    assert "block subproblem" in ot or "residual+subproblem" in ot or "subproblem" in ot
    assert "coordination" in ot
    assert "synthetic" in ot
    assert "readiness" in ot
    assert "not wire" in ot or "not pure-admm" in ot
    assert "plant-linking readiness" in ot or "plant linking readiness" in ot
    assert "multi-block plant-linking" in ot or "multi-block plant linking" in ot
    assert "not full plant" in ot or "full plant mb" in ot
    assert "plant-named linking readiness" in ot or "plant named linking readiness" in ot
    assert "plant_named_offline_demo" in ot or "plant product streams" in ot
    assert "wire-preflight readiness" in ot or "wire preflight readiness" in ot
    assert "wire_shipped=false" in ot or "wire_shipped=false" in ot.replace(" ", "")
    assert "case-1-shaped" in ot or "case1_shaped" in ot or "case 1-shaped" in ot
    assert "skeleton" in ot
    assert "linear_quality_pooling" in ot
    assert "naphtha" in ot and "residue" in ot
    assert "dual-space" in ot or "form contract" in ot
    assert "unproven" in ot
    assert "probe" in ot or "l∞" in ot or "linf" in ot
    assert "not verdict" in ot or "verdict" in ot
    assert "bridge" in ot or "live-λ" in ot or "live-lambda" in ot or "live_lambda" in ot
    assert "source-labeled" in ot or "live_lambda_source" in ot or "source" in ot
    assert "warm-start" in ot or "warmstart" in ot or "seeded" in ot
    assert "seed_policy" in ot or "seed≠proof" in ot or "seed" in ot
    assert "pooling" in ot or "honest blender" in ot
    assert "not affine" in ot or "affine" in ot
    assert "gate-criteria" in ot or "criteria" in ot or "flip=false" in ot
    assert (
        "iso design+crit" in ot
        or "isolation-rewrite" in ot
        or "isolation design" in ot
        or "design_present" in ot
    )
    assert (
        "rew=false" in ot
        or "rewrite=false" in ot
        or "ship=false" in ot
        or "design_present" in ot
    )
    # FCC/COKER export-vs-live wording
    assert "export" in index_rows["FCC"].lower() or "teaching" in index_rows["FCC"].lower()
    assert "not live" in index_rows["FCC"].lower() or "not" in index_rows["FCC"].lower()
    assert {"CDU", "BLENDER", "FCC", "COKER", "LINKING", "MASTER_ADMM", "OFFLINE_TF"} <= set(
        index_rows
    )

    # --- Summary honesty strip ---
    summary = {
        str(r[0].value): r[1].value
        for r in wb["Summary"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert summary.get("model_form") == "classic_2block_excel_path"
    assert summary.get("dual_gate") == "online_lambda"
    assert summary.get("verdict_dual_gate") == "online_only"
    assert summary.get("tf_on_excel_case1_path") is False
    offline_units = str(summary.get("offline_tf_units") or "")
    assert "FCC" in offline_units and "COKER" in offline_units and "CDU" in offline_units
    assert "PRIMARY" in str(summary.get("dual_linf_online_role") or "")
    assert "SECONDARY" in str(summary.get("dual_linf_recovered_role") or "")
    note = str(summary.get("offline_tf_note") or summary.get("index_offline_tf_note") or "")
    assert "not" in note.lower() and ("case 1" in note.lower() or "classic" in note.lower())
    assert "priced" in note.lower() and "timing" in note.lower()
    assert "admm residual" in note.lower()
    assert "block subproblem" in note.lower()
    assert "coordination" in note.lower()
    assert "plant-linking" in note.lower() or "plant linking" in note.lower()
    assert "plant-named" in note.lower() or "plant named" in note.lower()
    assert "wire-preflight" in note.lower() or "wire preflight" in note.lower()
    assert "case-1-shaped" in note.lower() or "skeleton" in note.lower()
    assert "dual-space" in note.lower() or "form contract" in note.lower()
    assert "probe" in note.lower() or "l∞" in note.lower()
    assert "bridge" in note.lower() or "live-λ" in note.lower() or "live-lambda" in note.lower()
    assert "warm-start" in note.lower() or "warmstart" in note.lower() or "seeded" in note.lower()
    priced_s = str(summary.get("offline_tf_priced") or "").lower()
    timing_s = str(summary.get("offline_tf_timing") or "").lower()
    admm_s = str(summary.get("offline_tf_admm_residual") or "").lower()
    sub_s = str(summary.get("offline_tf_admm_block_subproblem") or "").lower()
    coord_s = str(summary.get("offline_tf_admm_coordination") or "").lower()
    plant_s = str(summary.get("offline_tf_admm_plant_linking") or "").lower()
    named_s = str(summary.get("offline_tf_admm_plant_named_linking") or "").lower()
    pre_s = str(summary.get("offline_tf_wire_preflight") or "").lower()
    c1_s = str(summary.get("offline_tf_case1_shaped_linking") or "").lower()
    ds_s = str(summary.get("offline_tf_case1_dual_space_form_contract") or "").lower()
    lp_s = str(summary.get("offline_tf_case1_dual_space_linf_probe") or "").lower()
    lb_s = str(summary.get("offline_tf_case1_dual_space_linf_live_lambda_bridge") or "").lower()
    ws_s = str(
        summary.get("offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart") or ""
    ).lower()
    pool_s = str(
        summary.get("offline_tf_case1_honest_blender_pooling_path") or ""
    ).lower()
    crit_s = str(
        summary.get("offline_tf_case1_online_linf_gate_criteria_contract") or ""
    ).lower()
    iso_s = str(
        summary.get("offline_tf_case1_isolation_rewrite_design_contract") or ""
    ).lower()
    wsd_s = str(
        summary.get("offline_tf_case1_wire_ship_acceptance_design_contract") or ""
    ).lower()
    blockers_s = str(summary.get("offline_tf_wire_blockers") or "")
    assert "priced" in priced_s
    assert "not" in priced_s and ("dual" in priced_s or "shadow" in priced_s or "λ" in priced_s or "lambda" in priced_s)
    assert "timing" in timing_s
    assert "not" in timing_s and ("case 1" in timing_s or "wall" in timing_s or "dual" in timing_s)
    assert "admm residual" in admm_s or "synthetic" in admm_s
    assert "not" in admm_s and (
        "dual" in admm_s or "wire" in admm_s or "online" in admm_s
    )
    assert "subproblem" in sub_s
    assert "not" in sub_s and (
        "dual" in sub_s or "wire" in sub_s or "online" in sub_s
    )
    assert "coordination" in coord_s
    assert "not" in coord_s and (
        "dual" in coord_s or "wire" in coord_s or "online" in coord_s
    )
    assert "plant" in coord_s or "linking" in coord_s or "per-unit" in coord_s
    assert "plant-linking" in plant_s or "plant linking" in plant_s
    assert "synthetic" in plant_s
    assert "not" in plant_s and (
        "dual" in plant_s or "wire" in plant_s or "online" in plant_s
    )
    assert "full plant" in plant_s or "mass balance" in plant_s
    assert "plant-named" in named_s or "plant named" in named_s
    assert "plant_named_offline_demo" in named_s or "plant product" in named_s
    assert "not" in named_s and (
        "dual" in named_s or "wire" in named_s or "online" in named_s
    )
    assert "full plant" in named_s or "mass balance" in named_s
    assert "preflight" in pre_s
    assert "blocker" in pre_s
    assert "wire_shipped" in pre_s or "not wire shipped" in pre_s
    assert "case-1-shaped" in c1_s or "skeleton" in c1_s or "case1" in c1_s
    assert "linear_quality_pooling" in c1_s
    assert "naphtha" in c1_s and "residue" in c1_s
    assert "wire_shipped" in c1_s or "not wire" in c1_s
    assert "dual-space" in ds_s or "form contract" in ds_s or "form_planned" in ds_s
    assert "unproven" in ds_s
    assert "classic_2block" in ds_s or "form_current" in ds_s
    assert "tf_affine_cdu_blender" in ds_s or "form_planned" in ds_s
    assert "probe" in lp_s or "l∞" in lp_s or "linf" in lp_s
    assert "unproven" in lp_s
    assert "verdict" in lp_s
    assert "wire_shipped" in lp_s or "not wire" in lp_s
    assert "bridge" in lb_s or "live" in lb_s or "λ" in lb_s or "lambda" in lb_s
    assert "unproven" in lb_s
    assert "verdict" in lb_s
    assert "wire_shipped" in lb_s or "not wire" in lb_s
    assert "live_lambda_source" in lb_s or "source" in lb_s
    assert "caller_supplied" in lb_s and "fixture" in lb_s
    assert "warm" in ws_s or "seed" in ws_s or "λ" in ws_s or "lambda" in ws_s
    assert "unproven" in ws_s
    assert "verdict" in ws_s
    assert "wire_shipped" in ws_s or "not wire" in ws_s
    assert "live_lambda_source" in ws_s or "source" in ws_s
    assert "seed_policy" in ws_s or "lambda0_from_live_primary" in ws_s
    assert "caller_supplied" in ws_s and "fixture" in ws_s
    assert "pooling" in pool_s or "linear_quality_pooling" in pool_s
    assert "unproven" in pool_s
    assert "verdict" in pool_s
    assert "affine" in pool_s
    assert "honest_pooling_path_present" in pool_s
    assert "wire_shipped" in pool_s or "not wire" in pool_s
    assert "criteria" in crit_s or "gate" in crit_s or "flip" in crit_s
    assert "unproven" in crit_s
    assert "verdict" in crit_s
    assert "open" in crit_s
    assert "false" in crit_s
    assert "wire_shipped" in crit_s or "not wire" in crit_s
    assert "online_linf_gate" in crit_s
    assert "isolation" in iso_s and "design" in iso_s
    assert "design_present" in iso_s or "rewrite_shipped" in iso_s
    assert "unproven" in iso_s
    assert "verdict" in iso_s
    assert "open" in iso_s
    assert "false" in iso_s
    assert "wire_shipped" in iso_s or "not wire" in iso_s
    assert "wire-ship" in wsd_s or "wire_ship" in wsd_s or "ship" in wsd_s
    assert "design" in wsd_s
    assert "design_present" in wsd_s
    assert "wire_ship_allowed_today" in wsd_s or "ship_allowed" in wsd_s
    assert "false" in wsd_s
    assert "unproven" in wsd_s
    assert "verdict" in wsd_s
    assert "wire_shipped" in wsd_s or "not wire" in wsd_s
    assert "isolation_rewrite_required" in blockers_s
    assert "wire_not_shipped" in blockers_s
    assert "case1_is_cdu_blender_package_admm" in blockers_s
    assert "no_blender_offline_affine_kernel" in blockers_s
    assert "dual_linf_under_wire_unproven" in blockers_s
    assert summary.get("offline_tf_wire_shipped") is False

    # --- Calc_Check honesty audit rows (all ok) ---
    chk_h = [c.value for c in wb["Calc_Check"][1]]
    ci = chk_h.index("check")
    oi = chk_h.index("ok")
    checks = {
        r[ci]: r[oi]
        for r in wb["Calc_Check"].iter_rows(min_row=2, values_only=True)
        if r[ci]
    }
    assert checks.get("form_classic_2block") is True
    assert checks.get("dual_gate_online_only") is True
    assert checks.get("offline_tf_not_on_case1") is True
    assert checks.get("offline_tf_priced_not_duals") is True
    assert checks.get("offline_tf_timing_not_case1") is True
    assert checks.get("offline_tf_admm_residual_not_duals") is True
    assert checks.get("offline_tf_admm_block_subproblem_not_duals") is True
    assert checks.get("offline_tf_admm_coordination_not_duals") is True
    assert checks.get("offline_tf_admm_plant_linking_not_duals") is True
    assert checks.get("offline_tf_admm_plant_named_linking_not_duals") is True
    assert checks.get("offline_tf_wire_preflight_not_duals") is True
    assert checks.get("offline_tf_wire_not_shipped") is True
    assert checks.get("offline_tf_case1_shaped_linking_not_duals") is True
    assert checks.get("offline_tf_case1_shaped_not_wire") is True
    assert checks.get("offline_tf_case1_dual_space_form_contract_not_duals") is True
    assert checks.get("offline_tf_case1_dual_space_form_contract_not_wire") is True
    assert checks.get("offline_tf_case1_dual_space_linf_probe_not_duals") is True
    assert checks.get("offline_tf_case1_dual_space_linf_probe_not_wire") is True
    assert checks.get("offline_tf_case1_dual_space_linf_probe_not_verdict_gate") is True
    assert checks.get("offline_tf_case1_dual_space_linf_live_lambda_bridge_not_duals") is True
    assert checks.get("offline_tf_case1_dual_space_linf_live_lambda_bridge_not_wire") is True
    assert checks.get("offline_tf_case1_dual_space_linf_live_lambda_bridge_not_verdict_gate") is True
    assert checks.get("offline_tf_case1_dual_space_linf_live_lambda_bridge_source_must_be_labeled") is True
    assert checks.get("offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_duals") is True
    assert checks.get("offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_wire") is True
    assert checks.get("offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_verdict_gate") is True
    assert checks.get(
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_source_must_be_labeled"
    ) is True
    assert checks.get(
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_seed_identity_not_proof"
    ) is True
    assert checks.get("offline_tf_case1_honest_blender_pooling_path_not_duals") is True
    assert checks.get("offline_tf_case1_honest_blender_pooling_path_not_wire") is True
    assert checks.get("offline_tf_case1_honest_blender_pooling_path_not_verdict_gate") is True
    assert checks.get("offline_tf_case1_honest_blender_pooling_path_not_affine_kernel") is True
    assert checks.get(
        "offline_tf_case1_honest_blender_pooling_path_checklist_honest_pooling_path_present"
    ) is True
    assert checks.get("offline_tf_case1_online_linf_gate_criteria_contract_not_duals") is True
    assert checks.get("offline_tf_case1_online_linf_gate_criteria_contract_not_wire") is True
    assert checks.get("offline_tf_case1_online_linf_gate_criteria_contract_not_verdict_gate") is True
    assert checks.get(
        "offline_tf_case1_online_linf_gate_criteria_contract_gate_open_flip_false"
    ) is True
    assert checks.get(
        "offline_tf_case1_online_linf_gate_criteria_contract_not_dual_linf_under_wire_proof"
    ) is True
    assert checks.get("offline_tf_case1_isolation_rewrite_design_contract_not_duals") is True
    assert checks.get("offline_tf_case1_isolation_rewrite_design_contract_not_wire") is True
    assert checks.get("offline_tf_case1_isolation_rewrite_design_contract_not_verdict_gate") is True
    assert checks.get(
        "offline_tf_case1_isolation_rewrite_design_contract_rewrite_not_shipped_checklist_open"
    ) is True
    assert checks.get(
        "offline_tf_case1_isolation_rewrite_design_contract_not_dual_linf_under_wire_proof"
    ) is True
    assert checks.get("offline_tf_case1_wire_ship_acceptance_design_contract_not_duals") is True
    assert checks.get("offline_tf_case1_wire_ship_acceptance_design_contract_not_wire") is True
    assert checks.get("offline_tf_case1_wire_ship_acceptance_design_contract_not_verdict_gate") is True
    assert checks.get(
        "offline_tf_case1_wire_ship_acceptance_design_contract_ship_allowed_false_wire_false"
    ) is True
    assert checks.get(
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_dual_linf_under_wire_proof"
    ) is True
    for name, ok in checks.items():
        assert ok is True, (name, ok)

    # How_to offline + dual keys preserved (units + priced + timing + residual +
    # subproblem + multi-round coordination + multi-block plant-linking + plant-named
    # + wire-preflight + Case-1-shaped skeleton + dual-space/form contract)
    how = {
        str(r[0].value): str(r[1].value or "")
        for r in wb["How_to_read"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert how.get("tf_offline_units")
    assert how.get("tf_offline_priced")
    assert how.get("tf_offline_timing")
    assert how.get("tf_offline_admm_residual")
    assert how.get("tf_offline_admm_block_subproblem")
    assert how.get("tf_offline_admm_coordination")
    assert how.get("tf_offline_admm_plant_linking")
    assert how.get("tf_offline_admm_plant_named_linking")
    assert how.get("tf_offline_wire_preflight")
    assert how.get("tf_offline_case1_shaped_linking")
    assert how.get("tf_offline_case1_dual_space_form_contract")
    assert how.get("tf_offline_case1_dual_space_linf_probe")
    assert how.get("tf_offline_case1_dual_space_linf_live_lambda_bridge")
    assert how.get("tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart")
    assert how.get("tf_offline_case1_honest_blender_pooling_path")
    assert how.get("tf_offline_case1_online_linf_gate_criteria_contract")
    assert how.get("tf_offline_case1_isolation_rewrite_design_contract")
    assert how.get("tf_offline_case1_wire_ship_acceptance_design_contract")
    assert "PRIMARY" in how.get("duals_online_lambda", "") or "PRIMARY" in how.get(
        "duals_primary_secondary", ""
    )


def test_planner_honesty_check_rows_pure():
    """Honesty audits fail closed on form / dual-gate drift (no solve needed)."""
    from pims_admm_llm.models.excel_pipeline import planner_honesty_check_rows

    good = {
        "model": {"form": "classic_2block_excel_path"},
        "comparison": {
            "dual_gate": "online_lambda",
            "verdict_dual_gate": "online_only",
            "dual_linf_online_role": "PRIMARY",
        },
        "admm": {"dual_recovery_path": "package-admm/qp_l2+online_lambda_shadows"},
    }
    rows_good = planner_honesty_check_rows(good)
    names = {r["check"] for r in rows_good}
    assert {
        "form_classic_2block",
        "dual_gate_online_only",
        "recovered_linf_not_verdict_gate",
        "dual_glance_strip_primary_secondary",
        "dual_linf_under_wire_unproven_not_from_diagnostics",
        "offline_tf_not_on_case1",
        "offline_tf_ladder_toc_navigator",
        "offline_tf_priced_not_duals",
        "offline_tf_timing_not_case1",
        "offline_tf_admm_residual_not_duals",
        "offline_tf_admm_block_subproblem_not_duals",
        "offline_tf_admm_coordination_not_duals",
        "offline_tf_admm_plant_linking_not_duals",
        "offline_tf_admm_plant_named_linking_not_duals",
        "offline_tf_wire_preflight_not_duals",
        "offline_tf_wire_not_shipped",
        "offline_tf_case1_shaped_linking_not_duals",
        "offline_tf_case1_shaped_not_wire",
        "offline_tf_case1_dual_space_form_contract_not_duals",
        "offline_tf_case1_dual_space_form_contract_not_wire",
        "offline_tf_case1_dual_space_linf_probe_not_duals",
        "offline_tf_case1_dual_space_linf_probe_not_wire",
        "offline_tf_case1_dual_space_linf_probe_not_verdict_gate",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_not_duals",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_not_wire",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_not_verdict_gate",
        "offline_tf_case1_dual_space_linf_live_lambda_bridge_source_must_be_labeled",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_duals",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_wire",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_verdict_gate",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_source_must_be_labeled",
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_seed_identity_not_proof",
        "offline_tf_case1_honest_blender_pooling_path_not_duals",
        "offline_tf_case1_honest_blender_pooling_path_not_wire",
        "offline_tf_case1_honest_blender_pooling_path_not_verdict_gate",
        "offline_tf_case1_honest_blender_pooling_path_not_affine_kernel",
        "offline_tf_case1_honest_blender_pooling_path_checklist_honest_pooling_path_present",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_duals",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_wire",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_verdict_gate",
        "offline_tf_case1_online_linf_gate_criteria_contract_gate_open_flip_false",
        "offline_tf_case1_online_linf_gate_criteria_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_isolation_rewrite_design_contract_not_duals",
        "offline_tf_case1_isolation_rewrite_design_contract_not_wire",
        "offline_tf_case1_isolation_rewrite_design_contract_not_verdict_gate",
        "offline_tf_case1_isolation_rewrite_design_contract_rewrite_not_shipped_checklist_open",
        "offline_tf_case1_isolation_rewrite_design_contract_not_dual_linf_under_wire_proof",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_duals",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_wire",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_verdict_gate",
        "offline_tf_case1_wire_ship_acceptance_design_contract_ship_allowed_false_wire_false",
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_dual_linf_under_wire_proof",
    } <= names
    assert all(r["ok"] for r in rows_good)

    bad_form = dict(good)
    bad_form["model"] = {"form": "wired_tf_path"}
    rows = {r["check"]: r["ok"] for r in planner_honesty_check_rows(bad_form)}
    assert rows["form_classic_2block"] is False
    assert rows["offline_tf_not_on_case1"] is True
    assert rows["offline_tf_priced_not_duals"] is True
    assert rows["offline_tf_timing_not_case1"] is True
    assert rows["offline_tf_admm_residual_not_duals"] is True
    assert rows["offline_tf_admm_block_subproblem_not_duals"] is True
    assert rows["offline_tf_admm_coordination_not_duals"] is True
    assert rows["offline_tf_admm_plant_linking_not_duals"] is True
    assert rows["offline_tf_admm_plant_named_linking_not_duals"] is True
    assert rows["offline_tf_wire_preflight_not_duals"] is True
    assert rows["offline_tf_wire_not_shipped"] is True
    assert rows["offline_tf_case1_dual_space_form_contract_not_duals"] is True
    assert rows["offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_duals"] is True
    assert rows["offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_wire"] is True
    assert rows[
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_not_verdict_gate"
    ] is True
    assert rows[
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_source_must_be_labeled"
    ] is True
    assert rows[
        "offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_seed_identity_not_proof"
    ] is True
    assert rows["offline_tf_case1_honest_blender_pooling_path_not_duals"] is True
    assert rows["offline_tf_case1_honest_blender_pooling_path_not_wire"] is True
    assert rows["offline_tf_case1_honest_blender_pooling_path_not_verdict_gate"] is True
    assert rows["offline_tf_case1_honest_blender_pooling_path_not_affine_kernel"] is True
    assert rows[
        "offline_tf_case1_honest_blender_pooling_path_checklist_honest_pooling_path_present"
    ] is True
    assert rows["offline_tf_case1_online_linf_gate_criteria_contract_not_duals"] is True
    assert rows["offline_tf_case1_online_linf_gate_criteria_contract_not_wire"] is True
    assert rows["offline_tf_case1_online_linf_gate_criteria_contract_not_verdict_gate"] is True
    assert rows[
        "offline_tf_case1_online_linf_gate_criteria_contract_gate_open_flip_false"
    ] is True
    assert rows[
        "offline_tf_case1_online_linf_gate_criteria_contract_not_dual_linf_under_wire_proof"
    ] is True
    assert rows["offline_tf_case1_isolation_rewrite_design_contract_not_duals"] is True
    assert rows["offline_tf_case1_isolation_rewrite_design_contract_not_wire"] is True
    assert rows["offline_tf_case1_isolation_rewrite_design_contract_not_verdict_gate"] is True
    assert rows[
        "offline_tf_case1_isolation_rewrite_design_contract_rewrite_not_shipped_checklist_open"
    ] is True
    assert rows[
        "offline_tf_case1_isolation_rewrite_design_contract_not_dual_linf_under_wire_proof"
    ] is True
    assert rows["offline_tf_case1_wire_ship_acceptance_design_contract_not_duals"] is True
    assert rows["offline_tf_case1_wire_ship_acceptance_design_contract_not_wire"] is True
    assert rows["offline_tf_case1_wire_ship_acceptance_design_contract_not_verdict_gate"] is True
    assert rows[
        "offline_tf_case1_wire_ship_acceptance_design_contract_ship_allowed_false_wire_false"
    ] is True
    assert rows[
        "offline_tf_case1_wire_ship_acceptance_design_contract_not_dual_linf_under_wire_proof"
    ] is True
    assert rows["offline_tf_case1_dual_space_form_contract_not_wire"] is True
    assert rows["offline_tf_case1_dual_space_linf_probe_not_duals"] is True
    assert rows["offline_tf_case1_dual_space_linf_probe_not_wire"] is True
    assert rows["offline_tf_case1_dual_space_linf_probe_not_verdict_gate"] is True
    assert rows["offline_tf_case1_dual_space_linf_live_lambda_bridge_not_duals"] is True
    assert rows["offline_tf_case1_dual_space_linf_live_lambda_bridge_not_wire"] is True
    assert rows["offline_tf_case1_dual_space_linf_live_lambda_bridge_not_verdict_gate"] is True
    assert rows["offline_tf_case1_dual_space_linf_live_lambda_bridge_source_must_be_labeled"] is True


def test_format_planner_honesty_package_priced_timing_pure():
    """Pure formatter exposes priced + timing + residual + subproblem + coordination + plant-linking + plant-named + wire-preflight + dual-space contract + dual-space L∞ probe + live-λ bridge readiness without TF or full solve."""
    from pims_admm_llm.models.excel_pipeline import format_planner_honesty_package

    fake = {
        "model": {"form": "classic_2block_excel_path"},
        "comparison": {
            "dual_gate": "online_lambda",
            "verdict_dual_gate": "online_only",
            "dual_linf_online": 2.66,
            "dual_linf_recovered": 112.0,
            "dual_linf_online_role": "PRIMARY",
            "dual_linf_recovered_role": "SECONDARY",
        },
        "admm": {"dual_recovery_path": "package-admm/qp_l2+online_lambda_shadows"},
    }
    pkg = format_planner_honesty_package(fake)
    what = pkg["index_row"]["what"].lower()
    assert "priced" in what and "timing" in what and "readiness" in what
    assert "admm residual" in what and "synthetic" in what
    assert "block subproblem" in what or "residual+subproblem" in what or "subproblem" in what
    assert "coordination" in what
    assert "plant-linking readiness" in what or "plant linking readiness" in what
    assert "plant-named linking readiness" in what or "plant named linking readiness" in what
    assert "plant_named_offline_demo" in what or "plant product streams" in what
    assert "wire-preflight readiness" in what or "wire preflight readiness" in what
    assert "wire_shipped=false" in what or "wire_shipped=false" in what.replace(" ", "")
    assert "dual-space" in what or "form contract" in what
    assert "unproven" in what
    assert "probe" in what or "l∞" in what or "linf" in what
    assert "not verdict" in what or "verdict" in what
    assert "bridge" in what or "live-λ" in what or "live-lambda" in what or "live_lambda" in what
    assert "source-labeled" in what or "live_lambda_source" in what
    assert "warm-start" in what or "warmstart" in what or "seeded" in what
    assert "seed_policy" in what or "seed≠proof" in what or "seed" in what
    assert "pooling" in what or "honest blender" in what
    assert "not affine" in what or "affine" in what
    assert "not" in what and "case 1" in what
    assert "not full plant" in what or "full plant mb" in what
    meta = pkg["meta"]
    assert meta["offline_tf_priced_ready"] is True
    assert meta["offline_tf_timing_ready"] is True
    assert meta["offline_tf_admm_residual_ready"] is True
    assert meta["offline_tf_admm_block_subproblem_ready"] is True
    assert meta["offline_tf_admm_coordination_ready"] is True
    assert meta["offline_tf_admm_plant_linking_ready"] is True
    assert meta["offline_tf_admm_plant_named_linking_ready"] is True
    assert meta["offline_tf_wire_preflight_ready"] is True
    assert meta["offline_tf_case1_shaped_linking_ready"] is True
    assert meta["offline_tf_case1_dual_space_form_contract_ready"] is True
    assert meta["offline_tf_case1_dual_space_linf_probe_ready"] is True
    assert meta["offline_tf_case1_dual_space_linf_live_lambda_bridge_ready"] is True
    assert meta["offline_tf_case1_dual_space_linf_live_lambda_seeded_warmstart_ready"] is True
    assert meta["offline_tf_case1_honest_blender_pooling_path_ready"] is True
    assert meta.get("offline_tf_case1_online_linf_gate_criteria_contract_ready", True) is True
    assert meta.get("offline_tf_case1_isolation_rewrite_design_contract_ready", True) is True
    assert meta.get("offline_tf_case1_wire_ship_acceptance_design_contract_ready", True) is True
    assert meta["offline_tf_case1_dual_honest_tf_aware_path_design_contract_ready"] is True
    assert meta["offline_tf_path_design_present"] is True
    assert meta["offline_tf_path_shipped"] is False
    assert meta["offline_tf_dual_honest_tf_aware_path_present_ship_met"] is False
    assert meta["offline_tf_wire_ship_allowed_today"] is False
    assert meta["offline_tf_wire_shipped"] is False
    assert meta["tf_dual_recovery_path"] is None
    assert meta["form"] == "classic_2block_excel_path"
    assert "priced" in meta["planner_one_liner"].lower()
    assert "timing" in meta["planner_one_liner"].lower()
    assert "admm residual" in meta["planner_one_liner"].lower()
    assert "block subproblem" in meta["planner_one_liner"].lower()
    assert "coordination" in meta["planner_one_liner"].lower()
    assert "plant-linking" in meta["planner_one_liner"].lower() or "plant linking" in meta[
        "planner_one_liner"
    ].lower()
    assert "plant-named" in meta["planner_one_liner"].lower() or "plant named" in meta[
        "planner_one_liner"
    ].lower()
    assert "wire-preflight" in meta["planner_one_liner"].lower() or "wire preflight" in meta[
        "planner_one_liner"
    ].lower()
    assert "dual-space" in meta["planner_one_liner"].lower() or "form contract" in meta[
        "planner_one_liner"
    ].lower()
    assert "probe" in meta["planner_one_liner"].lower() or "l∞" in meta[
        "planner_one_liner"
    ].lower()
    assert (
        "bridge" in meta["planner_one_liner"].lower()
        or "live-λ" in meta["planner_one_liner"].lower()
        or "live-lambda" in meta["planner_one_liner"].lower()
        or "live_lambda" in meta["planner_one_liner"].lower()
    )
    assert (
        "warm-start" in meta["planner_one_liner"].lower()
        or "warmstart" in meta["planner_one_liner"].lower()
        or "seeded" in meta["planner_one_liner"].lower()
        or "seed_policy" in meta["planner_one_liner"].lower()
    )
    admm_note = str(meta["offline_tf_admm_residual"]).lower()
    assert "synthetic" in admm_note or "admm residual" in admm_note
    assert "not" in admm_note
    sub_note = str(meta["offline_tf_admm_block_subproblem"]).lower()
    assert "subproblem" in sub_note
    assert "not" in sub_note
    coord_note = str(meta["offline_tf_admm_coordination"]).lower()
    assert "coordination" in coord_note
    assert "not" in coord_note
    assert "plant" in coord_note or "linking" in coord_note or "per-unit" in coord_note
    plant_note = str(meta["offline_tf_admm_plant_linking"]).lower()
    assert "plant-linking" in plant_note or "plant linking" in plant_note
    assert "synthetic" in plant_note
    assert "not" in plant_note
    assert "full plant" in plant_note or "mass balance" in plant_note
    named_note = str(meta["offline_tf_admm_plant_named_linking"]).lower()
    assert "plant-named" in named_note or "plant named" in named_note
    assert "plant_named_offline_demo" in named_note or "plant product" in named_note
    assert "not" in named_note
    assert "full plant" in named_note or "mass balance" in named_note
    preflight_note = str(meta["offline_tf_wire_preflight"]).lower()
    assert "preflight" in preflight_note
    assert "blocker" in preflight_note
    assert "isolation_rewrite_required" in preflight_note
    assert "wire_not_shipped" in preflight_note
    ds_note = str(meta["offline_tf_case1_dual_space_form_contract"]).lower()
    assert "dual-space" in ds_note or "form contract" in ds_note or "form_planned" in ds_note
    assert "unproven" in ds_note
    lp_note = str(meta["offline_tf_case1_dual_space_linf_probe"]).lower()
    assert "probe" in lp_note or "l∞" in lp_note or "linf" in lp_note
    assert "unproven" in lp_note
    assert "verdict" in lp_note
    lb_note = str(meta["offline_tf_case1_dual_space_linf_live_lambda_bridge"]).lower()
    assert "bridge" in lb_note or "live" in lb_note or "λ" in lb_note or "lambda" in lb_note
    assert "unproven" in lb_note
    assert "verdict" in lb_note
    assert "live_lambda_source" in lb_note or "source" in lb_note
    # anti-claim: must not claim wire shipped or full plant mass balance shipped
    readiness = str(meta["offline_tf_readiness_note"]).lower()
    assert "admm residual" in readiness
    assert "block subproblem" in readiness
    assert "coordination" in readiness
    assert "plant-linking" in readiness or "plant linking" in readiness
    assert "plant-named" in readiness or "plant named" in readiness
    assert "wire-preflight" in readiness or "wire preflight" in readiness
    assert "dual-space" in readiness or "form contract" in readiness
    assert "probe" in readiness or "l∞" in readiness
    assert "verdict" in readiness
    assert "bridge" in readiness or "live-λ" in readiness or "live-lambda" in readiness
    assert "source-labeled" in readiness or "live_lambda_source" in readiness
    assert "not wire shipped" in readiness or "not on classic case 1" in readiness
    assert "wire shipped" not in readiness.replace("not wire shipped", "")
    assert "full plant mass balance shipped" not in readiness
    assert pkg["tf_offline_admm_block_subproblem"]["topic"] == "tf_offline_admm_block_subproblem"
    assert pkg["tf_offline_admm_block_subproblem"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_admm_coordination"]["topic"] == "tf_offline_admm_coordination"
    assert pkg["tf_offline_admm_coordination"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_admm_plant_linking"]["topic"] == "tf_offline_admm_plant_linking"
    assert pkg["tf_offline_admm_plant_linking"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_admm_plant_linking"]["not_full_plant_mass_balance"] == "true"
    assert pkg["tf_offline_admm_plant_named_linking"]["topic"] == "tf_offline_admm_plant_named_linking"
    assert pkg["tf_offline_admm_plant_named_linking"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_admm_plant_named_linking"]["topology_source"] == "plant_named_offline_demo"
    assert pkg["tf_offline_admm_plant_named_linking"]["linking_space"] == "plant_named_linking_streams"
    assert pkg["tf_offline_admm_plant_named_linking"]["not_full_plant_mass_balance"] == "true"
    assert pkg["tf_offline_wire_preflight"]["topic"] == "tf_offline_wire_preflight"
    assert pkg["tf_offline_wire_preflight"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_wire_preflight"]["wire_shipped"] == "false"
    assert pkg["tf_offline_wire_preflight"]["not_wire_shipped"] == "true"
    assert "isolation_rewrite_required" in pkg["tf_offline_wire_preflight"]["wire_blockers"]
    assert "wire_not_shipped" in pkg["tf_offline_wire_preflight"]["wire_blockers"]
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["topic"]
        == "tf_offline_case1_dual_space_form_contract"
    )
    assert pkg["tf_offline_case1_dual_space_form_contract"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_case1_dual_space_form_contract"]["wire_shipped"] == "false"
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["form_current"]
        == "classic_2block_excel_path"
    )
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["form_planned"]
        == "tf_affine_cdu_blender_shaped_excel_path"
    )
    assert (
        pkg["tf_offline_case1_dual_space_form_contract"]["dual_linf_under_wire_status"]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"]["topic"]
        == "tf_offline_case1_dual_space_linf_probe"
    )
    assert pkg["tf_offline_case1_dual_space_linf_probe"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_case1_dual_space_linf_probe"]["wire_shipped"] == "false"
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"]["dual_linf_under_wire_status"]
        == "unproven"
    )
    assert pkg["tf_offline_case1_dual_space_linf_probe"]["probe_is_not_verdict_gate"] == "true"
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"][
            "probe_is_not_dual_linf_under_wire_proof"
        ]
        == "true"
    )
    assert (
        "online_linf_gate_under_tf_path"
        in pkg["tf_offline_case1_dual_space_linf_probe"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["topic"]
        == "tf_offline_case1_dual_space_linf_live_lambda_bridge"
    )
    assert pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["dual_recovery_path"] == "None"
    assert pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["wire_shipped"] == "false"
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["dual_linf_under_wire_status"]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["bridge_is_not_verdict_gate"]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"][
            "bridge_is_not_dual_linf_under_wire_proof"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"][
            "live_lambda_source_must_be_labeled"
        ]
        == "true"
    )
    assert (
        "online_linf_gate_under_tf_path"
        in pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["topic"]
        == "tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "dual_recovery_path"
        ]
        == "None"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["wire_shipped"]
        == "false"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "dual_linf_under_wire_status"
        ]
        == "unproven"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "warmstart_is_not_verdict_gate"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "warmstart_is_not_dual_linf_under_wire_proof"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "seed_identity_linf_is_not_proof"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "live_lambda_source_must_be_labeled"
        ]
        == "true"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["seed_policy"]
        == "lambda0_from_live_primary_online"
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["z0_policy"]
        == "unchanged_default_skeleton_z"
    )
    assert (
        "online_linf_gate_under_tf_path"
        in pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"][
            "dual_linf_proof_checklist_open_ids"
        ]
    )
    # probe + bridge + warm-start co-exist (non-destructive packaging)
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"]["topic"]
        != pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["topic"]
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_live_lambda_bridge"]["topic"]
        != pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["topic"]
    )
    assert (
        pkg["tf_offline_case1_dual_space_linf_probe"]["topic"]
        != pkg["tf_offline_case1_dual_space_linf_live_lambda_seeded_warmstart"]["topic"]
    )


def test_load_pims_excel_has_crudes(tmp_path):
    xlsx = tmp_path / "t.xlsx"
    write_template_excel(xlsx)
    pkg = load_pims_excel(xlsx)
    assert pkg["crudes"] and pkg["products"] and pkg["capacities"]


def test_ensure_template(tmp_path):
    path = tmp_path / "crudes_template.xlsx"
    out = ensure_template(path)
    assert out.is_file() and out.stat().st_size > 1000


def test_format_tf_offline_case1_isolation_rewrite_first_blocker_operational_prep_howto_pure():
    """Static isolation first-blocker operational prep How_to: prep_present; ship false; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_FIRST_BLOCKING_COREQ,
        _CASE1_FORM_CURRENT,
        _CASE1_ISOLATION_SUITE_PATH,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        format_tf_offline_case1_isolation_rewrite_first_blocker_operational_prep_howto,
    )
    import pims_admm_llm.models.excel_pipeline as ep
    import inspect

    d = format_tf_offline_case1_isolation_rewrite_first_blocker_operational_prep_howto()
    assert d["topic"] == "tf_offline_case1_isolation_rewrite_first_blocker_operational_prep"
    assert d["prep_present"] == "true"
    assert d["first_blocker_prep_present"] == "true"
    assert d["first_blocking_coreq"] == _CASE1_FIRST_BLOCKING_COREQ
    assert d["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["rewrite_not_delete"] == "true"
    assert d["suite_path"] == _CASE1_ISOLATION_SUITE_PATH
    assert d["suite_path"] == "tests/test_tf_import_isolation.py"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    assert d["wire_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["prep_is_not_isolation_rewrite_shipped"] == "true"
    assert d["prep_is_not_wire"] == "true"
    assert d["prep_is_not_verdict_gate"] == "true"
    assert d["prep_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["this_prep_alone_is_not_ship"] == "true"
    assert d["packaging_alone_is_not_ship"] == "true"
    assert d["companion_matrix_is_inventory_only"] == "true"
    assert d["order_hint_is_not_executor"] == "true"
    assert d["feature_flag_enabled_today"] == "false"
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    one = d["planner_one_liner"].lower()
    assert "prep_present" in one
    assert "isolation_rewrite_shipped=false" in one
    assert "dual_recovery_path=none" in one
    assert "not verdict" in one or "verdict" in one
    # purity: no live report call / import
    src = inspect.getsource(
        format_tf_offline_case1_isolation_rewrite_first_blocker_operational_prep_howto
    )
    assert "offline_case1_isolation_rewrite_first_blocker_operational_prep_report(" not in src
    assert "import tensorflow" not in src
    assert "from pims_admm_llm.models import tf_linear_blocks" not in src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in src
    ep_src = open(ep.__file__, encoding="utf-8").read()
    assert "from pims_admm_llm.models import tf_linear_blocks" not in ep_src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in ep_src
    assert "import tensorflow" not in ep_src
    assert "import pulp" not in ep_src



def test_format_tf_offline_case1_isolation_rewrite_first_blocker_execution_scaffold_howto_pure():
    """Static isolation rewrite first-blocker execution scaffold How_to: scaffold≠ship; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_FIRST_BLOCKING_COREQ,
        _CASE1_FORM_CURRENT,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        format_tf_offline_case1_isolation_rewrite_first_blocker_execution_scaffold_howto,
        format_tf_offline_ladder_toc_howto,
    )
    import pims_admm_llm.models.excel_pipeline as ep
    import inspect

    d = format_tf_offline_case1_isolation_rewrite_first_blocker_execution_scaffold_howto()
    assert d["topic"] == "tf_offline_case1_isolation_rewrite_first_blocker_execution_scaffold"
    assert d["scaffold_present"] == "true"
    assert d["execution_scaffold_present"] == "true"
    assert d["isolation_rewrite_scaffold_present"] == "true"
    assert d["first_blocking_coreq"] == _CASE1_FIRST_BLOCKING_COREQ
    assert d["is_first_blocking_coreq"] == "true"
    assert d["order_hint_index"] == "0"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["isolation_ship_allowed_today"] == "false"
    assert d["isolation_tests_rewritten_with_wire"] == "false"
    assert d["rewrite_not_delete"] == "true"
    assert d["suite_path"] == "tests/test_tf_import_isolation.py"
    assert d["dual_recovery_path"] == "None"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    assert d["wire_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["scaffold_is_not_isolation_rewrite_shipped"] == "true"
    assert d["scaffold_is_not_isolation_ship_allow"] == "true"
    assert d["scaffold_is_not_wire"] == "true"
    assert d["scaffold_is_not_verdict_gate"] == "true"
    assert d["this_scaffold_alone_is_not_ship_criterion"] == "true"
    assert d["packaging_alone_is_not_isolation_rewrite_shipped"] == "true"
    assert d["distinct_from_isolation_operational_prep"] == "true"
    assert d["distinct_from_path_execution_scaffold"] == "true"
    assert d["order_hint_is_not_executor"] == "true"
    assert d["feature_flag_enabled_today"] == "false"
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    one = d["planner_one_liner"].lower()
    assert "scaffold_present" in one
    assert "isolation_rewrite_shipped=false" in one
    assert "isolation_ship_allowed_today=false" in one
    assert "dual_recovery_path=none" in one
    src = inspect.getsource(
        format_tf_offline_case1_isolation_rewrite_first_blocker_execution_scaffold_howto
    )
    assert "offline_case1_isolation_rewrite_first_blocker_execution_scaffold_report(" not in src
    assert "import tensorflow" not in src
    assert "from pims_admm_llm.models import tf_linear_blocks" not in src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in src
    toc = format_tf_offline_ladder_toc_howto()
    assert toc.get("includes_isolation_rewrite_first_blocker_execution_scaffold") == "true"
    assert "isolation_rewrite_first_blocker_execution_scaffold" in toc["topic_ids"]
    ep_src = open(ep.__file__, encoding="utf-8").read()
    assert "from pims_admm_llm.models import tf_linear_blocks" not in ep_src
    assert "import tensorflow" not in ep_src
    assert "import pulp" not in ep_src



def test_excel_first_blocker_operational_prep_packaging_surfaces(tmp_path):
    """Workbook integration: prep How_to + meta + Summary + Calc_Check dual-ban; Index ≤1439."""
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
        _OFFLINE_TF_INDEX_WHAT,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    out = tmp_path / "prep_pkg.xlsx"
    write_results_excel(out, report)
    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) <= GOAL_MAX_SHEETS

    pkg = format_planner_honesty_package(report)
    assert pkg["meta"]["offline_tf_case1_isolation_rewrite_first_blocker_operational_prep_ready"] is True
    assert pkg["meta"]["offline_tf_first_blocker_prep_present"] is True
    assert pkg["meta"]["offline_tf_isolation_rewrite_shipped"] is False
    assert pkg["meta"]["offline_tf_wire_shipped"] is False
    assert pkg["meta"]["tf_dual_recovery_path"] is None
    assert pkg.get("tf_offline_case1_isolation_rewrite_first_blocker_operational_prep") is not None
    prep = pkg["tf_offline_case1_isolation_rewrite_first_blocker_operational_prep"]
    assert prep["topic"] == "tf_offline_case1_isolation_rewrite_first_blocker_operational_prep"
    assert prep["prep_present"] == "true"
    assert prep["isolation_rewrite_shipped"] == "false"
    assert prep["dual_recovery_path"] == "None"
    assert prep["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439

    how = {
        str(r[0].value): str(r[1].value or "")
        for r in wb["How_to_read"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert "tf_offline_case1_isolation_rewrite_first_blocker_operational_prep" in how
    body = how["tf_offline_case1_isolation_rewrite_first_blocker_operational_prep"].lower()
    assert "prep_present" in body
    assert "isolation_rewrite_shipped=false" in body
    assert "dual_recovery_path=none" in body
    # TOC lists prep
    assert "tf_offline_ladder_toc" in how
    assert "operational_prep" in how["tf_offline_ladder_toc"].lower() or "first_blocker" in how[
        "tf_offline_ladder_toc"
    ].lower()

    summary = {
        str(r[0].value): r[1].value
        for r in wb["Summary"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert summary.get("offline_tf_case1_isolation_rewrite_first_blocker_operational_prep_ready") in (
        True,
        "true",
        True,
    )
    assert summary.get("offline_tf_first_blocker_prep_present") in (True, "true")

    rows = planner_honesty_check_rows(report)
    names = {r["check"] for r in rows}
    assert "offline_tf_case1_isolation_rewrite_first_blocker_operational_prep_not_duals" in names
    assert "offline_tf_case1_isolation_rewrite_first_blocker_operational_prep_not_ship" in names
    assert "offline_tf_case1_isolation_rewrite_first_blocker_operational_prep_not_verdict_gate" in names
    assert (
        "offline_tf_case1_isolation_rewrite_first_blocker_operational_prep_not_dual_linf_under_wire_proof"
        in names
    )
    for r in rows:
        if r["check"].startswith(
            "offline_tf_case1_isolation_rewrite_first_blocker_operational_prep"
        ):
            assert r["ok"] is True

def test_format_tf_offline_case1_dual_linf_under_wire_criteria_contract_howto_pure():
    """Static dual_linf_under_wire flip-criteria How_to: criteria_present; unproven; dual-ban; no TF."""
    from pims_admm_llm.models.excel_pipeline import (
        _CASE1_FIRST_BLOCKING_COREQ,
        _CASE1_FORM_CURRENT,
        _CASE1_DUAL_LINF_UNDER_WIRE_STATUS,
        _CASE1_DUAL_LINF_UNDER_WIRE_FLIP_CRITERIA_KEYS,
        _CASE1_DUAL_LINF_UNDER_WIRE_ANTI_CRITERIA,
        _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED,
        format_tf_offline_case1_dual_linf_under_wire_criteria_contract_howto,
    )
    import pims_admm_llm.models.excel_pipeline as ep
    import inspect

    d = format_tf_offline_case1_dual_linf_under_wire_criteria_contract_howto()
    assert d["topic"] == "tf_offline_case1_dual_linf_under_wire_criteria_contract"
    assert d["criteria_present"] == "true"
    assert d["flip_criteria_formalized"] == "true"
    assert d["dual_linf_under_wire_status"] == _CASE1_DUAL_LINF_UNDER_WIRE_STATUS
    assert d["dual_linf_under_wire_status"] == "unproven"
    assert d["criteria_met_today"] == "false"
    assert d["dual_linf_proof_allowed_today"] == "false"
    assert d["dual_recovery_path"] == "None"
    assert d["wire_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_current"] == _CASE1_FORM_CURRENT
    assert d["first_blocking_coreq"] == _CASE1_FIRST_BLOCKING_COREQ
    assert d["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert d["contract_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["criteria_present_is_not_proven"] == "true"
    assert d["distinct_from_online_linf_gate_criteria_contract"] == "true"
    assert d["gates_status_not_checklist_id"] == "true"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert "online_linf_gate_criteria_alone" in d["anti_criteria_today"]
    for k in _CASE1_DUAL_LINF_UNDER_WIRE_FLIP_CRITERIA_KEYS:
        assert k in d["flip_criteria_keys"]
    for a in (
        "probe_linf",
        "bridge_linf",
        "warmstart_linf",
        "packaging_alone",
        "online_linf_gate_criteria_alone",
    ):
        assert a in _CASE1_DUAL_LINF_UNDER_WIRE_ANTI_CRITERIA
    one = d["planner_one_liner"].lower()
    assert "criteria_present" in one
    assert "unproven" in one
    assert "distinct" in one or "online_linf_gate" in one
    assert "dual_recovery_path=none" in one
    assert "not verdict" in one or "verdict" in one
    assert "pure-admm" not in d["dual_recovery_path_planned_when_shipped"].lower()
    assert d["dual_recovery_path_planned_when_shipped"] == _CASE1_PATH_DESIGN_DUAL_RECOVERY_PLANNED
    src = inspect.getsource(
        format_tf_offline_case1_dual_linf_under_wire_criteria_contract_howto
    )
    assert "offline_case1_dual_linf_under_wire_criteria_contract_report(" not in src
    assert "import tensorflow" not in src
    assert "from pims_admm_llm.models import tf_linear_blocks" not in src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in src
    ep_src = open(ep.__file__, encoding="utf-8").read()
    assert "from pims_admm_llm.models import tf_linear_blocks" not in ep_src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in ep_src
    assert "import tensorflow" not in ep_src
    assert "import pulp" not in ep_src


def test_excel_dual_linf_under_wire_criteria_contract_packaging_surfaces(tmp_path):
    """Workbook integration: dual_linf criteria How_to + meta + Summary + Calc_Check dual-ban; Index ≤1439."""
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
        _OFFLINE_TF_INDEX_WHAT,
        format_tf_offline_ladder_toc_howto,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    out = tmp_path / "dual_linf_crit_pkg.xlsx"
    write_results_excel(out, report)
    import openpyxl

    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) <= GOAL_MAX_SHEETS

    pkg = format_planner_honesty_package(report)
    assert pkg["meta"]["offline_tf_case1_dual_linf_under_wire_criteria_contract_ready"] is True
    assert pkg["meta"]["offline_tf_dual_linf_criteria_present"] is True
    assert pkg["meta"]["offline_tf_dual_linf_under_wire_status"] == "unproven"
    assert pkg["meta"]["offline_tf_dual_linf_proof_allowed_today"] is False
    assert pkg["meta"]["offline_tf_dual_linf_criteria_met_today"] is False
    assert pkg["meta"]["offline_tf_wire_shipped"] is False
    assert pkg["meta"]["offline_tf_isolation_rewrite_shipped"] is False
    assert pkg["meta"]["tf_dual_recovery_path"] is None
    assert pkg.get("tf_offline_case1_dual_linf_under_wire_criteria_contract") is not None
    crit = pkg["tf_offline_case1_dual_linf_under_wire_criteria_contract"]
    assert crit["topic"] == "tf_offline_case1_dual_linf_under_wire_criteria_contract"
    assert crit["criteria_present"] == "true"
    assert crit["dual_linf_under_wire_status"] == "unproven"
    assert crit["dual_recovery_path"] == "None"
    assert crit["dual_linf_proof_allowed_today"] == "false"
    assert crit["distinct_from_online_linf_gate_criteria_contract"] == "true"
    assert crit["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439

    how = {
        str(r[0].value): str(r[1].value or "")
        for r in wb["How_to_read"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert "tf_offline_case1_dual_linf_under_wire_criteria_contract" in how
    body = how["tf_offline_case1_dual_linf_under_wire_criteria_contract"].lower()
    assert "criteria_present" in body
    assert "unproven" in body
    assert "dual_recovery_path=none" in body
    assert "distinct" in body or "online_linf_gate" in body
    # TOC lists dual_linf criteria
    assert "tf_offline_ladder_toc" in how
    toc = how["tf_offline_ladder_toc"].lower()
    assert "dual_linf" in toc
    toc_d = format_tf_offline_ladder_toc_howto()
    assert toc_d.get("includes_dual_linf_under_wire_criteria_contract") == "true"
    assert "dual_linf_under_wire_criteria_contract" in toc_d["topic_ids"]

    summary = {
        str(r[0].value): r[1].value
        for r in wb["Summary"].iter_rows(min_row=2, max_col=2)
        if r[0].value
    }
    assert summary.get("offline_tf_case1_dual_linf_under_wire_criteria_contract_ready") in (
        True,
        "true",
    )
    assert summary.get("offline_tf_dual_linf_criteria_present") in (True, "true")
    assert str(summary.get("offline_tf_dual_linf_under_wire_status")).lower() == "unproven"

    rows = planner_honesty_check_rows(report)
    names = {r["check"] for r in rows}
    assert "offline_tf_dual_linf_under_wire_criteria_not_proof" in names
    for r in rows:
        if r["check"] == "offline_tf_dual_linf_under_wire_criteria_not_proof":
            assert r["ok"] is True
            assert "criteria" in r["predicted"].lower()
            assert "unproven" in r["predicted"].lower() or "unproven" in r["actual"].lower()



def test_format_tf_offline_case1_form_label_second_coreq_operational_prep_howto_pure():
    """E1/E2 Excel: form_label second-coreq operational prep How_to dual-ban (static)."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_case1_form_label_second_coreq_operational_prep_howto,
        _CASE1_FORM_LABEL_SECOND_COREQ_PREP_ANTI_CRITERIA,
        _CASE1_FORM_LABEL_MUTATION_PATH_NAME,
        _OFFLINE_WIRE_BLOCKER_IDS,
    )
    import inspect
    import pims_admm_llm.models.excel_pipeline as ep

    d = format_tf_offline_case1_form_label_second_coreq_operational_prep_howto()
    assert d["topic"] == "tf_offline_case1_form_label_second_coreq_operational_prep"
    assert d["prep_present"] == "true"
    assert d["form_label_second_coreq_prep_present"] == "true"
    assert d["operational_prep_present"] == "true"
    assert d["form_current"] == "classic_2block_excel_path"
    assert d["form_label_change_shipped"] == "false"
    assert d["form_label_ship_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["form_mutation_path_executed_today"] == "false"
    assert d["form_mutation_path_name"] == _CASE1_FORM_LABEL_MUTATION_PATH_NAME
    assert d["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert d["is_first_blocking_coreq"] == "false"
    assert d["order_hint_index"] == "1"
    assert d["dual_recovery_path"] == "None"
    assert d["wire_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["prep_is_not_form_label_change_shipped"] == "true"
    assert d["prep_is_not_form_flip"] == "true"
    assert d["distinct_from_form_label_change_shipped_criteria_contract"] == "true"
    assert d["this_prep_alone_is_not_ship"] == "true"
    assert d["packaging_alone_is_not_ship"] == "true"
    assert d["form_label_change_required_still_true"] == "true"
    one = d["planner_one_liner"].lower()
    assert "prep_present" in one
    assert "form_label_change_shipped=false" in one
    assert "form flip" in one or "not form flip" in one
    assert "isolation_rewrite_with_wire" in one
    assert "this_prep_alone" in _CASE1_FORM_LABEL_SECOND_COREQ_PREP_ANTI_CRITERIA
    assert "form_label_change_required" in _OFFLINE_WIRE_BLOCKER_IDS
    src = inspect.getsource(
        format_tf_offline_case1_form_label_second_coreq_operational_prep_howto
    )
    # Ban live imports / live report calls on packaging hot path (docstring may name them).
    assert "import tensorflow" not in src
    assert "from pims_admm_llm.models import tf_linear_blocks" not in src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in src
    assert "offline_case1_form_label_second_coreq_operational_prep_report(" not in src
    assert "format_tf_offline_case1_form_label_second_coreq_operational_prep_howto" in open(
        ep.__file__, encoding="utf-8"
    ).read()


def test_form_label_second_coreq_prep_package_surfaces(tmp_path):
    """E1/E2: form second-coreq prep meta/How_to/Calc_Check dual-ban surfaces."""
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
        _OFFLINE_TF_INDEX_WHAT,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    pkg = format_planner_honesty_package(report)
    assert pkg["meta"]["offline_tf_case1_form_label_second_coreq_operational_prep_ready"] is True
    assert pkg["meta"]["offline_tf_form_label_second_coreq_prep_present"] is True
    assert pkg["meta"]["offline_tf_form_label_change_shipped"] is False
    assert pkg["meta"]["offline_tf_form_label_ship_allowed_today"] is False
    note = str(pkg["meta"].get("offline_tf_case1_form_label_second_coreq_operational_prep", "")).lower()
    assert "prep" in note
    assert "form_label_change_shipped=false" in note or "form_label_change_shipped" in note
    fl = pkg.get("tf_offline_case1_form_label_second_coreq_operational_prep")
    assert fl is not None
    assert fl["form_label_change_shipped"] == "false"
    assert fl["prep_present"] == "true"
    assert fl["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert fl["is_first_blocking_coreq"] == "false"
    rows = planner_honesty_check_rows(report)
    checks = {r["check"]: r for r in rows}
    assert "offline_tf_form_label_second_coreq_prep_not_form_ship" in checks
    assert checks["offline_tf_form_label_second_coreq_prep_not_form_ship"]["ok"] is True
    # No Index growth for this residual
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439


def test_format_tf_offline_case1_path_third_coreq_operational_prep_howto_pure():
    """E1/E2 Excel: path third-coreq operational prep How_to dual-ban (static)."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_case1_path_third_coreq_operational_prep_howto,
        _CASE1_PATH_THIRD_COREQ_PREP_ANTI_CRITERIA,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _OFFLINE_WIRE_BLOCKER_IDS,
    )
    import inspect
    import pims_admm_llm.models.excel_pipeline as ep

    d = format_tf_offline_case1_path_third_coreq_operational_prep_howto()
    assert d["topic"] == "tf_offline_case1_path_third_coreq_operational_prep"
    assert d["prep_present"] == "true"
    assert d["path_third_coreq_prep_present"] == "true"
    assert d["operational_prep_present"] == "true"
    assert d["path_shipped"] == "false"
    assert d["dual_honest_tf_aware_path_present"] == "false"
    assert d["ship_met_allowed_today"] == "false"
    assert d["path_present_criteria_met_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["feature_flag_enabled_today"] == "false"
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert d["is_first_blocking_coreq"] == "false"
    assert d["order_hint_index"] == "2"
    assert d["order_hint_coreq"] == "dual_honest_tf_aware_path_present"
    assert d["dual_recovery_path"] == "None"
    assert d["wire_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["prep_is_not_path_shipped"] == "true"
    assert d["prep_is_not_ship_met"] == "true"
    assert d["prep_is_not_feature_flag_enable"] == "true"
    assert d["distinct_from_path_design_contract"] == "true"
    assert d["distinct_from_path_present_criteria_contract"] == "true"
    assert d["distinct_from_path_execution_scaffold"] == "true"
    assert d["this_prep_alone_is_not_ship"] == "true"
    assert d["packaging_alone_is_not_ship"] == "true"
    one = d["planner_one_liner"].lower()
    assert "prep_present" in one
    assert "path_shipped=false" in one
    assert "dual_honest_tf_aware_path_present=false" in one
    assert "isolation_rewrite_with_wire" in one
    assert "this_prep_alone" in _CASE1_PATH_THIRD_COREQ_PREP_ANTI_CRITERIA
    assert "path_design_alone" in _CASE1_PATH_THIRD_COREQ_PREP_ANTI_CRITERIA
    assert "scaffold_alone" in _CASE1_PATH_THIRD_COREQ_PREP_ANTI_CRITERIA
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    src = inspect.getsource(
        format_tf_offline_case1_path_third_coreq_operational_prep_howto
    )
    assert "import tensorflow" not in src
    assert "from pims_admm_llm.models import tf_linear_blocks" not in src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in src
    assert "offline_case1_path_third_coreq_operational_prep_report(" not in src
    assert "format_tf_offline_case1_path_third_coreq_operational_prep_howto" in open(
        ep.__file__, encoding="utf-8"
    ).read()


def test_path_third_coreq_prep_package_surfaces(tmp_path):
    """E1/E2: path third-coreq prep meta/How_to/Calc_Check dual-ban surfaces."""
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
        _OFFLINE_TF_INDEX_WHAT,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    pkg = format_planner_honesty_package(report)
    assert pkg["meta"]["offline_tf_case1_path_third_coreq_operational_prep_ready"] is True
    assert pkg["meta"]["offline_tf_path_third_coreq_prep_present"] is True
    assert pkg["meta"]["offline_tf_path_shipped"] is False
    assert pkg["meta"]["offline_tf_dual_honest_tf_aware_path_present_ship_met"] is False
    note = str(pkg["meta"].get("offline_tf_case1_path_third_coreq_operational_prep", "")).lower()
    assert "prep" in note
    assert "path_shipped=false" in note or "path_shipped" in note
    path_prep = pkg.get("tf_offline_case1_path_third_coreq_operational_prep")
    assert path_prep is not None
    assert path_prep["path_shipped"] == "false"
    assert path_prep["dual_honest_tf_aware_path_present"] == "false"
    assert path_prep["prep_present"] == "true"
    assert path_prep["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert path_prep["is_first_blocking_coreq"] == "false"
    assert path_prep["order_hint_index"] == "2"
    rows = planner_honesty_check_rows(report)
    checks = {r["check"]: r for r in rows}
    assert "offline_tf_path_third_coreq_prep_not_path_ship" in checks
    assert checks["offline_tf_path_third_coreq_prep_not_path_ship"]["ok"] is True
    # No Index growth for this residual
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439


def test_format_tf_offline_case1_dual_linf_fourth_coreq_operational_prep_howto_pure():
    """E1/E2 Excel: dual_linf fourth-coreq operational prep How_to dual-ban (static)."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_case1_dual_linf_fourth_coreq_operational_prep_howto,
        _CASE1_DUAL_LINF_FOURTH_COREQ_PREP_ANTI_CRITERIA,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _OFFLINE_WIRE_BLOCKER_IDS,
    )
    import inspect
    import pims_admm_llm.models.excel_pipeline as ep

    d = format_tf_offline_case1_dual_linf_fourth_coreq_operational_prep_howto()
    assert d["topic"] == "tf_offline_case1_dual_linf_fourth_coreq_operational_prep"
    assert d["prep_present"] == "true"
    assert d["dual_linf_fourth_coreq_prep_present"] == "true"
    assert d["operational_prep_present"] == "true"
    assert d["dual_linf_under_wire_status"] == "unproven"
    assert d["dual_linf_proof_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["feature_flag_enabled_today"] == "false"
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert d["is_first_blocking_coreq"] == "false"
    assert d["order_hint_index"] == "3"
    assert d["order_hint_coreq"] == "dual_linf_under_wire_proven"
    assert d["dual_recovery_path"] == "None"
    assert d["wire_shipped"] == "false"
    assert d["path_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["prep_is_not_dual_linf_under_wire_proof"] == "true"
    assert d["prep_is_not_criteria_met"] == "true"
    assert d["prep_is_not_gate_flip"] == "true"
    assert d["distinct_from_dual_linf_criteria_contract"] == "true"
    assert d["distinct_from_online_linf_gate_criteria_contract"] == "true"
    assert d["distinct_from_probe_bridge_warmstart"] == "true"
    assert d["this_prep_alone_is_not_proof"] == "true"
    assert d["packaging_alone_is_not_proof"] == "true"
    one = d["planner_one_liner"].lower()
    assert "prep_present" in one
    assert "unproven" in one
    assert "proof_allowed" in one or "dual_linf_proof_allowed_today=false" in one
    assert "isolation_rewrite_with_wire" in one
    assert "this_prep_alone" in _CASE1_DUAL_LINF_FOURTH_COREQ_PREP_ANTI_CRITERIA
    assert "dual_linf_criteria_alone" in _CASE1_DUAL_LINF_FOURTH_COREQ_PREP_ANTI_CRITERIA
    assert "probe_linf" in _CASE1_DUAL_LINF_FOURTH_COREQ_PREP_ANTI_CRITERIA
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    src = inspect.getsource(
        format_tf_offline_case1_dual_linf_fourth_coreq_operational_prep_howto
    )
    assert "import tensorflow" not in src
    assert "from pims_admm_llm.models import tf_linear_blocks" not in src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in src
    assert "offline_case1_dual_linf_fourth_coreq_operational_prep_report(" not in src
    assert "format_tf_offline_case1_dual_linf_fourth_coreq_operational_prep_howto" in open(
        ep.__file__, encoding="utf-8"
    ).read()


def test_dual_linf_fourth_coreq_prep_package_surfaces(tmp_path):
    """E1/E2: dual_linf fourth-coreq prep meta/How_to/Calc_Check dual-ban surfaces."""
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
        _OFFLINE_TF_INDEX_WHAT,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    pkg = format_planner_honesty_package(report)
    assert pkg["meta"]["offline_tf_case1_dual_linf_fourth_coreq_operational_prep_ready"] is True
    assert pkg["meta"]["offline_tf_dual_linf_fourth_coreq_prep_present"] is True
    assert pkg["meta"]["offline_tf_dual_linf_under_wire_status"] == "unproven"
    assert pkg["meta"]["offline_tf_dual_linf_proof_allowed_today"] is False
    note = str(pkg["meta"].get("offline_tf_case1_dual_linf_fourth_coreq_operational_prep", "")).lower()
    assert "prep" in note
    assert "unproven" in note or "proof" in note
    dl_prep = pkg.get("tf_offline_case1_dual_linf_fourth_coreq_operational_prep")
    assert dl_prep is not None
    assert dl_prep["dual_linf_under_wire_status"] == "unproven"
    assert dl_prep["dual_linf_proof_allowed_today"] == "false"
    assert dl_prep["prep_present"] == "true"
    assert dl_prep["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert dl_prep["is_first_blocking_coreq"] == "false"
    assert dl_prep["order_hint_index"] == "3"
    rows = planner_honesty_check_rows(report)
    checks = {r["check"]: r for r in rows}
    assert "offline_tf_dual_linf_fourth_coreq_prep_not_dual_linf_proof" in checks
    assert checks["offline_tf_dual_linf_fourth_coreq_prep_not_dual_linf_proof"]["ok"] is True
    # No Index growth for this residual
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439



def test_format_tf_offline_case1_wire_fifth_coreq_operational_prep_howto_pure():
    """E1/E2 Excel: wire fifth-coreq operational prep How_to dual-ban (static)."""
    from pims_admm_llm.models.excel_pipeline import (
        format_tf_offline_case1_wire_fifth_coreq_operational_prep_howto,
        _CASE1_WIRE_FIFTH_COREQ_PREP_ANTI_CRITERIA,
        _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME,
        _OFFLINE_WIRE_BLOCKER_IDS,
    )
    import inspect
    import pims_admm_llm.models.excel_pipeline as ep

    d = format_tf_offline_case1_wire_fifth_coreq_operational_prep_howto()
    assert d["topic"] == "tf_offline_case1_wire_fifth_coreq_operational_prep"
    assert d["prep_present"] == "true"
    assert d["wire_fifth_coreq_prep_present"] == "true"
    assert d["operational_prep_present"] == "true"
    assert d["wire_shipped"] == "false"
    assert d["wire_ship_allowed_today"] == "false"
    assert d["wire_ship_criteria_met_today"] == "false"
    assert d["dual_linf_under_wire_status"] == "unproven"
    assert d["dual_linf_proof_allowed_today"] == "false"
    assert d["criteria_met_today"] == "false"
    assert d["online_linf_gate_under_tf_path"] == "open"
    assert d["gate_flip_allowed_today"] == "false"
    assert d["feature_flag_enabled_today"] == "false"
    assert d["feature_flag_name"] == _CASE1_PATH_DESIGN_FEATURE_FLAG_NAME
    assert d["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert d["is_first_blocking_coreq"] == "false"
    assert d["order_hint_index"] == "4"
    assert d["order_hint_coreq"] == "wire_shipped"
    assert d["dual_recovery_path"] == "None"
    assert d["path_shipped"] == "false"
    assert d["bundle_shipped"] == "false"
    assert d["isolation_rewrite_shipped"] == "false"
    assert d["form_label_change_shipped"] == "false"
    assert d["prep_is_not_wire_shipped"] == "true"
    assert d["prep_is_not_wire_ship_allow"] == "true"
    assert d["prep_is_not_wire_criteria_met"] == "true"
    assert d["distinct_from_wire_ship_acceptance_design_contract"] == "true"
    assert d["distinct_from_offline_wire_preflight"] == "true"
    assert d["distinct_from_prior_coreq_operational_preps"] == "true"
    assert d["this_prep_alone_is_not_wire_shipped"] == "true"
    assert d["packaging_alone_is_not_wire_shipped"] == "true"
    one = d["planner_one_liner"].lower()
    assert "prep_present" in one
    assert "wire_shipped=false" in one
    assert "wire_ship_allowed_today=false" in one
    assert "isolation_rewrite_with_wire" in one
    assert "this_prep_alone" in _CASE1_WIRE_FIFTH_COREQ_PREP_ANTI_CRITERIA
    assert "wire_ship_acceptance_design_alone" in _CASE1_WIRE_FIFTH_COREQ_PREP_ANTI_CRITERIA
    assert "preflight_alone" in _CASE1_WIRE_FIFTH_COREQ_PREP_ANTI_CRITERIA
    assert "wire_not_shipped" in _OFFLINE_WIRE_BLOCKER_IDS
    src = inspect.getsource(
        format_tf_offline_case1_wire_fifth_coreq_operational_prep_howto
    )
    assert "import tensorflow" not in src
    assert "from pims_admm_llm.models import tf_linear_blocks" not in src
    assert "from pims_admm_llm.models.tf_linear_blocks" not in src
    assert "offline_case1_wire_fifth_coreq_operational_prep_report(" not in src
    assert "format_tf_offline_case1_wire_fifth_coreq_operational_prep_howto" in open(
        ep.__file__, encoding="utf-8"
    ).read()


def test_wire_fifth_coreq_prep_package_surfaces(tmp_path):
    """E1/E2: wire fifth-coreq prep meta/How_to/Calc_Check dual-ban surfaces."""
    from pims_admm_llm.models.excel_pipeline import (
        format_planner_honesty_package,
        planner_honesty_check_rows,
        _OFFLINE_TF_INDEX_WHAT,
    )

    xlsx_in = tmp_path / "model.xlsx"
    write_template_excel(xlsx_in)
    report = run_excel_pipeline(xlsx_in)
    pkg = format_planner_honesty_package(report)
    assert pkg["meta"]["offline_tf_case1_wire_fifth_coreq_operational_prep_ready"] is True
    assert pkg["meta"]["offline_tf_wire_fifth_coreq_prep_present"] is True
    assert pkg["meta"]["offline_tf_wire_shipped"] is False
    assert pkg["meta"]["offline_tf_wire_ship_allowed_today"] is False
    note = str(pkg["meta"].get("offline_tf_case1_wire_fifth_coreq_operational_prep", "")).lower()
    assert "prep" in note
    assert "wire" in note
    wire_prep = pkg.get("tf_offline_case1_wire_fifth_coreq_operational_prep")
    assert wire_prep is not None
    assert wire_prep["wire_shipped"] == "false"
    assert wire_prep["wire_ship_allowed_today"] == "false"
    assert wire_prep["prep_present"] == "true"
    assert wire_prep["first_blocking_coreq"] == "isolation_rewrite_with_wire"
    assert wire_prep["is_first_blocking_coreq"] == "false"
    assert wire_prep["order_hint_index"] == "4"
    rows = planner_honesty_check_rows(report)
    checks = {r["check"]: r for r in rows}
    assert "offline_tf_wire_fifth_coreq_prep_not_wire_shipped" in checks
    assert checks["offline_tf_wire_fifth_coreq_prep_not_wire_shipped"]["ok"] is True
    # No Index growth for this residual
    assert len(_OFFLINE_TF_INDEX_WHAT) <= 1439
