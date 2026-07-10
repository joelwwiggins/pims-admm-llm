"""Import vendor crude assay Excel files (ExxonMobil / BP) → detailed cut JSON.

Supported layouts
-----------------
* **ExxonMobil / EMTEC** summary sheet (e.g. WTI Light Export, Cold Lake PDF twin):
  rows Start/End °C, Yield % vol/wt, Density, API, Sulfur, N, RON, PNA, MCR, metals.
* **BP** Summary sheet (e.g. Basrah Medium, Mars):
  product-style cuts (Lt Naph, Hvy Naph, Kero, LGO, HGO, LVGO, HVGO, VacRes)
  mapped into the same TBP cut list used by assay_swing.

Output matches ``data/assays/cold_lake_blend_clkbl23b.json`` so
``import_detailed_assay_json`` / cut-point CDU work unchanged.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

PathLike = Union[str, Path]


def _num(x: Any, default: float = 0.0) -> float:
    if x is None or x == "" or x == "-":
        return default
    try:
        return float(x)
    except (TypeError, ValueError):
        return default


def _api_to_sg(api: float) -> float:
    return 141.5 / (api + 131.5) if api > -131 else 0.9


# ---------------------------------------------------------------------------
# ExxonMobil layout
# ---------------------------------------------------------------------------

def parse_exxon_summary_xlsx(path: PathLike) -> Dict[str, Any]:
    """Parse EMTEC-style Summary (C) sheet → assay package dict."""
    from openpyxl import load_workbook

    path = Path(path)
    wb = load_workbook(path, data_only=True)
    # prefer Summary sheet
    name = next((n for n in wb.sheetnames if "summary" in n.lower()), wb.sheetnames[0])
    ws = wb[name]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    def find_row(label: str) -> Optional[List[Any]]:
        lab = label.lower()
        for r in rows:
            for c in r[:4]:
                if c is not None and lab in str(c).lower():
                    return r
        return None

    def prop_value(label: str) -> float:
        """Whole-crude property: label in col ~12, value nearby."""
        lab = label.lower()
        for r in rows:
            for i, c in enumerate(r):
                if c is not None and lab in str(c).lower().replace("°", ""):
                    # value often at i+3 or last numeric in row after i
                    for j in range(i + 1, min(i + 6, len(r))):
                        if isinstance(r[j], (int, float)) and not isinstance(r[j], bool):
                            return float(r[j])
        return 0.0

    ref_row = find_row("Reference:")
    name_row = find_row("Name:")
    origin_row = find_row("Origin:")
    date_row = find_row("Assay Date:")

    reference = ""
    crude_name = path.stem
    origin = ""
    assay_date = ""
    if ref_row:
        for c in ref_row[1:6]:
            if c and str(c).strip() and "reference" not in str(c).lower():
                reference = str(c).strip()
                break
    if name_row:
        for c in name_row[1:6]:
            if c and str(c).strip() and "name" not in str(c).lower():
                crude_name = str(c).strip()
                break
    if origin_row:
        for c in origin_row[1:6]:
            if c and str(c).strip() and "origin" not in str(c).lower():
                origin = str(c).strip()
                break
    if date_row:
        for c in date_row[1:6]:
            if c is not None and "assay" not in str(c).lower():
                assay_date = str(c)[:10]
                break

    # Cut grid: find Start (°C) row
    start_r = end_r = None
    for i, r in enumerate(rows):
        if r[1] and "start" in str(r[1]).lower() and "°c" in str(r[1]).lower().replace("c", "c"):
            start_r = i
        if r[1] and str(r[1]).lower().startswith("end") and "°c" in str(r[1]).lower().replace("c", "c"):
            end_r = i
    if start_r is None or end_r is None:
        raise ValueError(f"Exxon cut Start/End rows not found in {path}")

    start_row = rows[start_r]
    end_row = rows[end_r]

    def row_by_label(label: str) -> Optional[List[Any]]:
        lab = label.lower()
        for r in rows:
            if r[1] is not None and lab in str(r[1]).lower():
                return r
        return None

    y_vol = row_by_label("Yield (% vol)")
    y_wt = row_by_label("Yield (% wt)")
    dens = row_by_label("Density @ 15°C")
    api_r = row_by_label("API Gravity")
    sul = row_by_label("Total Sulfur")
    ntot = row_by_label("Total Nitrogen")
    ron = row_by_label("RON")
    par = row_by_label("Paraffins")
    nap = row_by_label("Naphthenes")
    aro = row_by_label("Aromatics")
    mcr = row_by_label("Micro Carbon Residue")
    ni = row_by_label("Nickel")
    v = row_by_label("Vanadium")
    asp = row_by_label("C7 Asphaltenes")

    # Columns: skip whole-crude col2; take atmospheric narrow cuts C5..350-370 and vac 370-450..550-FBP
    # Typical indices 4..11 atm, 13..16 vac (skip col12 370-FBP bulk which double-counts vac)
    cuts: List[Dict[str, Any]] = []
    # detect which cols are narrow cuts
    for col in range(3, min(len(start_row), 20)):
        s = start_row[col]
        e = end_row[col]
        if s is None or e is None:
            continue
        # skip pure lights C4-only if yield tiny and end is C4
        try:
            if isinstance(s, str) and s.upper() in ("IBP",):
                if isinstance(e, str) and str(e).upper() in ("C4", "FBP"):
                    # IBP-C4 lights: optional include as c4_lights
                    if str(e).upper() == "C4":
                        ts, te = 0.0, 5.0
                    else:
                        continue
                else:
                    continue
            elif isinstance(s, str) and s.upper() == "C5":
                ts, te = 5.0, float(e) if not isinstance(e, str) else 65.0
            else:
                ts = float(s) if not isinstance(s, str) else 0.0
                if isinstance(e, str) and e.upper() == "FBP":
                    # 370-FBP bulk atm double-counts vac slices — skip only ~370 bulk
                    if 360 <= ts <= 400:
                        continue
                    te = 750.0
                else:
                    te = float(e)
        except (TypeError, ValueError):
            continue

        yv = _num(y_vol[col] if y_vol else 0) / 100.0  # sheet is %
        yw = _num(y_wt[col] if y_wt else 0) / 100.0
        if yv < 1e-6 and yw < 1e-6:
            continue
        api = _num(api_r[col] if api_r else 30, 30.0)
        d15 = _num(dens[col] if dens else 0) or _api_to_sg(api)
        cut = {
            "id": f"{int(ts)}_{int(te) if te < 700 else 'fbp'}",
            "tbp_start_c": ts,
            "tbp_end_c": te,
            "yield_vol": yv,
            "yield_wt": yw,
            "api": api,
            "density_15c_g_cc": d15,
            "sulfur_wt": _num(sul[col] if sul else 0),
            "nitrogen_ppm": _num(ntot[col] if ntot else 0),
            "ccr_wt": _num(mcr[col] if mcr else 0),
            "ron": _num(ron[col] if ron else 0),
            "paraffins_vol": _num(par[col] if par else 0) / 100.0 if par and _num(par[col]) > 1 else _num(par[col] if par else 0.33),
            "naphthenes_vol": _num(nap[col] if nap else 0) / 100.0 if nap and _num(nap[col]) > 1 else _num(nap[col] if nap else 0.33),
            "aromatics_vol": _num(aro[col] if aro else 0) / 100.0 if aro and _num(aro[col]) > 1 else _num(aro[col] if aro else 0.34),
            "nickel_ppm": _num(ni[col] if ni else 0),
            "vanadium_ppm": _num(v[col] if v else 0),
            "asphaltenes_wt": _num(asp[col] if asp else 0),
        }
        # PNA as wt% on sheet → store as vol proxy (planning)
        if par and _num(par[col]) > 1:
            p_ = _num(par[col]) / 100.0
            n_ = _num(nap[col]) / 100.0 if nap else 0.0
            a_ = _num(aro[col]) / 100.0 if aro else 0.0
            s = p_ + n_ + a_ or 1.0
            cut["paraffins_vol"] = p_ / s
            cut["naphthenes_vol"] = n_ / s
            cut["aromatics_vol"] = a_ / s
        cuts.append(cut)

    # renorm vol
    s = sum(c["yield_vol"] for c in cuts) or 1.0
    for c in cuts:
        c["yield_vol"] = c["yield_vol"] / s
    if sum(c["yield_wt"] for c in cuts) > 0:
        sw = sum(c["yield_wt"] for c in cuts)
        for c in cuts:
            c["yield_wt"] = c["yield_wt"] / sw

    api_w = prop_value("API Gravity") or _num(api_r[2] if api_r else 0, 40.0)
    whole = {
        "api": api_w,
        "density_15c_g_cc": prop_value("Density @ 15") or _api_to_sg(api_w),
        "sulfur_wt": prop_value("Total Sulfur") or _num(sul[2] if sul else 0),
        "ccr_wt": prop_value("Micro Carbon Residue") or _num(mcr[2] if mcr else 0),
        "nitrogen_ppm": prop_value("Total Nitrogen") or _num(ntot[2] if ntot else 0),
        "nickel_ppm": prop_value("Nickel") or _num(ni[2] if ni else 0),
        "vanadium_ppm": prop_value("Vanadium") or _num(v[2] if v else 0),
        "metals_ni_v_ppm": (prop_value("Nickel") or 0) + (prop_value("Vanadium") or 0),
        "tan_mgkoh_g": prop_value("Total Acid Number") or prop_value("Acidity"),
        "viscosity_cst_20c": prop_value("Viscosity @ 20"),
        "pour_point_c": prop_value("Pour Point"),
        "asphaltenes_wt": prop_value("C7 Asphaltenes") or _num(asp[2] if asp else 0),
        "paraffins_vol": _num(par[2] if par else 0.33) / (100.0 if par and _num(par[2]) > 1 else 1.0),
        "naphthenes_vol": _num(nap[2] if nap else 0.33) / (100.0 if nap and _num(nap[2]) > 1 else 1.0),
        "aromatics_vol": _num(aro[2] if aro else 0.34) / (100.0 if aro and _num(aro[2]) > 1 else 1.0),
        "price_usd_per_bbl": 78.0 if api_w > 40 else 70.0,
        "max_supply_kbd": 100.0,
    }
    # fix PNA whole if wt%
    if whole["paraffins_vol"] + whole["naphthenes_vol"] + whole["aromatics_vol"] > 1.5:
        p_, n_, a_ = whole["paraffins_vol"] / 100, whole["naphthenes_vol"] / 100, whole["aromatics_vol"] / 100
        s = p_ + n_ + a_ or 1
        whole["paraffins_vol"], whole["naphthenes_vol"], whole["aromatics_vol"] = p_ / s, n_ / s, a_ / s

    return {
        "meta": {
            "reference": reference,
            "name": crude_name,
            "origin": origin,
            "assay_date": assay_date,
            "source": f"ExxonMobil assay file {path.name}",
            "comments": "Imported for cut-point CDU heart/swing library",
            "vendor": "ExxonMobil/EMTEC",
        },
        "whole_crude": whole,
        "cuts": cuts,
        "default_cut_points_c": {"naphtha_ep": 200.0, "distillate_ep": 370.0, "gasoil_ep": 550.0},
        "product_map": {
            "cdu_naphtha": {"tbp_lo_c": 5, "tbp_hi_c": 200},
            "cdu_distillate": {"tbp_lo_c": 200, "tbp_hi_c": 370},
            "cdu_gasoil": {"tbp_lo_c": 370, "tbp_hi_c": 550},
            "cdu_resid": {"tbp_lo_c": 550, "tbp_hi_c": 750},
        },
    }


# ---------------------------------------------------------------------------
# BP layout
# ---------------------------------------------------------------------------

# Map BP product columns (by header fragment) → TBP window for our library
_BP_CUT_MAP = [
    # (header_substr, id, t_lo, t_hi)
    ("light naphtha", "c5_95", 5.0, 95.0),
    ("heavy naphtha", "95_149", 95.0, 149.0),
    # sometimes split HN further — BP has 149-175 as empty name col
    ("kero", "175_232", 175.0, 232.0),
    ("light gas oil", "232_342", 232.0, 342.0),
    ("heavy gas oil", "342_369", 342.0, 369.0),
    ("light vacuum gas oil", "369_509", 369.0, 509.0),
    ("heavy vacuum gas oil", "509_550", 509.0, 550.0),
    ("vacres", "550_fbp", 550.0, 750.0),
]


def parse_bp_summary_xls(path: PathLike) -> Dict[str, Any]:
    """Parse BP Summary sheet → assay package dict."""
    import xlrd

    path = Path(path)
    book = xlrd.open_workbook(str(path))
    sh = book.sheet_by_name("Summary") if "Summary" in book.sheet_names() else book.sheet_by_index(0)

    def cell(r: int, c: int) -> Any:
        if r < sh.nrows and c < sh.ncols:
            return sh.cell_value(r, c)
        return None

    # header names row ~27
    headers = [str(cell(27, c) or "").strip().lower() for c in range(sh.ncols)]
    # start/end rows 29-30
    starts = [cell(29, c) for c in range(sh.ncols)]
    ends = [cell(30, c) for c in range(sh.ncols)]

    def row_label(label: str) -> Optional[int]:
        lab = label.lower()
        for r in range(sh.nrows):
            v = cell(r, 1)
            if v is not None and lab in str(v).lower():
                return r
        return None

    y_vol_r = row_label("Yield on crude (% vol)")
    y_wt_r = row_label("Yield on crude (% wt)")
    dens_r = row_label("Density at 15")
    sul_r = row_label("Total Sulphur")
    n_r = row_label("Total Nitrogen")
    par_r = row_label("Paraffins")
    nap_r = row_label("Naphthenes")
    aro_r = row_label("Aromatics")
    ron_r = row_label("Research Octane")

    # whole crude
    name = str(cell(10, 2) or path.stem)
    reference = str(cell(8, 2) or "")
    origin = str(cell(12, 2) or "")
    api = _num(cell(8, 14))
    sul_w = _num(cell(9, 14))
    dens_w = _num(cell(35, 2)) if dens_r else _api_to_sg(api)

    # Build cuts from product columns 3..
    cuts: List[Dict[str, Any]] = []
    for c in range(3, sh.ncols):
        hdr = headers[c] if c < len(headers) else ""
        # also use start/end
        ts = starts[c] if c < len(starts) else None
        te = ends[c] if c < len(ends) else None
        # skip AtRes intermediate residues that double-count
        if "atres" in hdr or hdr in ("", " "):
            # try identify by TBP
            if isinstance(ts, (int, float)) and isinstance(te, str) and te.upper() == "FBP":
                if float(ts) < 500:
                    continue  # atm residue bulk
            if not hdr.strip():
                # unnamed mid cut 149-175
                if isinstance(ts, (int, float)) and isinstance(te, (int, float)):
                    cid = f"{int(ts)}_{int(te)}"
                    t_lo, t_hi = float(ts), float(te)
                else:
                    continue
            else:
                continue
        else:
            t_lo = t_hi = None
            cid = None
            for key, cid0, lo, hi in _BP_CUT_MAP:
                if key in hdr:
                    cid, t_lo, t_hi = cid0, lo, hi
                    break
            if cid is None:
                if isinstance(ts, (int, float)) and isinstance(te, (int, float)):
                    t_lo, t_hi = float(ts), float(te)
                    cid = f"{int(t_lo)}_{int(t_hi)}"
                elif isinstance(ts, (int, float)) and isinstance(te, str) and te.upper() == "FBP":
                    t_lo, t_hi = float(ts), 750.0
                    cid = f"{int(t_lo)}_fbp"
                else:
                    continue
            # prefer numeric start/end when present
            if isinstance(ts, (int, float)):
                t_lo = float(ts)
            if isinstance(te, (int, float)):
                t_hi = float(te)
            elif isinstance(te, str) and te.upper() == "FBP":
                t_hi = 750.0

        yv = _num(cell(y_vol_r, c) if y_vol_r is not None else 0) / 100.0
        yw = _num(cell(y_wt_r, c) if y_wt_r is not None else 0) / 100.0
        if yv < 1e-6:
            continue
        d15 = _num(cell(dens_r, c) if dens_r is not None else 0)
        # BP density kg/litre = g/cc
        api_c = (141.5 / d15 - 131.5) if d15 > 0.5 else 30.0
        cut = {
            "id": cid or f"cut_{c}",
            "tbp_start_c": t_lo or 0.0,
            "tbp_end_c": t_hi or 100.0,
            "yield_vol": yv,
            "yield_wt": yw,
            "api": api_c,
            "density_15c_g_cc": d15 or _api_to_sg(api_c),
            "sulfur_wt": _num(cell(sul_r, c) if sul_r is not None else 0),
            "nitrogen_ppm": _num(cell(n_r, c) if n_r is not None else 0),
            "ccr_wt": 0.0,
            "ron": _num(cell(ron_r, c) if ron_r is not None else 0),
            "paraffins_vol": 0.33,
            "naphthenes_vol": 0.33,
            "aromatics_vol": 0.34,
            "nickel_ppm": 0.0,
            "vanadium_ppm": 0.0,
            "asphaltenes_wt": 0.0,
        }
        if par_r is not None:
            p_ = _num(cell(par_r, c))
            n_ = _num(cell(nap_r, c) if nap_r else 0)
            a_ = _num(cell(aro_r, c) if aro_r else 0)
            if p_ + n_ + a_ > 1:
                p_, n_, a_ = p_ / 100, n_ / 100, a_ / 100
            s = p_ + n_ + a_
            if s > 0.1:
                cut["paraffins_vol"] = p_ / s
                cut["naphthenes_vol"] = n_ / s
                cut["aromatics_vol"] = a_ / s
        # heavy resid CCR estimate from sulfur/API if missing
        if cut["tbp_start_c"] >= 500:
            cut["ccr_wt"] = max(5.0, 0.5 * cut["sulfur_wt"] * 3)
        cuts.append(cut)

    s = sum(c["yield_vol"] for c in cuts) or 1.0
    for c in cuts:
        c["yield_vol"] = c["yield_vol"] / s

    whole = {
        "api": api,
        "density_15c_g_cc": dens_w or _api_to_sg(api),
        "sulfur_wt": sul_w,
        "ccr_wt": 6.0 if api < 30 else 3.0,
        "nitrogen_ppm": _num(cell(n_r, 2) if n_r else 1400),
        "nickel_ppm": 15.0 if api < 30 else 5.0,
        "vanadium_ppm": 40.0 if api < 30 else 10.0,
        "metals_ni_v_ppm": 55.0 if api < 30 else 15.0,
        "tan_mgkoh_g": _num(cell(row_label("Acidity") or 0, 2) if row_label("Acidity") else 0.16),
        "viscosity_cst_20c": _num(cell(row_label("Viscosity at 20") or 0, 2) if row_label("Viscosity at 20") else 0),
        "paraffins_vol": 0.30,
        "naphthenes_vol": 0.30,
        "aromatics_vol": 0.40,
        "price_usd_per_bbl": 68.0 if api < 32 else 74.0,
        "max_supply_kbd": 80.0,
    }

    return {
        "meta": {
            "reference": reference,
            "name": name,
            "origin": origin,
            "assay_date": "",
            "source": f"BP assay file {path.name}",
            "comments": "Imported from BP public crude assays; cut points drive CDU",
            "vendor": "BP",
            "source_url": "https://www.bp.com/en/global/bp-supply-trading-and-shipping/documents-and-downloads/technical-downloads/crudes-assays.html",
        },
        "whole_crude": whole,
        "cuts": cuts,
        "default_cut_points_c": {"naphtha_ep": 175.0, "distillate_ep": 369.0, "gasoil_ep": 550.0},
        "product_map": {
            "cdu_naphtha": {"tbp_lo_c": 5, "tbp_hi_c": 175},
            "cdu_distillate": {"tbp_lo_c": 175, "tbp_hi_c": 369},
            "cdu_gasoil": {"tbp_lo_c": 369, "tbp_hi_c": 550},
            "cdu_resid": {"tbp_lo_c": 550, "tbp_hi_c": 750},
        },
    }


def import_vendor_file(path: PathLike) -> Dict[str, Any]:
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        return parse_exxon_summary_xlsx(path)
    if path.suffix.lower() == ".xls":
        return parse_bp_summary_xls(path)
    raise ValueError(f"unsupported assay file type: {path}")


def write_assay_json(pkg: Dict[str, Any], out_path: PathLike) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(pkg, indent=2) + "\n", encoding="utf-8")
    return out_path
