"""JSON + Excel (PIMS-export-shaped) assay loaders.

Property-rich crude / intermediate assays drive unit yields. Legacy
``load_crude_data`` consumers stay on synthetic_crudes.json by default;
use ``assays_to_refinery_data`` or ``load_crude_data(path=assays_json)``
for the dual path.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from .properties import FeedProperties, crude_to_props

PathLike = Union[str, Path]

# Map full-plant CDU stream names → classic toy intermediate names
_CLASSIC_YIELD_MAP = {
    "cdu_naphtha": "naphtha",
    "cdu_distillate": "distillate",
    "cdu_gasoil": "gasoil",
    "cdu_resid": "residue",
    "naphtha": "naphtha",
    "distillate": "distillate",
    "gasoil": "gasoil",
    "residue": "residue",
}


def _repo_roots(here: Path) -> List[Path]:
    return [
        here.parents[3],
        here.parents[2],
        Path.cwd(),
        Path("/home/joel/projects/pims-admm-llm"),
    ]


def _find_rel(rel: Path) -> Path:
    here = Path(__file__).resolve()
    for root in _repo_roots(here):
        p = root / rel
        if p.is_file():
            return p
    return _repo_roots(here)[0] / rel


def default_assays_path() -> Path:
    return _find_rel(Path("data") / "assays" / "crudes.json")


def default_intermediates_path() -> Path:
    return _find_rel(Path("data") / "assays" / "intermediates.json")


def default_routing_path() -> Path:
    return _find_rel(Path("data") / "routing.json")


def load_json(path: PathLike) -> Dict[str, Any]:
    path = Path(path)
    with path.open() as f:
        return json.load(f)


def load_assays_json(path: PathLike | None = None) -> Dict[str, Any]:
    """Load crude assay package (crudes + products + capacities + tanks)."""
    return load_json(path or default_assays_path())


def load_intermediates_json(path: PathLike | None = None) -> Dict[str, Any]:
    """Load intermediate / unit-feed assay package."""
    return load_json(path or default_intermediates_path())


def load_routing(path: PathLike | None = None) -> Dict[str, Any]:
    return load_json(path or default_routing_path())


def is_assay_shaped(raw: Dict[str, Any]) -> bool:
    """True if JSON looks like Wave2 assays (TBP / sulfur_wt) rather than legacy yields-only."""
    crudes = raw.get("crudes") or []
    if not crudes:
        return False
    c0 = crudes[0]
    if "tbp_cut_vol" in c0 or "sulfur_wt" in c0 or "ccr_wt" in c0:
        return True
    if "yields" in c0 and "sulfur_wt_pct" in c0 and "tbp_cut_vol" not in c0:
        return False
    return "yields" not in c0


def intermediate_properties_list(inter_pkg: Dict[str, Any] | None = None) -> List[FeedProperties]:
    inter_pkg = inter_pkg or load_intermediates_json()
    out: List[FeedProperties] = []
    for row in inter_pkg.get("intermediates") or []:
        out.append(crude_to_props(row))
    return out


def intermediate_props_by_name(
    inter_pkg: Dict[str, Any] | None = None,
) -> Dict[str, FeedProperties]:
    return {p.name: p for p in intermediate_properties_list(inter_pkg)}


def crude_properties_list(assays: Dict[str, Any] | None = None) -> List[FeedProperties]:
    assays = assays or load_assays_json()
    return [crude_to_props(c) for c in assays.get("crudes", [])]


def cdu_yields_classic(crude: Dict[str, Any]) -> Dict[str, float]:
    """Property/TBP-driven CDU yields mapped to classic intermediate names."""
    from .yields import cdu_yields_from_assay

    props = crude_to_props(crude)
    y_full = cdu_yields_from_assay(props, crude.get("tbp_cut_vol"))
    classic: Dict[str, float] = {
        "naphtha": 0.0,
        "distillate": 0.0,
        "gasoil": 0.0,
        "residue": 0.0,
    }
    for k, v in y_full.items():
        classic[_CLASSIC_YIELD_MAP.get(k, k)] = classic.get(
            _CLASSIC_YIELD_MAP.get(k, k), 0.0
        ) + float(v)
    # keep only classic four
    classic = {k: float(classic.get(k, 0.0)) for k in ("naphtha", "distillate", "gasoil", "residue")}
    s = sum(classic.values()) or 1.0
    return {k: v / s for k, v in classic.items()}


def assays_to_refinery_data(
    assays: Dict[str, Any] | None = None,
    *,
    classic_intermediates: bool = True,
) -> "RefineryData":
    """Convert assay-shaped package → RefineryData for legacy mono/ADMM blocks.

    When classic_intermediates=True (default), CDU yields are mapped to
    naphtha/distillate/gasoil/residue so existing toy LP consumers work.
    """
    from .data import (
        CrudeAssay,
        InventorySpec,
        ProductSpec,
        RefineryData,
        UtilitySpec,
    )

    assays = assays or load_assays_json()
    intermediates = ["naphtha", "distillate", "gasoil", "residue"]

    crudes: List[CrudeAssay] = []
    for c in assays.get("crudes", []):
        if classic_intermediates:
            if "yields" in c and not c.get("tbp_cut_vol"):
                yields = {k: float(v) for k, v in c["yields"].items()}
            else:
                yields = cdu_yields_classic(c)
        else:
            yields = {k: float(v) for k, v in (c.get("yields") or {}).items()}
        for i in intermediates:
            yields.setdefault(i, 0.0)
        crudes.append(
            CrudeAssay(
                name=str(c["name"]),
                price_usd_per_bbl=float(c.get("price_usd_per_bbl") or c.get("price") or 70),
                max_supply_kbd=float(c.get("max_supply_kbd") or c.get("max_supply") or 50),
                yields=yields,
                api=float(c.get("api", 0.0)),
                sulfur_wt_pct=float(
                    c.get("sulfur_wt_pct", c.get("sulfur_wt", 0.0))
                ),
                utility_use={
                    k: float(v) for k, v in (c.get("utility_use") or {}).items()
                }
                or {"fuel_gas": 0.015, "steam": 0.02, "power": 0.01},
            )
        )

    products: Dict[str, ProductSpec] = {}
    for name, p in (assays.get("products") or {}).items():
        products[name] = ProductSpec(
            name=name,
            price_usd_per_bbl=float(p.get("price_usd_per_bbl") or p.get("price") or 100),
            max_demand_kbd=float(p.get("max_demand_kbd") or p.get("max_demand") or 50),
            utility_use={
                k: float(v) for k, v in (p.get("utility_use") or {}).items()
            },
        )
    if not products:
        products = {
            "gasoline": ProductSpec("gasoline", 105.0, 90.0),
            "diesel": ProductSpec("diesel", 112.0, 80.0),
            "fuel_oil": ProductSpec("fuel_oil", 68.0, 70.0),
        }

    # Classic blend recipes if full-plant keys present
    blend_raw = assays.get("blend_recipes") or {}
    if blend_raw and any("naphtha" in r for r in blend_raw.values()):
        blend_recipes = {
            prod: {comp: float(f) for comp, f in recipe.items()}
            for prod, recipe in blend_raw.items()
        }
    else:
        blend_recipes = {
            "gasoline": {"naphtha": 0.85, "distillate": 0.15},
            "diesel": {"distillate": 0.70, "gasoil": 0.30},
            "fuel_oil": {"gasoil": 0.40, "residue": 0.60},
        }

    caps = assays.get("capacities") or {}
    cdu_cap = float(
        assays.get("cdu_capacity_kbd")
        or caps.get("cdu_kbd")
        or 120.0
    )

    inventory: Dict[str, InventorySpec] = {}
    inv_raw = assays.get("inventory") or {}
    if inv_raw:
        for stream, spec in inv_raw.items():
            inventory[stream] = InventorySpec(
                stream=stream,
                start_kbd=float(spec.get("start_kbd", 0.0)),
                capacity_kbd=float(spec.get("capacity_kbd", 1e9)),
                holding_cost_usd_per_bbl=float(
                    spec.get("holding_cost_usd_per_bbl", spec.get("holding_cost", 0.0))
                ),
            )
    else:
        # Map tanks → classic intermediates when possible
        tanks = assays.get("tanks") or {}
        stream_to_classic = {
            "cdu_gasoil": "gasoil",
            "cdu_resid": "residue",
            "cdu_naphtha": "naphtha",
            "cdu_distillate": "distillate",
        }
        for _tname, tspec in tanks.items():
            stream = str(tspec.get("stream", ""))
            classic = stream_to_classic.get(stream)
            if classic and classic not in inventory:
                inventory[classic] = InventorySpec(
                    stream=classic,
                    start_kbd=float(tspec.get("start_kbd", 0.0)),
                    capacity_kbd=float(tspec.get("capacity_kbd", 1e6)),
                    holding_cost_usd_per_bbl=float(
                        tspec.get("holding_cost", tspec.get("holding_cost_usd_per_bbl", 0.1))
                    ),
                )
        defaults = {
            "naphtha": (5.0, 40.0, 0.15),
            "distillate": (6.0, 45.0, 0.12),
            "gasoil": (4.0, 35.0, 0.10),
            "residue": (8.0, 50.0, 0.08),
        }
        for stream, (start, cap, hold) in defaults.items():
            inventory.setdefault(
                stream,
                InventorySpec(stream, start, cap, hold),
            )

    utilities: Dict[str, UtilitySpec] = {}
    util_raw = assays.get("utilities") or {
        "fuel_gas": {"capacity": 6.0, "cost_usd_per_unit": 18.0, "unit": "kbd_equiv"},
        "steam": {"capacity": 5.0, "cost_usd_per_unit": 12.0, "unit": "kbd_equiv"},
        "power": {"capacity": 7.0, "cost_usd_per_unit": 25.0, "unit": "mwh_per_day_scaled"},
    }
    for uname, uspec in util_raw.items():
        utilities[uname] = UtilitySpec(
            name=uname,
            capacity=float(uspec.get("capacity", 1e9)),
            cost_usd_per_unit=float(uspec.get("cost_usd_per_unit", 0.0)),
            unit=str(uspec.get("unit", "unit")),
        )
    utility_names = list(utilities.keys()) or ["fuel_gas", "steam", "power"]

    return RefineryData(
        crudes=crudes,
        products=products,
        cdu_capacity_kbd=cdu_cap,
        blend_recipes=blend_recipes,
        intermediates=intermediates,
        inventory=inventory,
        utilities=utilities,
        utility_names=utility_names,
    )


def load_assays_excel(path: PathLike) -> Dict[str, Any]:
    """Load PIMS-export-shaped Excel.

    Expected sheet 'Crudes' columns (case-insensitive):
      name, api, sulfur_wt|sulfur, ccr_wt|ccr, nitrogen_ppm|nitrogen,
      paraffins_vol|P, naphthenes_vol|N, aromatics_vol|A,
      price|price_usd_per_bbl, max_supply|max_supply_kbd,
      y_naphtha, y_distillate, y_gasoil, y_resid (optional TBP cuts)
    Optional sheet 'Products': name, price, max_demand
    Optional sheet 'Capacities': unit, capacity_kbd
    Optional sheet 'Intermediates': same property columns as crudes (+ stream)
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "openpyxl required for Excel assay load: pip install openpyxl"
        ) from e

    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Crudes" not in wb.sheetnames:
        raise ValueError(f"Excel {path} missing sheet 'Crudes'")

    def norm(h: Any) -> str:
        return str(h or "").strip().lower().replace(" ", "_")

    def sheet_dicts(sheet_name: str) -> List[Dict[str, Any]]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [norm(h) for h in rows[0]]
        out: List[Dict[str, Any]] = []
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            out.append(
                {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            )
        return out

    crudes: List[Dict[str, Any]] = []
    for d in sheet_dicts("Crudes"):
        name = d.get("name") or d.get("crude")
        if not name:
            continue
        tbp = {}
        if d.get("y_naphtha") is not None:
            tbp = {
                "naphtha_ibp_350f": float(d.get("y_naphtha") or 0),
                "distillate_350_650f": float(d.get("y_distillate") or 0),
                "gasoil_650_1050f": float(d.get("y_gasoil") or 0),
                "resid_1050f_plus": float(d.get("y_resid") or 0),
            }
        row = {
            "name": str(name),
            "api": float(d.get("api") or 30),
            "sulfur_wt": float(d.get("sulfur_wt") or d.get("sulfur") or 1.0),
            "ccr_wt": float(d.get("ccr_wt") or d.get("ccr") or 2.0),
            "nitrogen_ppm": float(d.get("nitrogen_ppm") or d.get("nitrogen") or 1000),
            "paraffins_vol": float(d.get("paraffins_vol") or d.get("p") or 0.33),
            "naphthenes_vol": float(d.get("naphthenes_vol") or d.get("n") or 0.33),
            "aromatics_vol": float(d.get("aromatics_vol") or d.get("a") or 0.34),
            "price_usd_per_bbl": float(
                d.get("price_usd_per_bbl") or d.get("price") or 70
            ),
            "max_supply_kbd": float(
                d.get("max_supply_kbd") or d.get("max_supply") or 50
            ),
            "utility_use": {"fuel_gas": 0.015, "steam": 0.02, "power": 0.01},
        }
        if tbp:
            row["tbp_cut_vol"] = tbp
        crudes.append(row)

    products = {
        "gasoline": {"price_usd_per_bbl": 105.0, "max_demand_kbd": 90.0},
        "diesel": {"price_usd_per_bbl": 112.0, "max_demand_kbd": 80.0},
        "fuel_oil": {"price_usd_per_bbl": 68.0, "max_demand_kbd": 70.0},
    }
    prod_rows = sheet_dicts("Products")
    if prod_rows:
        products = {}
        for dd in prod_rows:
            nm = str(dd.get("name"))
            products[nm] = {
                "price_usd_per_bbl": float(
                    dd.get("price") or dd.get("price_usd_per_bbl") or 100
                ),
                "max_demand_kbd": float(
                    dd.get("max_demand") or dd.get("max_demand_kbd") or 50
                ),
            }

    capacities = {
        "cdu_kbd": 140.0,
        "fcc_kbd": 55.0,
        "coker_kbd": 40.0,
        "reformer_kbd": 45.0,
    }
    if "Capacities" in wb.sheetnames:
        cws = wb["Capacities"]
        crows = list(cws.iter_rows(values_only=True))
        if crows:
            for row in crows[1:]:
                if not row or row[0] is None:
                    continue
                unit = str(row[0]).strip().lower()
                cap = float(row[1])
                if "cdu" in unit:
                    capacities["cdu_kbd"] = cap
                elif "fcc" in unit:
                    capacities["fcc_kbd"] = cap
                elif "coker" in unit:
                    capacities["coker_kbd"] = cap
                elif "reform" in unit:
                    capacities["reformer_kbd"] = cap

    intermediates: List[Dict[str, Any]] = []
    for d in sheet_dicts("Intermediates"):
        name = d.get("name") or d.get("stream")
        if not name:
            continue
        intermediates.append(
            {
                "name": str(name),
                "stream": str(d.get("stream") or name),
                "api": float(d.get("api") or 30),
                "sulfur_wt": float(d.get("sulfur_wt") or d.get("sulfur") or 1.0),
                "ccr_wt": float(d.get("ccr_wt") or d.get("ccr") or 2.0),
                "nitrogen_ppm": float(
                    d.get("nitrogen_ppm") or d.get("nitrogen") or 1000
                ),
                "paraffins_vol": float(d.get("paraffins_vol") or d.get("p") or 0.33),
                "naphthenes_vol": float(d.get("naphthenes_vol") or d.get("n") or 0.33),
                "aromatics_vol": float(d.get("aromatics_vol") or d.get("a") or 0.34),
                "price_usd_per_bbl": float(
                    d.get("price_usd_per_bbl") or d.get("price") or 70
                ),
            }
        )

    return {
        "meta": {"source": str(path), "format": "excel_pims_shaped"},
        "crudes": crudes,
        "products": products,
        "capacities": capacities,
        "tanks": {},
        "intermediates": intermediates,
    }


def write_template_excel(path: PathLike, *, include_intermediates: bool = True) -> Path:
    """Write a PIMS-shaped template workbook from JSON assays (+ optional intermediates)."""
    import openpyxl

    data = load_assays_json()
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crudes"
    headers = [
        "name",
        "api",
        "sulfur_wt",
        "ccr_wt",
        "nitrogen_ppm",
        "paraffins_vol",
        "naphthenes_vol",
        "aromatics_vol",
        "price",
        "max_supply",
        "y_naphtha",
        "y_distillate",
        "y_gasoil",
        "y_resid",
    ]
    ws.append(headers)
    for c in data["crudes"]:
        tbp = c.get("tbp_cut_vol") or {}
        ws.append(
            [
                c["name"],
                c["api"],
                c["sulfur_wt"],
                c["ccr_wt"],
                c["nitrogen_ppm"],
                c["paraffins_vol"],
                c["naphthenes_vol"],
                c["aromatics_vol"],
                c["price_usd_per_bbl"],
                c["max_supply_kbd"],
                tbp.get("naphtha_ibp_350f"),
                tbp.get("distillate_350_650f"),
                tbp.get("gasoil_650_1050f"),
                tbp.get("resid_1050f_plus"),
            ]
        )
    pws = wb.create_sheet("Products")
    pws.append(["name", "price", "max_demand"])
    for name, spec in data.get("products", {}).items():
        pws.append([name, spec["price_usd_per_bbl"], spec["max_demand_kbd"]])
    cws = wb.create_sheet("Capacities")
    cws.append(["unit", "capacity_kbd"])
    for k, v in data.get("capacities", {}).items():
        cws.append([k, v])

    if include_intermediates:
        ipath = default_intermediates_path()
        if ipath.is_file():
            ipkg = load_intermediates_json(ipath)
            iws = wb.create_sheet("Intermediates")
            iws.append(
                [
                    "name",
                    "stream",
                    "api",
                    "sulfur_wt",
                    "ccr_wt",
                    "nitrogen_ppm",
                    "paraffins_vol",
                    "naphthenes_vol",
                    "aromatics_vol",
                    "price",
                ]
            )
            for row in ipkg.get("intermediates") or []:
                iws.append(
                    [
                        row.get("name"),
                        row.get("stream", row.get("name")),
                        row.get("api"),
                        row.get("sulfur_wt"),
                        row.get("ccr_wt"),
                        row.get("nitrogen_ppm"),
                        row.get("paraffins_vol"),
                        row.get("naphthenes_vol"),
                        row.get("aromatics_vol"),
                        row.get("price_usd_per_bbl"),
                    ]
                )

    wb.save(path)
    return path
